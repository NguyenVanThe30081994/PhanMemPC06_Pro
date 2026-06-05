# -*- coding: utf-8 -*-
import html
import json
import os
import re
import unicodedata
import ast
import operator as op
from copy import copy
from datetime import date, datetime

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries

CELL_REF_RE = re.compile(r'(?<![A-Z0-9_"])\$?[A-Z]{1,3}\$?\d+')
RANGE_REF_RE = re.compile(r'\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+')
ROW_RANGE_REF_RE = re.compile(r'(?<![A-Z0-9_"])\$?\d+:\$?\d+(?![A-Z0-9_"])')
_FORMULA_BIN_OPS = {
    ast.Add: op.add,
    ast.Sub: op.sub,
    ast.Mult: op.mul,
    ast.Div: op.truediv,
    ast.Pow: op.pow,
    ast.Mod: op.mod,
}
_FORMULA_UNARY_OPS = {
    ast.UAdd: op.pos,
    ast.USub: op.neg,
}


def normalize_code(value):
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("đ", "d").replace("Đ", "D").lower().strip()
    text = "".join(ch if ch.isalnum() else "_" for ch in text)
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


def safe_filename(value):
    text = unicodedata.normalize("NFKD", str(value or "report"))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace(" ", "_")
    return "".join(ch for ch in text if ch.isalnum() or ch in "_-.")


def _load_workbook(file_path):
    try:
        return openpyxl.load_workbook(file_path, rich_text=True, data_only=False)
    except Exception:
        return openpyxl.load_workbook(file_path, data_only=False)


def _safe_color(color_obj):
    try:
        if not color_obj:
            return None
        if color_obj.type == "rgb" and color_obj.rgb:
            rgb = str(color_obj.rgb).upper()
            if len(rgb) == 8:
                rgb = rgb[2:]
            if rgb in {"FFFFFF", "000000", "00000000"}:
                return None
            return rgb[-6:]
        if color_obj.type == "theme" and color_obj.theme is not None:
            return f"THEME_{color_obj.theme}"
    except Exception:
        return None
    return None


def _safe_formula_eval(expression, allowed_functions):
    def eval_node(node):
        if isinstance(node, ast.Expression):
            return eval_node(node.body)
        if isinstance(node, ast.Constant):
            if isinstance(node.value, (int, float, str)):
                return node.value
            raise ValueError("Unsupported constant in formula")
        if isinstance(node, ast.Num):
            return node.n
        if isinstance(node, ast.BinOp) and type(node.op) in _FORMULA_BIN_OPS:
            return _FORMULA_BIN_OPS[type(node.op)](eval_node(node.left), eval_node(node.right))
        if isinstance(node, ast.UnaryOp) and type(node.op) in _FORMULA_UNARY_OPS:
            return _FORMULA_UNARY_OPS[type(node.op)](eval_node(node.operand))
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            func = allowed_functions.get(node.func.id)
            if not func:
                raise ValueError("Unsupported function in formula")
            return func(*[eval_node(arg) for arg in node.args])
        raise ValueError("Unsupported formula expression")

    parsed = ast.parse(expression, mode="eval")
    return eval_node(parsed)


def is_input_cell(cell):
    try:
        fill = cell.fill
        if not fill or fill.patternType is None:
            return False
        color = _safe_color(fill.start_color)
        if not color:
            return False
        if color.startswith("THEME_"):
            return True
        return color not in {"FFFFFF", "000000"}
    except Exception:
        return False


def format_excel_number(value, number_format):
    if value is None or value == "":
        return ""
    if not isinstance(value, (int, float)):
        return str(value).strip()
    if not number_format:
        return str(int(value)) if float(value).is_integer() else f"{value:.2f}".rstrip("0").rstrip(".")
    fmt = str(number_format).lower().strip()
    try:
        if "%" in fmt:
            decimals = 0
            if "." in fmt:
                decimals = len(fmt.split(".")[1].split("%")[0])
            return f"{value * 100:.{decimals}f}%"
        if "#,##0" in fmt or "," in fmt:
            decimals = 0
            if "." in fmt:
                decimals = len(fmt.split(".")[1])
            result = f"{value:,.{decimals}f}"
            return result.rstrip("0").rstrip(".") if decimals == 0 else result
        if "0" in fmt:
            if "." in fmt:
                decimals = len(fmt.split(".")[1])
                return f"{value:.{decimals}f}"
            return str(int(round(value)))
    except Exception:
        pass
    return str(int(value)) if float(value).is_integer() else str(value)


def _extract_styles(cell):
    styles = []
    if cell.fill and cell.fill.patternType and cell.fill.patternType != "none":
        color = _safe_color(cell.fill.start_color)
        if color and not color.startswith("THEME_") and color not in {"FFFFFF", "000000"}:
            styles.append(f"background-color:#{color};")
    f = cell.font
    if f:
        if f.bold:
            styles.append("font-weight:bold;")
        if f.italic:
            styles.append("font-style:italic;")
        if f.sz:
            styles.append(f"font-size:{f.sz}pt;")
        if f.color and getattr(f.color, "rgb", None):
            rgb = str(f.color.rgb)
            if len(rgb) == 8:
                styles.append(f"color:#{rgb[2:]};")
    a = cell.alignment
    if a:
        if a.horizontal:
            styles.append(f"text-align:{a.horizontal};")
        if a.vertical:
            styles.append(f"vertical-align:{a.vertical};")
        styles.append("white-space:normal;" if a.wrapText else "white-space:nowrap;")
    return "".join(styles)


def _active_bounds(ws):
    dim = ws.calculate_dimension()
    min_col, min_row, max_col, max_row = range_boundaries(dim)
    for merge in ws.merged_cells.ranges:
        min_col = min(min_col, merge.min_col)
        min_row = min(min_row, merge.min_row)
        max_col = max(max_col, merge.max_col)
        max_row = max(max_row, merge.max_row)
    return min_col, min_row, max_col, max_row


def _merge_lookup(ws):
    spans = {}
    shadows = set()
    anchor_by_coord = {}
    for merge in ws.merged_cells.ranges:
        r1, c1, r2, c2 = merge.min_row, merge.min_col, merge.max_row, merge.max_col
        spans[(r1, c1)] = (r2 - r1 + 1, c2 - c1 + 1)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if r == r1 and c == c1:
                    anchor_by_coord[(r, c)] = (r1, c1)
                    continue
                shadows.add((r, c))
                anchor_by_coord[(r, c)] = (r1, c1)
    return spans, shadows, anchor_by_coord


def _anchor_value(ws, r, c, anchor_by_coord):
    ar, ac = anchor_by_coord.get((r, c), (r, c))
    return ws.cell(row=ar, column=ac).value


def _resolve_header_range(min_row, max_row, header_rows=2, header_start_row=None, header_end_row=None):
    start_row = int(header_start_row or min_row or 1)
    if header_end_row:
        end_row = int(header_end_row)
    else:
        end_row = start_row + max(int(header_rows or 1), 1) - 1
    if end_row < start_row:
        end_row = start_row
    if max_row < start_row:
        start_row = max_row
    if max_row < end_row:
        end_row = max_row
    return max(1, start_row), max(1, end_row)


def _resolve_column_range(min_col, max_col, start_column=None, end_column=None):
    start_col = min_col
    end_col = max_col
    if start_column:
        try:
            start_col = column_index_from_string(str(start_column).strip().upper())
        except Exception:
            start_col = min_col
    if end_column:
        try:
            end_col = column_index_from_string(str(end_column).strip().upper())
        except Exception:
            end_col = max_col
    start_col = max(min_col, start_col)
    end_col = min(max_col, end_col)
    if end_col < start_col:
        end_col = start_col
    return start_col, end_col


def _coerce_numeric(value):
    if value in (None, ""):
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if not text:
        return 0
    try:
        return float(text)
    except Exception:
        return value


def _split_formula_args(expr):
    args = []
    current = []
    depth = 0
    in_string = False
    i = 0
    while i < len(expr):
        ch = expr[i]
        if ch == '"':
            in_string = not in_string
            current.append(ch)
        elif not in_string and ch == "(":
            depth += 1
            current.append(ch)
        elif not in_string and ch == ")":
            depth = max(depth - 1, 0)
            current.append(ch)
        elif not in_string and depth == 0 and ch == ",":
            args.append("".join(current).strip())
            current = []
        else:
            current.append(ch)
        i += 1
    if current:
        args.append("".join(current).strip())
    return args


def _evaluate_sheet_formulas(ws):
    cache = {}
    visiting = set()

    def _column_bound_row_range(range_ref, current_coord):
        if not current_coord or ":" not in str(range_ref or ""):
            return None
        current_column = "".join(ch for ch in str(current_coord or "").upper() if ch.isalpha())
        if not current_column:
            return None
        start_row_text, end_row_text = [part.replace("$", "").strip() for part in str(range_ref).split(":", 1)]
        if not start_row_text.isdigit() or not end_row_text.isdigit():
            return None
        return f"{current_column}{int(start_row_text)}:{current_column}{int(end_row_text)}"

    def cell_value(coord):
        coord = coord.replace("$", "").upper()
        if coord in cache:
            return cache[coord]
        if coord in visiting:
            return ""
        visiting.add(coord)
        cell = ws[coord]
        value = cell.value
        if cell.data_type == "f" and isinstance(value, str):
            result = eval_formula(value, current_coord=coord)
        else:
            result = value
            coerced = _coerce_numeric(result)
            if isinstance(coerced, (int, float)):
                result = coerced
        visiting.discard(coord)
        cache[coord] = result
        return result

    def range_values(range_ref):
        start, end = range_ref.split(":")
        start = start.replace("$", "").upper()
        end = end.replace("$", "").upper()
        start_col, start_row, end_col, end_row = range_boundaries(f"{start}:{end}")
        values = []
        for row_index in range(start_row, end_row + 1):
            for col_index in range(start_col, end_col + 1):
                values.append(cell_value(f"{get_column_letter(col_index)}{row_index}"))
        return values

    def func_sum(*items):
        total = 0
        for item in items:
            if isinstance(item, list):
                total += func_sum(*item)
                continue
            number = _coerce_numeric(item)
            if isinstance(number, (int, float)):
                total += number
        return total

    def func_iferror(primary, fallback):
        try:
            if isinstance(primary, Exception):
                raise primary
            return primary
        except Exception:
            return fallback

    def eval_formula(raw_formula, current_coord=None):
        formula = str(raw_formula or "").strip()
        if formula.startswith("="):
            formula = formula[1:]
        if not formula:
            return ""

        transformed = formula.replace("^", "**")
        transformed = re.sub(r'\bSUM\s*\(', 'FUNC_SUM(', transformed, flags=re.IGNORECASE)
        transformed = re.sub(r'\bIFERROR\s*\(', 'FUNC_IFERROR(', transformed, flags=re.IGNORECASE)
        range_placeholders = {}
        def replace_range(match):
            key = f"__RANGE_{len(range_placeholders)}__"
            range_placeholders[key] = f'RANGE("{match.group(0).replace("$", "").upper()}")'
            return key
        transformed = RANGE_REF_RE.sub(replace_range, transformed)
        transformed = ROW_RANGE_REF_RE.sub(
            lambda match: replace_range(match)
            if _column_bound_row_range(match.group(0), current_coord)
            else match.group(0),
            transformed,
        )
        for key, replacement in list(range_placeholders.items()):
            raw_ref = replacement[len('RANGE("'):-2]
            normalized_row_range = _column_bound_row_range(raw_ref, current_coord)
            if normalized_row_range:
                range_placeholders[key] = f'RANGE("{normalized_row_range}")'
        transformed = CELL_REF_RE.sub(lambda m: f'CELL("{m.group(0).replace("$", "").upper()}")', transformed)
        for key, replacement in range_placeholders.items():
            transformed = transformed.replace(key, replacement)
        allowed_functions = {
            "CELL": cell_value,
            "RANGE": range_values,
            "FUNC_SUM": func_sum,
            "FUNC_IFERROR": func_iferror,
        }
        try:
            result = _safe_formula_eval(transformed, allowed_functions)
        except ZeroDivisionError:
            return ""
        except Exception:
            if transformed.startswith('FUNC_IFERROR(') and transformed.endswith(')'):
                inner = transformed[len('FUNC_IFERROR('):-1]
                args = _split_formula_args(inner)
                if len(args) == 2:
                    try:
                        return _safe_formula_eval(args[1], allowed_functions)
                    except Exception:
                        return ""
            return formula
        return result

    overrides = {}
    min_col, min_row, max_col, max_row = _active_bounds(ws)
    for row_index in range(min_row, max_row + 1):
        for col_index in range(min_col, max_col + 1):
            cell = ws.cell(row=row_index, column=col_index)
            if cell.data_type == "f":
                overrides[cell.coordinate] = cell_value(cell.coordinate)
    return overrides


def _suggest_hidden_field(field):
    label = normalize_code(field.get("field_name") or field.get("path_code") or "")
    if not label:
        return False
    hidden_markers = {
        "stt",
        "so_thu_tu",
        "thu_tu",
        "don_vi",
        "ten_don_vi",
        "ma_don_vi",
    }
    return any(marker in label for marker in hidden_markers)


def parse_workbook(
    file_path,
    header_rows=2,
    data_start_row=3,
    header_start_row=None,
    header_end_row=None,
    data_end_row=None,
    total_start_row=None,
    total_end_row=None,
    start_column=None,
    end_column=None,
    sheet_options=None,
):
    wb = _load_workbook(file_path)
    metadata = {"sheets": [], "parser_version": "1.0"}
    sheet_options = sheet_options or {}

    visible_order = 0
    for ws in wb.worksheets:
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue
        order = visible_order
        visible_order += 1
        sheet_option = sheet_options.get(ws.title, {})
        min_col, min_row, max_col, max_row = _active_bounds(ws)
        resolved_min_col, resolved_max_col = _resolve_column_range(
            min_col,
            max_col,
            start_column=sheet_option.get("start_column") or start_column,
            end_column=sheet_option.get("end_column") or end_column,
        )
        spans, shadows, anchors = _merge_lookup(ws)
        resolved_header_start, resolved_header_end = _resolve_header_range(
            min_row,
            max_row,
            header_rows=sheet_option.get("header_rows") or header_rows,
            header_start_row=sheet_option.get("header_start_row") or header_start_row,
            header_end_row=sheet_option.get("header_end_row") or header_end_row,
        )
        resolved_data_start = max(
            1,
            int(sheet_option.get("data_start_row") or data_start_row or (resolved_header_end + 1)),
        )
        resolved_data_end = min(max_row, int(sheet_option.get("data_end_row") or data_end_row or max_row))
        if resolved_data_end < resolved_data_start:
            resolved_data_end = resolved_data_start
        resolved_total_start = int(sheet_option.get("total_start_row") or total_start_row or 0)
        resolved_total_end = int(sheet_option.get("total_end_row") or total_end_row or 0)
        if resolved_total_start:
            resolved_total_start = min(max_row, max(1, resolved_total_start))
        if resolved_total_end:
            resolved_total_end = min(max_row, max(resolved_total_start or 1, resolved_total_end))
        display_end_row = max(resolved_header_end, resolved_data_end, resolved_total_end or 0)
        header_rows_meta = []
        field_lookup = {}
        fields = []
        used_codes = {}

        for r in range(resolved_header_start, resolved_header_end + 1):
            row_meta = {"row": r, "cells": []}
            for c in range(resolved_min_col, resolved_max_col + 1):
                if (r, c) in shadows:
                    continue
                cell = ws.cell(row=r, column=c)
                value = _anchor_value(ws, r, c, anchors)
                if value is None or str(value).strip() == "":
                    continue
                rowspan, colspan = spans.get((r, c), (1, 1))
                row_meta["cells"].append({
                    "label": str(value),
                    "row": r,
                    "col": c,
                    "rowspan": rowspan,
                    "colspan": colspan,
                    "coord": cell.coordinate,
                })
            header_rows_meta.append(row_meta)

        for c in range(resolved_min_col, resolved_max_col + 1):
            path_labels = []
            for r in range(resolved_header_start, resolved_header_end + 1):
                value = _anchor_value(ws, r, c, anchors)
                if value is not None and str(value).strip():
                    path_labels.append(str(value).strip())
            leaf_label = path_labels[-1] if path_labels else get_column_letter(c)
            base_code = normalize_code(leaf_label) or get_column_letter(c).lower()
            count = used_codes.get(base_code, 0)
            field_code = base_code if count == 0 else f"{base_code}_{get_column_letter(c).lower()}"
            used_codes[base_code] = count + 1
            inferred_type = "text"
            for r in range(resolved_data_start, max_row + 1):
                if ws.row_dimensions[r].hidden:
                    continue
                sample = ws.cell(row=r, column=c).value
                if sample in (None, ""):
                    continue
                if isinstance(sample, (int, float)):
                    inferred_type = "number"
                    break
                if isinstance(sample, (datetime, date)):
                    inferred_type = "date"
                    break
            field = {
                "field_code": field_code,
                "field_name": leaf_label,
                "column_index": c,
                "column_letter": get_column_letter(c),
                "data_type": inferred_type,
                "input_mode": "text",
                "is_required": False,
                "is_visible": True,
                "is_editable": True,
                "default_value": "",
                "validation_rule": "",
                "dictionary_source": "",
                "formula_expression": "",
                "aggregation_type": "",
                "display_order": c - min_col + 1,
                "path_code": " > ".join(path_labels) if path_labels else get_column_letter(c),
            }
            fields.append(field)
            field_lookup[c] = field_code

        input_cells = []
        for r in range(resolved_data_start, resolved_data_end + 1):
            if ws.row_dimensions[r].hidden:
                continue
            for c in range(resolved_min_col, resolved_max_col + 1):
                cell = ws.cell(row=r, column=c)
                if is_input_cell(cell):
                    input_cells.append({
                        "sheet_name": ws.title,
                        "cell_address": cell.coordinate,
                        "row_index": r,
                        "column_index": c,
                        "field_code": field_lookup.get(c),
                    })

        if not input_cells:
            for r in range(resolved_data_start, resolved_data_end + 1):
                if ws.row_dimensions[r].hidden:
                    continue
                for c in range(resolved_min_col, resolved_max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if (r, c) in shadows:
                        continue
                    if cell.value in (None, ""):
                        input_cells.append({
                            "sheet_name": ws.title,
                            "cell_address": cell.coordinate,
                            "row_index": r,
                            "column_index": c,
                            "field_code": field_lookup.get(c),
                        })

        editable_columns = {cell["column_index"] for cell in input_cells if cell.get("column_index")}
        for field in fields:
            is_input_column = field["column_index"] in editable_columns
            field["is_visible"] = not _suggest_hidden_field(field)
            field["is_editable"] = is_input_column

        metadata["sheets"].append({
            "sheet_name": ws.title,
            "order_index": order,
            "min_row": min_row,
            "min_col": resolved_min_col,
            "max_row": display_end_row,
            "max_col": resolved_max_col,
            "header_start_row": resolved_header_start,
            "header_end_row": resolved_header_end,
            "header_rows": max(0, resolved_header_end - resolved_header_start + 1),
            "data_start_row": resolved_data_start,
            "data_end_row": resolved_data_end,
            "unit_start_row": resolved_data_start,
            "unit_end_row": resolved_data_end,
            "total_start_row": resolved_total_start,
            "total_end_row": resolved_total_end,
            "start_column": get_column_letter(resolved_min_col),
            "end_column": get_column_letter(resolved_max_col),
            "merges": [str(m) for m in ws.merged_cells.ranges],
            "hidden_rows": [idx for idx in range(min_row, display_end_row + 1) if ws.row_dimensions[idx].hidden],
            "hidden_cols": [get_column_letter(i) for i in range(resolved_min_col, resolved_max_col + 1) if ws.column_dimensions[get_column_letter(i)].hidden],
            "header_rows_meta": header_rows_meta,
            "fields": fields,
            "field_lookup": field_lookup,
            "input_cells": input_cells,
        })

    return metadata


def render_sheet_html(
    ws,
    editable_values=None,
    field_lookup=None,
    editable=True,
    start_row=None,
    end_row=None,
    min_col=None,
    max_col=None,
    header_end_row=None,
    sticky_first_col=None,
):
    editable_values = editable_values or {}
    field_lookup = field_lookup or {}
    active_min_col, active_min_row, active_max_col, active_max_row = _active_bounds(ws)
    min_col = max(active_min_col, min_col or active_min_col)
    max_col = min(active_max_col, max_col or active_max_col)
    min_row = max(active_min_row, start_row or active_min_row)
    max_row = min(active_max_row, end_row or active_max_row)
    spans, shadows, anchors = _merge_lookup(ws)
    html_rows = []
    sticky_top_by_row = {}
    running_header_top = 0

    for r in range(min_row, max_row + 1):
        if ws.row_dimensions[r].hidden:
            continue
        row_height = ws.row_dimensions[r].height or 15
        rendered_row_height = int(row_height * 1.33)
        if header_end_row is not None and r <= int(header_end_row):
            sticky_top_by_row[r] = running_header_top
            running_header_top += rendered_row_height
        html_rows.append(f'<tr style="height:{rendered_row_height}px">')
        for c in range(min_col, max_col + 1):
            if (r, c) in shadows:
                continue
            cell = ws.cell(row=r, column=c)
            rowspan, colspan = spans.get((r, c), (1, 1))
            coord = cell.coordinate
            value = editable_values.get(coord, cell.value)
            if value is None:
                value = ""
            if cell.data_type == "f" and coord not in editable_values:
                value = cell.value or ""
            if isinstance(value, (datetime, date)):
                display = value.strftime("%d/%m/%Y")
            elif isinstance(value, (int, float)):
                display = format_excel_number(value, cell.number_format)
            else:
                display = str(value)
            css = _extract_styles(cell)
            field_meta = field_lookup.get(c, {})
            if isinstance(field_meta, dict):
                field_code = field_meta.get("code", "")
                field_label = field_meta.get("label", "")
            else:
                field_code = field_meta or ""
                field_label = ""
            is_input = editable and is_input_cell(cell)
            attrs = [f'data-cell="{coord}"']
            if field_code:
                attrs.append(f'data-field-code="{field_code}"')
            if field_label:
                attrs.append(f'data-field-label="{html.escape(field_label, quote=True)}"')
                attrs.append(f'title="{html.escape(field_label, quote=True)}"')
            if is_input:
                attrs.append('contenteditable="true"')
                attrs.append('spellcheck="false"')
            td_class = "report-cell"
            is_header_cell = header_end_row is not None and r <= int(header_end_row)
            is_first_col_cell = sticky_first_col is not None and c == int(sticky_first_col)
            if is_header_cell:
                td_class += " report-cell--header"
            if is_first_col_cell:
                td_class += " report-cell--first-col"
            if is_input:
                td_class += " report-input"
            sticky_style = ""
            if is_header_cell:
                sticky_style = f"--report-sticky-top:{sticky_top_by_row.get(r, 0)}px;"
            rs_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ""
            cs_attr = f' colspan="{colspan}"' if colspan > 1 else ""
            html_rows.append(
                f'<td class="{td_class}" {rs_attr}{cs_attr} style="padding:4px 6px;border:1px solid #d1d5db;overflow:hidden;box-sizing:border-box;{sticky_style}{css}" {" ".join(attrs)}>{html.escape(display)}</td>'
            )
        html_rows.append("</tr>")

    return "\n".join(html_rows)


def build_preview_workbook(template_path, values_by_sheet=None):
    wb = _load_workbook(template_path)
    values_by_sheet = values_by_sheet or {}
    formula_values = {}
    for ws in wb.worksheets:
        sheet_values = values_by_sheet.get(ws.title, {})
        for cell_address, value in sheet_values.items():
            ws[cell_address] = value
        formula_values[ws.title] = _evaluate_sheet_formulas(ws)
    return wb, formula_values


def write_workbook_copy(template_path, output_path, values_by_sheet):
    wb = _load_workbook(template_path)
    for ws in wb.worksheets:
        sheet_values = values_by_sheet.get(ws.title, {})
        for cell_address, value in sheet_values.items():
            ws[cell_address] = value
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    wb.save(output_path)
    return output_path
