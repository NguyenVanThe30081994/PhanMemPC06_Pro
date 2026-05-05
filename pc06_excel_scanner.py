# -*- coding: utf-8 -*-
# Excel Scanner for V2 Reports
# Uses raw values from Excel - no conversion to avoid float precision errors
import io
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries


def _is_blank(value):
    return value is None or str(value).strip() == ''


def _is_formula_text(value):
    return isinstance(value, str) and value.strip().startswith('=')


def _is_text_like(value):
    if _is_blank(value) or _is_formula_text(value):
        return False
    return isinstance(value, str)


def _is_numeric_like(value):
    if _is_blank(value):
        return False
    return isinstance(value, (int, float))


def _normalized_text(value):
    if _is_blank(value):
        return ''
    return ' '.join(str(value).replace('\n', ' ').split()).strip().lower()


def _row_profile(ws, row, min_col, max_col):
    entries = []
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row, col)
        entries.append((col, cell.value, cell.data_type))
    values = [value for _, value, _ in entries]
    non_empty = [value for _, value, _ in entries if not _is_blank(value)]
    text_like = [value for _, value, _ in entries if not _is_blank(value) and _is_text_like(value)]
    numeric_like = [value for _, value, _ in entries if not _is_blank(value) and _is_numeric_like(value)]
    formula_like = [
        value for _, value, data_type in entries
        if not _is_blank(value) and (data_type == 'f' or _is_formula_text(value))
    ]
    normalized = [_normalized_text(value) for _, value, _ in entries if _normalized_text(value)]
    return {
        'values': values,
        'non_empty': non_empty,
        'text_like_count': len(text_like),
        'numeric_like_count': len(numeric_like),
        'formula_like_count': len(formula_like),
        'normalized_texts': normalized,
        'non_empty_count': len(non_empty),
    }


def _looks_like_detail_row(ws, row, min_col, max_col):
    first_value = ws.cell(row, min_col).value
    second_value = ws.cell(row, min_col + 1).value if min_col + 1 <= max_col else None
    second_text = _normalized_text(second_value)
    if not _is_numeric_like(first_value):
        return False
    if not _is_text_like(second_value):
        return False
    if not second_text or second_text in {'tên đơn vị', 'đơn vị'}:
        return False
    if any(token in second_text for token in ('toàn tỉnh', 'tổng', 'cộng')):
        return False

    numeric_or_formula = 0
    for col in range(min_col + 2, max_col + 1):
        cell = ws.cell(row, col)
        value = cell.value
        if _is_numeric_like(value) or cell.data_type == 'f' or _is_formula_text(value):
            numeric_or_formula += 1
    return numeric_or_formula >= 1


def _looks_like_numbering_row(ws, row, min_col, max_col):
    profile = _row_profile(ws, row, min_col, max_col)
    if profile['non_empty_count'] < 3:
        return False
    if profile['formula_like_count'] > 0 or profile['text_like_count'] > 0:
        return False
    seq = []
    for col in range(min_col, max_col + 1):
        value = ws.cell(row, col).value
        if _is_numeric_like(value):
            try:
                seq.append(int(value))
            except Exception:
                return False
    if len(seq) < 3:
        return False
    return seq == list(range(seq[0], seq[0] + len(seq)))


def _looks_like_summary_row(ws, row, min_col, max_col):
    profile = _row_profile(ws, row, min_col, max_col)
    if profile['non_empty_count'] == 0:
        return False
    first_text = _normalized_text(ws.cell(row, min_col).value)
    second_text = _normalized_text(ws.cell(row, min_col + 1).value if min_col + 1 <= max_col else None)
    leading_text = ' '.join([part for part in (first_text, second_text) if part])
    if profile['formula_like_count'] >= 1 and not _looks_like_detail_row(ws, row, min_col, max_col):
        return True
    texts = ' '.join(profile['normalized_texts'])
    if any(token in leading_text for token in ('toàn tỉnh', 'tổng', 'cộng')):
        return True
    return any(token in texts for token in ('toàn tỉnh', 'tổng', 'cộng'))


def _max_merge_span_for_row(merged_ranges, row):
    max_span = 1
    for merged in merged_ranges:
        if merged['row'] <= row < merged['row'] + merged.get('rowspan', 1):
            max_span = max(max_span, merged.get('colspan', 1))
    return max_span


def _looks_like_title_row(ws, row, min_col, max_col, merged_ranges, used_width):
    profile = _row_profile(ws, row, min_col, max_col)
    if profile['non_empty_count'] == 0:
        return False
    if profile['numeric_like_count'] > 0 and profile['text_like_count'] == 0:
        return False
    max_merge = _max_merge_span_for_row(merged_ranges, row)
    if max_merge >= max(4, int(used_width * 0.6)) and profile['text_like_count'] <= 2:
        return True
    if max_merge >= max(4, int(used_width * 0.25)) and profile['text_like_count'] == 1 and profile['non_empty_count'] == 1:
        return True
    if profile['text_like_count'] == 1 and profile['non_empty_count'] == 1 and len(profile['normalized_texts'][0]) >= 20:
        return True
    return False


def _looks_like_header_row(ws, row, min_col, max_col, merged_ranges, used_width):
    profile = _row_profile(ws, row, min_col, max_col)
    if profile['non_empty_count'] == 0:
        return False
    if _looks_like_numbering_row(ws, row, min_col, max_col):
        return False
    if _looks_like_summary_row(ws, row, min_col, max_col):
        return False
    max_merge = _max_merge_span_for_row(merged_ranges, row)
    if profile['text_like_count'] >= 2:
        return True
    if max_merge >= 2 and profile['text_like_count'] >= 1 and max_merge < max(4, int(used_width * 0.6)):
        return True
    return False


def _load_workbook_utf8(buffer, **kwargs):
    """Load workbook with UTF-8 encoding support"""
    kwargs.setdefault('read_only', False)
    kwargs.setdefault('keep_vba', True)
    # Try with rich_text for better UTF-8 handling
    try:
        return openpyxl.load_workbook(buffer, rich_text=True, **kwargs)
    except:
        return openpyxl.load_workbook(buffer, **kwargs)


def scan_excel_structure(excel_blob):
    """
    Quét cấu trúc file Excel và trả về gợi ý cấu hình.
    Chỉ quét vùng có dữ liệu thực tế (used range), không quét vùng trắng.
    """
    wb = _load_workbook_utf8(io.BytesIO(excel_blob))
    ws = wb.active
    
    # Tìm vùng thực tế có dữ liệu (used range)
    # Thay vì dùng max_row/max_column (có thể bao gồm vùng trắng), 
    # ta tìm last row/col có dữ liệu
    used_min_row = ws.min_row
    used_max_row = ws.max_row
    used_min_col = ws.min_column
    used_max_col = ws.max_column
    
    # Nếu worksheet có nhiều dòng trống ở cuối, tìn lại
    # Duyệt từ cuối ngược lên để tìm dòng có dữ liệu
    if used_max_row > used_min_row:
        for row in range(used_max_row, used_min_row, -1):
            has_data = False
            for col in range(used_min_col, used_max_col + 1):
                cell_val = ws.cell(row, col).value
                if cell_val is not None and str(cell_val).strip() != '':
                    has_data = True
                    break
            if has_data:
                used_max_row = row
                break
            else:
                used_max_row = row - 1  # Giảm max_row nếu dòng trống
    
    # Tương tự cho cột
    if used_max_col > used_min_col:
        for col in range(used_max_col, used_min_col, -1):
            has_data = False
            for row in range(used_min_row, used_max_row + 1):
                cell_val = ws.cell(row, col).value
                if cell_val is not None and str(cell_val).strip() != '':
                    has_data = True
                    break
            if has_data:
                used_max_col = col
                break
            else:
                used_max_col = col - 1
    
    used_width = max(1, used_max_col - used_min_col + 1)
    hidden_rows = []
    hidden_columns = []
    visible_columns = []
    for row in range(used_min_row, used_max_row + 1):
        if ws.row_dimensions[row].hidden:
            hidden_rows.append(row)
    for col in range(used_min_col, used_max_col + 1):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].hidden:
            hidden_columns.append(letter)
        else:
            visible_columns.append(letter)

    result = {
        'sheet_name': ws.title,
        'total_rows': used_max_row,
        'total_cols': used_max_col,
        'used_range': f"{get_column_letter(used_min_col)}{used_min_row}:{get_column_letter(used_max_col)}{used_max_row}",
        'columns': [],  # ['A', 'B', 'C', ...]
        'visible_columns': [],
        'hidden_rows': hidden_rows,
        'hidden_columns': hidden_columns,
        'title_rows': [],
        'header_rows': [],
        'helper_rows': [],
        'summary_rows': [],
        'data_start_row': 4,  # Default
        'headers': {},  # {row: {col: value}}
        'merged_cells': [],
        'numeric_columns': [],
        'formulas': {},  # {col_letter: formula_type}
        'original_max_row': ws.max_row,  # Lưu lại để debug
        'original_max_col': ws.max_column,
    }
    
    # Get columns (chỉ vùng có dữ liệu)
    for col in range(used_min_col, used_max_col + 1):
        result['columns'].append(get_column_letter(col))
    result['visible_columns'] = visible_columns
    
    # Scan merged cells
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        result['merged_cells'].append({
            'range': str(merged_range),
            'row': min_row,
            'col_start': get_column_letter(min_col),
            'col_end': get_column_letter(max_col),
            'colspan': max_col - min_col + 1,
            'rowspan': max_row - min_row + 1,
            'value': ws.cell(min_row, min_col).value
        })
    
    # Detect table structure by first finding detail data rows, then walk upward
    detected_detail_row = None
    for row in range(used_min_row, min(used_max_row, used_min_row + 79) + 1):
        if ws.row_dimensions[row].hidden:
            continue
        if _looks_like_detail_row(ws, row, used_min_col, used_max_col):
            detected_detail_row = row
            break

    header_rows = []
    helper_rows = []
    summary_rows = []
    title_rows = []

    if detected_detail_row:
        result['data_start_row'] = detected_detail_row
        blank_streak = 0
        for row in range(detected_detail_row - 1, used_min_row - 1, -1):
            if ws.row_dimensions[row].hidden:
                continue
            if all(_is_blank(ws.cell(row, col).value) for col in range(used_min_col, used_max_col + 1)):
                blank_streak += 1
                if header_rows and blank_streak >= 1:
                    break
                continue

            if _looks_like_summary_row(ws, row, used_min_col, used_max_col):
                summary_rows.append(row)
                continue
            if _looks_like_numbering_row(ws, row, used_min_col, used_max_col):
                helper_rows.append(row)
                continue
            if _looks_like_header_row(ws, row, used_min_col, used_max_col, result['merged_cells'], used_width):
                header_rows.append(row)
                blank_streak = 0
                continue
            if header_rows:
                break

        header_rows = sorted(header_rows)
        helper_rows = sorted(helper_rows)
        summary_rows = sorted(summary_rows)

        for row in range(used_min_row, (min(header_rows) - 1) if header_rows else detected_detail_row):
            if ws.row_dimensions[row].hidden:
                continue
            if all(_is_blank(ws.cell(row, col).value) for col in range(used_min_col, used_max_col + 1)):
                continue
            if _looks_like_title_row(ws, row, used_min_col, used_max_col, result['merged_cells'], used_width):
                title_rows.append(row)
    else:
        # Fallback for files where the detail row is not recognizable.
        header_candidates = []
        detected_data_row = None
        scan_limit = min(used_max_row, used_min_row + 39)
        for row in range(used_min_row, scan_limit + 1):
            if ws.row_dimensions[row].hidden:
                continue
            profile = _row_profile(ws, row, used_min_col, used_max_col)
            if not profile['non_empty_count']:
                continue

            text_like_count = profile['text_like_count']
            numeric_like_count = profile['numeric_like_count']
            formula_like_count = profile['formula_like_count']
            max_merge = _max_merge_span_for_row(result['merged_cells'], row)
            first_cell = ws.cell(row, used_min_col).value
            numeric_ratio = numeric_like_count / max(1, profile['non_empty_count'])

            looks_like_data = (
                numeric_like_count >= 2
                and numeric_ratio >= 0.35
                and text_like_count <= max(2, profile['non_empty_count'] // 3)
            ) or (
                header_candidates
                and _is_numeric_like(first_cell)
                and numeric_like_count >= 2
                and text_like_count <= 2
                and formula_like_count <= 2
            ) or (
                header_candidates
                and (numeric_like_count > 0 or formula_like_count > 0)
                and text_like_count <= max(1, profile['non_empty_count'] // 2)
            )

            if header_candidates and looks_like_data:
                detected_data_row = row
                break

            if text_like_count > 0 or max_merge >= max(3, int(used_width * 0.45)):
                header_candidates.append(row)
                continue

            if header_candidates and numeric_like_count > 0:
                detected_data_row = row
                break

        header_rows = header_candidates
        result['data_start_row'] = detected_data_row or result['data_start_row']

    result['title_rows'] = title_rows
    result['header_rows'] = header_rows
    result['helper_rows'] = helper_rows
    result['summary_rows'] = summary_rows

    # Scan headers - use raw value directly, don't convert
    for row in result['header_rows']:
        if ws.row_dimensions[row].hidden:
            continue
        result['headers'][row] = {}
        for col in range(used_min_col, used_max_col + 1):
            cell_val = ws.cell(row, col).value
            if cell_val:
                # Use raw value - don't convert to avoid float precision issues
                result['headers'][row][get_column_letter(col)] = cell_val

    # Detect numeric columns
    data_row = result['data_start_row']
    if data_row <= ws.max_row:
        numeric_cols = set()
        sample_rows = [
            row for row in range(data_row, min(data_row + 8, used_max_row + 1))
            if not ws.row_dimensions[row].hidden and any(not _is_blank(ws.cell(row, col).value) for col in range(used_min_col, used_max_col + 1))
        ]
        
        for col in range(used_min_col, used_max_col + 1):
            all_numeric = True
            for row in sample_rows:
                cell_val = ws.cell(row, col).value
                if cell_val is not None and cell_val != '':
                    if not isinstance(cell_val, (int, float)):
                        if isinstance(cell_val, str) and cell_val.startswith('='):
                            continue
                        all_numeric = False
                        break
            if all_numeric:
                numeric_cols.add(get_column_letter(col))
        
        result['numeric_columns'] = list(numeric_cols)
    
    # Detect formulas
    formula_cols = {}
    if data_row <= ws.max_row:
        for col in range(used_min_col, used_max_col + 1):
            seen_formula = None
            for row in range(data_row, min(data_row + 8, used_max_row + 1)):
                if ws.row_dimensions[row].hidden:
                    continue
                cell = ws.cell(row, col)
                if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                    formula = str(cell.value or '')
                    if 'SUM' in formula.upper():
                        seen_formula = 'SUM'
                    elif 'AVG' in formula.upper():
                        seen_formula = 'AVG'
                    elif '/' in formula and '%' not in formula.upper():
                        seen_formula = 'RATIO'
                    else:
                        seen_formula = 'CUSTOM'
                    break
            if seen_formula:
                formula_cols[get_column_letter(col)] = seen_formula
    
    result['formulas'] = formula_cols
    
    return result
