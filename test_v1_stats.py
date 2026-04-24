#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test V1 Stats - Kiểm tra render data rows với submissions
"""

import io
import openpyxl
from excel_renderer import build_stats_table_html

# Mock ReportConfig
class MockReportConfig:
    def __init__(self):
        self.header_start = 1
        self.header_rows = 2
        self.config_json = '[{"idx": 2, "is_visible": true}, {"idx": 3, "is_visible": true}]'
        self.is_daily = False
        self.file_blob = None

# Tạo Excel file test
def create_test_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    
    # Header rows
    ws['A1'] = "Đơn vị"
    ws['B1'] = "Số lượng"
    ws['C1'] = "Tỷ lệ"
    
    ws['A2'] = "Unit"
    ws['B2'] = "Qty"
    ws['C2'] = "Rate"
    
    # Data rows (template)
    ws['A3'] = ""
    ws['B3'] = 0
    ws['C3'] = 0
    
    ws['A4'] = ""
    ws['B4'] = 0
    ws['C4'] = 0
    
    # Format
    ws['B3'].number_format = '0'
    ws['C3'].number_format = '0.00'
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Test
print("=" * 80)
print("TEST: V1 Stats - Render data rows with submissions")
print("=" * 80)

config = MockReportConfig()
config.file_blob = create_test_excel()

# Test Case 1: Submissions WITH values
print("\n✓ Test Case 1: Submissions WITH values key")
submissions_with_values = [
    {
        'unit': 'Đơn vị A',
        'sender': 'Người A',
        'date': '23/04/2026',
        'values': {'2': '491', '3': '0.5'}  # ← Có values
    },
    {
        'unit': 'Đơn vị B',
        'sender': 'Người B',
        'date': '23/04/2026',
        'values': {'2': '543', '3': '0.75'}  # ← Có values
    }
]

try:
    html = build_stats_table_html(config.file_blob, config, submissions_with_values)
    if '491' in html and '543' in html:
        print("  ✅ PASS: Data rows rendered with values")
        print(f"  Found: 491, 543 in HTML")
    else:
        print("  ❌ FAIL: Data not found in HTML")
        print(f"  HTML length: {len(html)}")
except Exception as e:
    print(f"  ❌ FAIL: {e}")

# Test Case 2: Submissions WITHOUT values (should not crash)
print("\n✓ Test Case 2: Submissions WITHOUT values key (defensive check)")
submissions_without_values = [
    {
        'unit': 'Đơn vị C',
        'sender': 'Người C',
        'date': '23/04/2026'
        # ← Không có values key
    }
]

try:
    html = build_stats_table_html(config.file_blob, config, submissions_without_values)
    print("  ✅ PASS: No crash when values key missing")
    print(f"  HTML rendered successfully (length: {len(html)})")
except KeyError as e:
    print(f"  ❌ FAIL: KeyError - {e}")
except Exception as e:
    print(f"  ❌ FAIL: {e}")

# Test Case 3: Mixed submissions
print("\n✓ Test Case 3: Mixed submissions (some with values, some without)")
submissions_mixed = [
    {
        'unit': 'Đơn vị D',
        'sender': 'Người D',
        'date': '23/04/2026',
        'values': {'2': '100', '3': '0.1'}
    },
    {
        'unit': 'Đơn vị E',
        'sender': 'Người E',
        'date': '23/04/2026'
        # ← Không có values
    }
]

try:
    html = build_stats_table_html(config.file_blob, config, submissions_mixed)
    if '100' in html:
        print("  ✅ PASS: Mixed submissions handled correctly")
        print(f"  Found: 100 in HTML")
    else:
        print("  ⚠️  WARNING: Data with values not found")
except Exception as e:
    print(f"  ❌ FAIL: {e}")

print("\n" + "=" * 80)
print("✅ All tests completed")
print("=" * 80)
