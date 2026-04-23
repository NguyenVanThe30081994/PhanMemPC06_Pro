#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Integration test: Kiểm tra format_excel_number với Excel file thực tế
"""

import io
import openpyxl
from excel_renderer import format_excel_number

# Tạo Excel file test
def create_test_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    
    # Row 1: Headers
    ws['A1'] = "Integer"
    ws['B1'] = "Decimal"
    ws['C1'] = "Percentage"
    ws['D1'] = "Thousand"
    
    # Row 2: Data with formats
    ws['A2'] = 491
    ws['A2'].number_format = '0'
    
    ws['B2'] = 491.14
    ws['B2'].number_format = '0.00'
    
    ws['C2'] = 0.5
    ws['C2'].number_format = '0%'
    
    ws['D2'] = 1234.56
    ws['D2'].number_format = '#,##0.00'
    
    # Row 3: More test cases
    ws['A3'] = 543
    ws['A3'].number_format = '0'
    
    ws['B3'] = 543.11
    ws['B3'].number_format = '0.00'
    
    ws['C3'] = 0.125
    ws['C3'].number_format = '0.0%'
    
    ws['D3'] = 3441.6
    ws['D3'].number_format = '0.0'
    
    return wb

# Test
print("=" * 80)
print("INTEGRATION TEST: format_excel_number with real Excel file")
print("=" * 80)

wb = create_test_excel()
ws = wb.active

test_data = [
    (2, 'A', '0', '491'),
    (2, 'B', '0.00', '491.14'),
    (2, 'C', '0%', '50%'),
    (2, 'D', '#,##0.00', '1,234.56'),
    (3, 'A', '0', '543'),
    (3, 'B', '0.00', '543.11'),
    (3, 'C', '0.0%', '12.5%'),
    (3, 'D', '0.0', '3441.6'),
]

passed = 0
failed = 0

for row, col, fmt, expected in test_data:
    cell = ws[f'{col}{row}']
    value = cell.value
    number_format = cell.number_format
    result = format_excel_number(value, number_format)
    
    status = "✓ PASS" if result == expected else "✗ FAIL"
    if result == expected:
        passed += 1
    else:
        failed += 1
    
    print(f"\n{status}: Cell {col}{row}")
    print(f"  Value:         {value}")
    print(f"  Format:        {number_format}")
    print(f"  Expected:      {expected}")
    print(f"  Got:           {result}")

print("\n" + "=" * 80)
print(f"Results: {passed} passed, {failed} failed out of {len(test_data)} tests")
print("=" * 80)

if failed == 0:
    print("✓ Integration test passed!")
else:
    print(f"✗ {failed} test(s) failed")
