# -*- coding: utf-8 -*-
"""
Report Exporter
Xuất dữ liệu báo cáo ra Excel từ template đã lưu trong DB.
"""
from io import BytesIO
from datetime import datetime
import json
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Protection
from excel_renderer import format_excel_number

from models_reporting import db, ReportInstance, ReportFieldValue, FormField
from services.excel_recalc_service import ExcelRecalcService
from services.excel_formula_engine import ExcelFormulaEngine


class ReportExporter:
    """Xuất báo cáo ra Excel"""

    @staticmethod
    def _effective_reporting_unit(instance):
        from models import User

        user = db.session.get(User, instance.user_id) if instance else None
        user_fullname = (getattr(user, 'fullname', '') or '').strip()
        stored_unit = (getattr(instance, 'org_unit', '') or '').strip()
        if user_fullname and user_fullname != stored_unit:
            return user_fullname
        return stored_unit or user_fullname

    @staticmethod
    def _resolve_target_cell(ws, cell_ref):
        """Nếu ô nằm trong merged range (không phải top-left) thì chuyển về top-left để tránh lỗi ghi."""
        from openpyxl.utils.cell import coordinate_to_tuple
        try:
            r_idx, c_idx = coordinate_to_tuple(cell_ref)
        except Exception:
            return cell_ref

        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= r_idx <= merged_range.max_row and \
               merged_range.min_col <= c_idx <= merged_range.max_col:
                return merged_range.start_cell.coordinate
        return cell_ref

    @staticmethod
    def _find_target_row(ws, org_unit):
        from utils import is_unit_match

        for column_index in (2, 1):
            for r in range(1, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=column_index).value
                if cell_val and isinstance(cell_val, str) and is_unit_match(cell_val, org_unit):
                    return r
        return None

    @staticmethod
    def _fallback_target_row(instance, ws):
        metadata = {}
        try:
            metadata = json.loads(instance.version.metadata_json) if instance and instance.version and instance.version.metadata_json else {}
        except Exception:
            metadata = {}

        effective = metadata.get('effective_structure') or {}
        scan_summary = metadata.get('scan_summary') or {}
        candidate = (
            effective.get('data_start_row') or
            scan_summary.get('data_start_row') or
            metadata.get('data_start_row') or
            2
        )
        try:
            candidate = int(candidate)
        except Exception:
            candidate = 2
        max_row = ws.max_row or candidate
        return max(1, min(candidate, max_row))

    def resolve_target_row(self, instance, ws):
        matched = self._find_target_row(ws, self._effective_reporting_unit(instance))
        if matched:
            return matched, 'matched'
        return self._fallback_target_row(instance, ws), 'fallback'

    @staticmethod
    def _safe_filename(instance):
        import re
        safe_unit = (instance.org_unit or 'unit')
        safe_unit = re.sub(r'[^A-Za-z0-9._-]+', '_', safe_unit).strip('_') or 'unit'
        return f"report_{instance.id}_{safe_unit}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    @staticmethod
    def _materialize_formula_cells(wb):
        evaluator = ExcelFormulaEngine(wb)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        try:
                            cell.value = evaluator.evaluate_cell(ws, cell.coordinate)
                        except Exception:
                            pass
        return wb

    def build_workbook(self, instance_id, protect_cells=False):
        instance = ReportInstance.query.get(instance_id)
        if not instance:
            raise ValueError(f"Report instance {instance_id} không tồn tại")

        values_map = {fv.field_code: fv.value for fv in instance.field_values}
        fields = FormField.query.filter_by(version_id=instance.version_id).all()

        if instance.template.excel_template_blob:
            wb = load_workbook(BytesIO(instance.template.excel_template_blob))
        else:
            wb = Workbook()

        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.calculation.calcOnSave = True
        except Exception:
            pass

        ws = wb.active
        target_row, target_row_source = self.resolve_target_row(instance, ws)

        for field in fields:
            raw_value = values_map.get(field.field_code)
            if raw_value in (None, '') or not field.excel_cell_ref:
                continue

            if field.field_type == 'number' or field.data_type in ['integer', 'decimal']:
                try:
                    raw_str = str(raw_value).replace(',', '')
                    raw_value = float(raw_str) if '.' in raw_str else int(raw_str)
                except ValueError:
                    pass

            cell_ref_str = str(field.excel_cell_ref).strip()
            if re.match(r'^[A-Za-z]+$', cell_ref_str):
                if target_row:
                    cell_ref = f"{cell_ref_str}{target_row}"
                else:
                    continue
            else:
                cell_ref = cell_ref_str

            target_cell = self._resolve_target_cell(ws, cell_ref)
            try:
                ws[target_cell] = raw_value
            except Exception:
                continue

        if protect_cells:
            # Lock calculated/read-only cells so the spreadsheet behaves like a form on desktop.
            try:
                for field in fields:
                    if not field.excel_cell_ref:
                        continue
                    cell_ref_str = str(field.excel_cell_ref).strip()
                    if re.match(r'^[A-Za-z]+$', cell_ref_str):
                        if not target_row:
                            continue
                        cell_ref = f"{cell_ref_str}{target_row}"
                    else:
                        cell_ref = cell_ref_str
                    target_cell = self._resolve_target_cell(ws, cell_ref)
                    ws[target_cell].protection = Protection(locked=bool(field.is_readonly or field.is_calculated))
                ws.protection.sheet = True
                ws.protection.enable()
            except Exception:
                pass

        return wb

    def build_workbook_bytes(self, instance_id, protect_cells=False, recalculate=False):
        wb = self.build_workbook(instance_id, protect_cells=protect_cells)
        if recalculate and not ExcelRecalcService.is_available():
            wb = self._materialize_formula_cells(wb)
        output = BytesIO()
        wb.save(output)
        workbook_bytes = output.getvalue()
        if recalculate and ExcelRecalcService.is_available():
            workbook_bytes = ExcelRecalcService.recalc_xlsx_bytes(workbook_bytes)
        return workbook_bytes

    def import_workbook_values(self, instance_id, file_bytes):
        instance = ReportInstance.query.get(instance_id)
        if not instance:
            raise ValueError(f"Report instance {instance_id} không tồn tại")

        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        fields = FormField.query.filter_by(version_id=instance.version_id).all()
        target_row = self._find_target_row(ws, self._effective_reporting_unit(instance))

        values_written = {}
        for field in fields:
            if not field.excel_cell_ref:
                continue

            cell_ref_str = str(field.excel_cell_ref).strip()
            if re.match(r'^[A-Za-z]+$', cell_ref_str):
                if target_row:
                    cell_ref = f"{cell_ref_str}{target_row}"
                else:
                    continue
            else:
                cell_ref = cell_ref_str

            target_cell = self._resolve_target_cell(ws, cell_ref)
            value = ws[target_cell].value
            if value is None:
                continue

            if field.field_type == 'number' or field.data_type in ['integer', 'decimal']:
                try:
                    value = format_excel_number(float(str(value).replace(',', '').strip()), None)
                except Exception:
                    value = str(value).strip()
            else:
                value = str(value).strip()

            values_written[field.field_code] = value
            fv = ReportFieldValue.query.filter_by(
                instance_id=instance_id,
                field_code=field.field_code
            ).first()
            if fv:
                fv.value = value
                fv.value_type = 'decimal' if field.field_type == 'number' else 'text'
            else:
                db.session.add(ReportFieldValue(
                    instance_id=instance_id,
                    field_code=field.field_code,
                    value=value,
                    value_type='decimal' if field.field_type == 'number' else 'text'
                ))

        db.session.commit()
        return values_written

    def export_to_excel_bytes(self, instance_id):
        instance = ReportInstance.query.get(instance_id)
        if not instance:
            raise ValueError(f"Report instance {instance_id} không tồn tại")

        workbook_bytes = self.build_workbook_bytes(instance_id, protect_cells=False, recalculate=True)
        output = BytesIO(workbook_bytes)
        output.seek(0)

        return output, self._safe_filename(instance)
