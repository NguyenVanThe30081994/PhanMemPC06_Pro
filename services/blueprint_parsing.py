# -*- coding: utf-8 -*-
"""
Phân tích tài liệu tham chiếu thành blueprint công việc (Word đề cương,
Excel mẫu, Google Form từ xa).

Tách từ routes/tasks.py (Pha 2). routes/tasks.py vẫn re-export toàn bộ tên cũ
nên các chỗ gọi hiện có không đổi.
"""

import io
import json
import os
import re
from decimal import Decimal, InvalidOperation

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from flask import current_app, jsonify, request, session

from google_forms import (
    build_google_forms_service,
    extract_google_form_id,
    load_google_form_into_builder,
    parse_google_form_definition,
)
from permissions import current_is_admin
from services.outline_engine import _parse_outline_upload_titles
from services.task_import_drafts import _parse_task_workflow_blueprint_payload
from services.task_permissions import _can_process_task_module, _current_perms
from task_blueprints import (
    normalize_task_workflow_blueprint,
    workflow_blueprint_preview_data,
)
from utils import remove_accents

TASK_BLUEPRINT_IMPORT_ALLOWED_EXTENSIONS = {".docx", ".txt", ".xlsx"}
TASK_BLUEPRINT_IMPORT_MODES = {
    "docx_outline": {
        "source_kind": "directive",
        "collection_mode": "outline",
        "default_title": "Đề cương công tác",
    },
    "docx_report_outline": {
        "source_kind": "sectioned_report",
        "collection_mode": "outline",
        "default_title": "Đề cương báo cáo",
    },
    "xlsx_form": {
        "source_kind": "excel_template",
        "collection_mode": "form",
        "default_title": "Biểu mẫu số liệu",
    },
    "google_form_remote": {
        "source_kind": "google_form",
        "collection_mode": "form",
        "default_title": "Biểu mẫu Google Form",
    },
}

def _blueprint_title_from_filename(filename, fallback):
    stem = os.path.splitext(os.path.basename(str(filename or "").strip()))[0]
    stem = stem.replace("_", " ").replace("-", " ")
    stem = re.sub(r"\s+", " ", stem).strip()
    return (stem or fallback or "Điều hành và thu báo cáo")[:255]

def _coerce_excel_sample_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Có" if value else "Không"
    return str(value).strip()

def _looks_like_number(value):
    text = _coerce_excel_sample_text(value).replace(",", "").strip()
    if not text:
        return False
    try:
        Decimal(text)
        return True
    except (InvalidOperation, ValueError):
        return False

def _infer_excel_blueprint_field_type(label, samples):
    compact_label = remove_accents(str(label or "")).strip().lower()
    if any(token in compact_label for token in ("so ", "số ", "tong", "tổng", "ty le", "tỷ lệ", "%", "chi tieu", "chỉ tiêu")):
        return "number"

    non_empty_samples = [_coerce_excel_sample_text(value) for value in (samples or []) if _coerce_excel_sample_text(value)]
    if not non_empty_samples:
        return "text"

    numeric_ratio = sum(1 for value in non_empty_samples if _looks_like_number(value)) / max(len(non_empty_samples), 1)
    if numeric_ratio >= 0.7:
        return "number"

    if max(len(value) for value in non_empty_samples) >= 80:
        return "textarea"
    return "text"

def _pick_excel_header_row(rows):
    best_index = None
    best_score = -1
    for index, row in enumerate(rows[:10]):
        non_empty = [cell for cell in row if cell]
        if len(non_empty) < 2:
            continue
        unique_count = len({cell.lower() for cell in non_empty})
        score = unique_count * 10 - index
        if score > best_score:
            best_score = score
            best_index = index
    return 0 if best_index is None and rows else best_index

def _parse_excel_template_blueprint(file_storage):
    if load_workbook is None:
        raise ValueError("Máy chủ chưa cài thư viện đọc file Excel (.xlsx).")

    extension = os.path.splitext(file_storage.filename or "")[1].lower()
    if extension == ".xls":
        raise ValueError("Hiện mới hỗ trợ file Excel .xlsx. Hãy chuyển file .xls sang .xlsx trước khi nạp.")

    try:
        file_storage.stream.seek(0)
        workbook = load_workbook(io.BytesIO(file_storage.stream.read()), data_only=True)
    except Exception:
        raise ValueError("Không đọc được file Excel. Hãy kiểm tra lại định dạng .xlsx.")

    worksheet = None
    for candidate in workbook.worksheets:
        if candidate.max_row <= 0 or candidate.max_column <= 0:
            continue
        worksheet = candidate
        break
    if worksheet is None:
        raise ValueError("Không tìm thấy sheet dữ liệu hợp lệ trong file Excel.")

    rows = []
    for row in worksheet.iter_rows(values_only=True):
        normalized_row = [_coerce_excel_sample_text(value) for value in row]
        if any(normalized_row):
            rows.append(normalized_row)
    if not rows:
        raise ValueError("File Excel chưa có dữ liệu để suy luận biểu mẫu.")

    header_index = _pick_excel_header_row(rows)
    if header_index is None:
        raise ValueError("Không xác định được dòng tiêu đề trong file Excel.")

    header_row = rows[header_index]
    header_cells = [
        (column_index, label.strip())
        for column_index, label in enumerate(header_row)
        if str(label or "").strip()
    ]
    if not header_cells:
        raise ValueError("Không tìm thấy cột hợp lệ trong dòng tiêu đề Excel.")

    sample_rows = rows[header_index + 1 : header_index + 16]
    form_fields = []
    for column_index, label in header_cells:
        samples = [
            sample_row[column_index]
            for sample_row in sample_rows
            if column_index < len(sample_row) and _coerce_excel_sample_text(sample_row[column_index])
        ]
        form_fields.append(
            {
                "label": label[:255],
                "type": _infer_excel_blueprint_field_type(label, samples),
                "required": False,
            }
        )

    blueprint = normalize_task_workflow_blueprint(
        {
            "title": _blueprint_title_from_filename(file_storage.filename, worksheet.title or "Biểu mẫu số liệu"),
            "source_kind": "excel_template",
            "collection_mode": "form",
            "form_fields": form_fields,
            "meta": {
                "sheet_name": worksheet.title,
                "header_row_index": header_index + 1,
            },
        }
    )
    if not blueprint:
        raise ValueError("Không thể chuyển file Excel thành blueprint hợp lệ.")
    return blueprint

def _blueprint_form_fields_from_google_form_payload(form_payload):
    raw_fields = []
    field_defs, _question_map = parse_google_form_definition(form_payload)
    for field_def in field_defs:
        options_payload = {}
        raw_options = field_def.get("field_options_json")
        if raw_options:
            try:
                options_payload = json.loads(raw_options)
            except Exception:
                options_payload = {}

        raw_field = {
            "label": field_def.get("field_label", ""),
            "type": field_def.get("field_type") or "text",
            "required": bool(field_def.get("is_required")),
        }
        if raw_field["type"] in {"radio", "checkbox"}:
            raw_field["choices"] = list(options_payload.get("choices") or [])
        elif raw_field["type"] == "table":
            raw_field["columns"] = list(options_payload.get("columns") or [])
        raw_fields.append(raw_field)
    return raw_fields

def _parse_google_form_reference_to_blueprint(form_reference):
    form_id = extract_google_form_id(form_reference)
    if not form_id:
        raise ValueError("Không nhận diện được Google Form URL hoặc form ID.")

    try:
        service = build_google_forms_service(current_app.config)
        imported = load_google_form_into_builder(service, form_id)
    except Exception as exc:
        raise ValueError(str(exc) or "Không thể đọc cấu trúc Google Form.") from exc

    form_payload = imported.get("form_payload") if isinstance(imported, dict) else {}
    info = form_payload.get("info") if isinstance(form_payload, dict) else {}
    title = str((info or {}).get("title") or "").strip()[:255]
    if not title:
        title = TASK_BLUEPRINT_IMPORT_MODES["google_form_remote"]["default_title"]

    blueprint = normalize_task_workflow_blueprint(
        {
            "title": title,
            "source_kind": "google_form",
            "collection_mode": "form",
            "form_fields": _blueprint_form_fields_from_google_form_payload(form_payload),
            "meta": {
                "google_form_id": form_id,
                "google_form_url": form_reference,
            },
        }
    )
    if not blueprint:
        raise ValueError("Không thể chuyển Google Form thành blueprint hợp lệ.")
    return blueprint

def _parse_reference_file_to_blueprint(file_storage, import_mode, form_reference=""):
    import_config = TASK_BLUEPRINT_IMPORT_MODES.get(str(import_mode or "").strip())
    if not import_config:
        raise ValueError("Chưa chọn kiểu phân tích tài liệu tham chiếu.")

    if import_mode == "google_form_remote":
        return _parse_google_form_reference_to_blueprint(form_reference)

    if not file_storage or not getattr(file_storage, "filename", ""):
        raise ValueError("Cần chọn tài liệu tham chiếu trước khi phân tích.")

    extension = os.path.splitext(file_storage.filename or "")[1].lower()
    if extension == ".doc":
        raise ValueError("File .doc chưa được hỗ trợ. Hãy chuyển sang .docx trước khi nạp.")
    if extension == ".xls":
        raise ValueError("Hiện mới hỗ trợ file Excel .xlsx. Hãy chuyển file .xls sang .xlsx trước khi nạp.")
    if extension not in TASK_BLUEPRINT_IMPORT_ALLOWED_EXTENSIONS:
        raise ValueError("Chỉ hỗ trợ tài liệu .docx, .txt hoặc .xlsx.")

    if import_mode == "xlsx_form":
        return _parse_excel_template_blueprint(file_storage)

    titles = _parse_outline_upload_titles(file_storage)
    if not titles:
        raise ValueError("Không tìm thấy đầu mục hợp lệ trong tài liệu tham chiếu.")

    blueprint = normalize_task_workflow_blueprint(
        {
            "title": _blueprint_title_from_filename(file_storage.filename, import_config["default_title"]),
            "source_kind": import_config["source_kind"],
            "collection_mode": import_config["collection_mode"],
            "items": [
                {
                    "title": title,
                    "report_kind": "narrative",
                    "attachment_required": False,
                }
                for title in titles
            ],
        }
    )
    if not blueprint:
        raise ValueError("Không thể chuyển tài liệu tham chiếu thành blueprint hợp lệ.")
    return blueprint


def _parse_task_workflow_blueprint_from_request(form):
    """Đọc + chuẩn hóa blueprint điều hành từ form (Pha 2 đợt 11: chuyển từ routes/tasks.py)."""
    raw_blueprint = (form.get("workflow_blueprint_json") or "").strip()
    if not raw_blueprint:
        return None

    try:
        parsed = json.loads(raw_blueprint)
    except Exception as exc:
        raise ValueError("Blueprint điều hành không hợp lệ.") from exc

    normalized = normalize_task_workflow_blueprint(parsed)
    if not normalized:
        raise ValueError("Blueprint điều hành chưa có nội dung hợp lệ.")
    return normalized


def _preview_workflow_blueprint():
    """Xem trước blueprint từ dữ liệu JSON gửi lên (Pha 2 đợt 12: tách từ routes/tasks.py)."""
    if not session.get("uid"):
        return jsonify({"ok": False, "error": "Phiên làm việc đã hết hạn."}), 401

    perms = _current_perms()
    is_admin = bool(current_is_admin())
    if not (is_admin or _can_process_task_module(perms)):
        return jsonify({"ok": False, "error": "Bạn không có quyền phân tích blueprint."}), 403

    try:
        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = request.form.to_dict(flat=True)
            raw_blueprint = (payload.get("workflow_blueprint_json") or "").strip()
            if raw_blueprint:
                payload = json.loads(raw_blueprint)
        blueprint = _parse_task_workflow_blueprint_payload(payload)
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    except Exception:
        return jsonify({"ok": False, "error": "Blueprint điều hành không hợp lệ."}), 400

    return jsonify({"ok": True, "preview": workflow_blueprint_preview_data(blueprint)})


def _import_workflow_blueprint():
    """Import blueprint từ file/doc/form tham chiếu (Pha 2 đợt 12: tách từ routes/tasks.py)."""
    if not session.get("uid"):
        return jsonify({"ok": False, "error": "Phiên làm việc đã hết hạn."}), 401

    perms = _current_perms()
    is_admin = bool(current_is_admin())
    if not (is_admin or _can_process_task_module(perms)):
        return jsonify({"ok": False, "error": "Bạn không có quyền phân tích tài liệu tham chiếu."}), 403

    file_storage = request.files.get("blueprint_source_file")
    import_mode = (request.form.get("blueprint_import_mode") or "").strip()
    form_reference = (request.form.get("blueprint_form_reference") or "").strip()
    try:
        blueprint = _parse_reference_file_to_blueprint(
            file_storage, import_mode, form_reference=form_reference
        )
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400

    return jsonify(
        {
            "ok": True,
            "workflow_blueprint": blueprint,
            "preview": workflow_blueprint_preview_data(blueprint),
        }
    )

