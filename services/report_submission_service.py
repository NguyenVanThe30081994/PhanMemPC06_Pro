# -*- coding: utf-8 -*-
"""
Submission-centric reporting workflow aligned with the Excel-upload architecture.
"""
import datetime
import json
import re
import uuid
from io import BytesIO
from pathlib import Path

from flask import current_app
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from excel_renderer import format_excel_number
from models_reporting import (
    db,
    FormField,
    FormTemplate,
    FormVersion,
    ReportSubmission,
    ReportDataRow,
    ReportDataCell,
    ReportValidationError,
    ReportWorkflowHistory,
    ReportingPeriod,
)
from services.excel_recalc_service import ExcelRecalcService


class ReportSubmissionService:
    STATUS_DRAFT = 'DRAFT'
    STATUS_IMPORT_FAILED = 'IMPORT_FAILED'
    STATUS_SUBMITTED = 'SUBMITTED'
    STATUS_UNDER_REVIEW = 'UNDER_REVIEW'
    STATUS_RETURNED = 'RETURNED'
    STATUS_APPROVED = 'APPROVED'
    STATUS_REJECTED = 'REJECTED'
    STATUS_LOCKED = 'LOCKED'
    STATUS_CANCELLED = 'CANCELLED'

    ERROR_FILL = PatternFill(fill_type='solid', fgColor='FFF1F2')

    def _storage_root(self):
        root = Path(current_app.root_path).resolve().parent / 'uploads' / 'reporting'
        root.mkdir(parents=True, exist_ok=True)
        for name in ('originals', 'processed', 'errors'):
            (root / name).mkdir(parents=True, exist_ok=True)
        return root

    @staticmethod
    def _normalize_header_text(text):
        text = str(text or '').strip().upper()
        text = re.sub(r'\s+', ' ', text)
        text = re.sub(r'[^0-9A-ZÀ-Ỵ _./()-]+', ' ', text)
        return ' '.join(text.split())

    @staticmethod
    def _field_type_name(field):
        field_type = (field.field_type or '').strip().lower()
        data_type = (field.data_type or '').strip().lower()
        if field_type == 'number' or data_type in ('integer', 'decimal'):
            return 'decimal'
        if field_type == 'date' or data_type == 'date':
            return 'date'
        return 'string'

    @staticmethod
    def _safe_float(value):
        if value in (None, ''):
            return None
        raw = str(value).replace(',', '').strip()
        if not raw:
            return None
        return float(raw)

    def _normalize_value(self, value, column_config):
        col_type = (column_config.get('type') or 'string').strip().lower()
        if value is None:
            return None, 'string'
        if isinstance(value, datetime.datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S'), 'datetime'
        if isinstance(value, datetime.date):
            return value.isoformat(), 'date'
        if col_type in ('number', 'decimal', 'integer'):
            try:
                numeric = self._safe_float(value)
                if numeric is None:
                    return None, 'decimal'
                return format_excel_number(numeric, None), 'decimal'
            except Exception:
                return str(value).strip(), 'string'
        return str(value).strip(), 'string'

    def _serialize_error(self, error_code, message, severity='ERROR', **kwargs):
        payload = {
            'error_code': error_code,
            'error_message': message,
            'severity': severity,
        }
        payload.update(kwargs)
        return payload

    def get_published_version(self, template_id):
        version = FormVersion.query.filter_by(
            template_id=template_id,
            is_published=True
        ).order_by(FormVersion.created_at.desc()).first()
        if not version:
            raise ValueError('Biểu mẫu chưa có phiên bản đang áp dụng.')
        return version

    def _build_fingerprint(self, sheet_name, columns):
        header_signature = []
        for column in columns:
            header_signature.append({
                'column_letter': column.get('column_letter'),
                'field': column.get('field'),
                'excel_header': column.get('excel_header'),
            })
        return {
            'sheet_names': [sheet_name] if sheet_name else [],
            'header_signature': header_signature,
        }

    def build_auto_config(self, template, version):
        metadata = json.loads(version.metadata_json) if version.metadata_json else {}
        scan_summary = metadata.get('scan_summary', {})
        fields = FormField.query.filter_by(version_id=version.id).order_by(FormField.display_order.asc()).all()

        columns = []
        for field in fields:
            rules = json.loads(field.validation_rules_json) if field.validation_rules_json else {}
            column_letter = str(field.excel_cell_ref or '').strip().upper() or None
            if column_letter and re.match(r'^[A-Z]+[0-9]+$', column_letter):
                column_letter = re.sub(r'[0-9]+$', '', column_letter)
            columns.append({
                'excel_header': field.field_name,
                'field': field.field_code,
                'type': self._field_type_name(field),
                'required': bool(field.is_required),
                'readonly': bool(field.is_readonly),
                'calculated': bool(field.is_calculated),
                'hidden': bool(rules.get('hidden', False)),
                'section': field.section or 'Dữ liệu chính',
                'column_letter': column_letter,
                'min': rules.get('min'),
                'max': rules.get('max'),
            })

        header_rows = scan_summary.get('header_rows') or [metadata.get('header_rows') or 1]
        header_start = min(header_rows) if header_rows else 1
        header_end = max(header_rows) if header_rows else header_start
        data_start = scan_summary.get('data_start_row') or metadata.get('data_start_row') or (header_end + 1)
        sheet_name = scan_summary.get('sheet_name') or 'Sheet1'

        config = {
            'template_code': template.code,
            'version': version.version_number,
            'name': template.name,
            'entry_mode': 'excel_upload',
            'fingerprint': self._build_fingerprint(sheet_name, columns),
            'sheets': [
                {
                    'sheet_code': 'MAIN',
                    'sheet_name': sheet_name,
                    'metadata': [],
                    'sections': [
                        {
                            'section_code': 'MAIN_TABLE',
                            'section_name': 'Dữ liệu chính',
                            'header_start_row': header_start,
                            'header_end_row': header_end,
                            'data_start_row': data_start,
                            'stop_conditions': [
                                {'column': 'A', 'contains': 'Tổng'},
                                {'column': 'B', 'contains': 'Tổng'},
                                {'column': 'A', 'contains': 'Cộng'},
                                {'column': 'B', 'contains': 'Cộng'},
                            ],
                            'columns': columns,
                        }
                    ],
                }
            ],
            'ui': {
                'layout': 'tabs',
                'tabs': [
                    {'title': 'Thông tin chung', 'display': 'form', 'source': 'metadata'},
                    {'title': 'Dữ liệu chi tiết', 'display': 'table', 'section_code': 'MAIN_TABLE'},
                    {'title': 'Lỗi / cảnh báo', 'display': 'errors'},
                    {'title': 'Lịch sử xử lý', 'display': 'history'},
                ],
            },
        }
        return config

    def ensure_version_config(self, template, version, force=False):
        metadata = json.loads(version.metadata_json) if version.metadata_json else {}
        config = metadata.get('template_config')
        if force or not config:
            config = self.build_auto_config(template, version)
            metadata['template_config'] = config
            version.metadata_json = json.dumps(metadata, ensure_ascii=False)
            db.session.flush()
        return config

    def _merged_value(self, worksheet, merged_ranges, row_idx, col_idx):
        for merged in merged_ranges:
            min_col, min_row, max_col, max_row = merged.bounds
            if min_row <= row_idx <= max_row and min_col <= col_idx <= max_col:
                return worksheet.cell(min_row, min_col).value
        return worksheet.cell(row_idx, col_idx).value

    def _extract_header_map(self, worksheet, section_cfg):
        header_map = {}
        header_start = int(section_cfg.get('header_start_row') or 1)
        header_end = int(section_cfg.get('header_end_row') or header_start)
        merged_ranges = list(worksheet.merged_cells.ranges)

        for column_cfg in section_cfg.get('columns', []):
            col_letter = (column_cfg.get('column_letter') or '').strip().upper()
            if not col_letter:
                continue
            col_idx = column_index_from_string(col_letter)
            parts = []
            for row_idx in range(header_start, header_end + 1):
                text = self._merged_value(worksheet, merged_ranges, row_idx, col_idx)
                normalized = str(text).strip() if text is not None else ''
                if normalized and normalized not in parts:
                    parts.append(normalized)
            header_map[col_letter] = ' / '.join(parts).strip()
        return header_map

    def _check_stop_conditions(self, worksheet, row_idx, stop_conditions):
        for condition in stop_conditions or []:
            col_letter = (condition.get('column') or '').strip().upper()
            if not col_letter:
                continue
            cell_text = str(worksheet[f'{col_letter}{row_idx}'].value or '').strip()
            if not cell_text:
                continue
            contains = condition.get('contains')
            equals = condition.get('equals')
            if contains and contains.lower() in cell_text.lower():
                return True
            if equals and equals.lower() == cell_text.lower():
                return True
        return False

    def _validate_column(self, normalized_value, column_cfg, sheet_name, section_code, row_idx, col_idx, cell_address):
        errors = []
        col_type = (column_cfg.get('type') or 'string').strip().lower()
        if column_cfg.get('required') and normalized_value in (None, ''):
            errors.append(self._serialize_error(
                'REQUIRED_VALUE_MISSING',
                'Thiếu giá trị bắt buộc',
                sheet_name=sheet_name,
                section_code=section_code,
                row_index=row_idx,
                column_index=col_idx,
                cell_address=cell_address,
                field_code=column_cfg.get('field'),
            ))
            return errors

        if normalized_value in (None, ''):
            return errors

        if col_type in ('number', 'decimal', 'integer'):
            try:
                numeric_value = self._safe_float(normalized_value)
            except Exception:
                numeric_value = None
            if numeric_value is None:
                errors.append(self._serialize_error(
                    'INVALID_NUMBER',
                    'Giá trị không đúng định dạng số',
                    sheet_name=sheet_name,
                    section_code=section_code,
                    row_index=row_idx,
                    column_index=col_idx,
                    cell_address=cell_address,
                    field_code=column_cfg.get('field'),
                ))
                return errors
            min_value = column_cfg.get('min')
            max_value = column_cfg.get('max')
            if min_value is not None and numeric_value < float(min_value):
                errors.append(self._serialize_error(
                    'VALUE_BELOW_MIN',
                    f'Giá trị nhỏ hơn mức tối thiểu {min_value}',
                    sheet_name=sheet_name,
                    section_code=section_code,
                    row_index=row_idx,
                    column_index=col_idx,
                    cell_address=cell_address,
                    field_code=column_cfg.get('field'),
                ))
            if max_value is not None and numeric_value > float(max_value):
                errors.append(self._serialize_error(
                    'VALUE_ABOVE_MAX',
                    f'Giá trị lớn hơn mức tối đa {max_value}',
                    sheet_name=sheet_name,
                    section_code=section_code,
                    row_index=row_idx,
                    column_index=col_idx,
                    cell_address=cell_address,
                    field_code=column_cfg.get('field'),
                ))
        return errors

    def _parse_metadata(self, worksheet, metadata_cfg):
        values = {}
        for item in metadata_cfg or []:
            cell_ref = (item.get('cell') or '').strip()
            field_code = (item.get('field') or '').strip()
            if not cell_ref or not field_code:
                continue
            values[field_code] = worksheet[cell_ref].value
        return values

    def parse_workbook_bytes(self, file_bytes, config):
        workbook_bytes = file_bytes
        recalc_mode = 'cached'
        if ExcelRecalcService.is_available():
            workbook_bytes = ExcelRecalcService.recalc_xlsx_bytes(file_bytes)
            recalc_mode = 'libreoffice'

        wb_formula = load_workbook(BytesIO(workbook_bytes), data_only=False)
        wb_values = load_workbook(BytesIO(workbook_bytes), data_only=True)

        parsed_rows = []
        validation_errors = []
        metadata_values = {}
        total_rows = 0
        valid_rows = 0
        invalid_rows = 0
        warning_count = 0

        expected_sheets = {sheet_cfg.get('sheet_name') for sheet_cfg in config.get('sheets', []) if sheet_cfg.get('sheet_name')}
        actual_sheets = set(wb_formula.sheetnames)
        for expected_sheet in sorted(expected_sheets - actual_sheets):
            validation_errors.append(self._serialize_error(
                'SHEET_NOT_FOUND',
                f'Không tìm thấy sheet "{expected_sheet}" trong file upload',
                severity='ERROR',
                sheet_name=expected_sheet,
            ))

        for sheet_cfg in config.get('sheets', []):
            sheet_name = sheet_cfg.get('sheet_name')
            if not sheet_name or sheet_name not in wb_formula.sheetnames:
                continue

            ws_formula = wb_formula[sheet_name]
            ws_values = wb_values[sheet_name] if sheet_name in wb_values.sheetnames else wb_values[wb_values.sheetnames[0]]
            metadata_values[sheet_name] = self._parse_metadata(ws_values, sheet_cfg.get('metadata'))

            for section_cfg in sheet_cfg.get('sections', []):
                section_code = section_cfg.get('section_code') or 'MAIN_TABLE'
                header_map = self._extract_header_map(ws_formula, section_cfg)

                for column_cfg in section_cfg.get('columns', []):
                    col_letter = (column_cfg.get('column_letter') or '').strip().upper()
                    if not col_letter:
                        validation_errors.append(self._serialize_error(
                            'REQUIRED_COLUMN_MISSING',
                            f'Chưa cấu hình cột Excel cho trường {column_cfg.get("field")}',
                            sheet_name=sheet_name,
                            section_code=section_code,
                        ))
                        continue
                    actual_header = self._normalize_header_text(header_map.get(col_letter))
                    expected_header = self._normalize_header_text(column_cfg.get('excel_header'))
                    if expected_header and actual_header and expected_header not in actual_header and actual_header not in expected_header:
                        validation_errors.append(self._serialize_error(
                            'HEADER_NOT_FOUND',
                            f'Header cột {col_letter} không khớp cấu hình ({header_map.get(col_letter) or "trống"})',
                            sheet_name=sheet_name,
                            section_code=section_code,
                            column_index=column_index_from_string(col_letter),
                            cell_address=f'{col_letter}{section_cfg.get("header_end_row")}',
                            field_code=column_cfg.get('field'),
                        ))

                blank_streak = 0
                data_start_row = int(section_cfg.get('data_start_row') or (int(section_cfg.get('header_end_row') or 1) + 1))
                for row_idx in range(data_start_row, ws_formula.max_row + 1):
                    if self._check_stop_conditions(ws_formula, row_idx, section_cfg.get('stop_conditions')):
                        break

                    row_payload = []
                    row_empty = True
                    row_errors = []
                    for column_cfg in section_cfg.get('columns', []):
                        col_letter = (column_cfg.get('column_letter') or '').strip().upper()
                        if not col_letter:
                            continue
                        col_idx = column_index_from_string(col_letter)
                        formula_cell = ws_formula.cell(row=row_idx, column=col_idx)
                        value_cell = ws_values.cell(row=row_idx, column=col_idx)

                        normalized_value, value_type = self._normalize_value(value_cell.value, column_cfg)
                        formula_text = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith('=') else None
                        if normalized_value not in (None, '') or formula_text:
                            row_empty = False

                        cell_address = f'{col_letter}{row_idx}'
                        column_errors = self._validate_column(
                            normalized_value,
                            column_cfg,
                            sheet_name,
                            section_code,
                            row_idx,
                            col_idx,
                            cell_address
                        )
                        if formula_text and value_cell.value is None and not ExcelRecalcService.is_available():
                            column_errors.append(self._serialize_error(
                                'FORMULA_MISMATCH',
                                'Ô công thức chưa có giá trị tính sẵn trên máy chủ.',
                                severity='WARNING',
                                sheet_name=sheet_name,
                                section_code=section_code,
                                row_index=row_idx,
                                column_index=col_idx,
                                cell_address=cell_address,
                                field_code=column_cfg.get('field'),
                            ))

                        row_errors.extend(column_errors)
                        row_payload.append({
                            'field_code': column_cfg.get('field'),
                            'excel_address': cell_address,
                            'raw_value': '' if formula_cell.value is None else str(formula_cell.value),
                            'normalized_value': normalized_value,
                            'value_type': value_type,
                            'formula_text': formula_text,
                        })

                    if row_empty:
                        blank_streak += 1
                        if blank_streak >= 2:
                            break
                        continue

                    blank_streak = 0
                    total_rows += 1
                    row_status = 'VALID'
                    if any(err['severity'] == 'ERROR' for err in row_errors):
                        row_status = 'INVALID'
                        invalid_rows += 1
                    else:
                        valid_rows += 1
                    warning_count += sum(1 for err in row_errors if err['severity'] == 'WARNING')

                    parsed_rows.append({
                        'sheet_code': sheet_cfg.get('sheet_code') or sheet_name,
                        'sheet_name': sheet_name,
                        'section_code': section_code,
                        'row_index': row_idx,
                        'status': row_status,
                        'cells': row_payload,
                    })
                    validation_errors.extend(row_errors)

        if total_rows == 0:
            validation_errors.append(self._serialize_error(
                'NO_DATA_FOUND',
                'Không đọc được dòng dữ liệu nào từ file Excel.',
                severity='ERROR',
            ))

        return {
            'processed_bytes': workbook_bytes,
            'recalc_mode': recalc_mode,
            'rows': parsed_rows,
            'errors': validation_errors,
            'metadata_values': metadata_values,
            'total_rows': total_rows,
            'valid_rows': valid_rows,
            'invalid_rows': invalid_rows,
            'warning_count': warning_count,
        }

    def _write_bytes(self, folder_name, filename, file_bytes):
        target = self._storage_root() / folder_name / filename
        target.write_bytes(file_bytes)
        return str(target)

    @staticmethod
    def _resolve_target_cell(ws, cell_ref):
        from openpyxl.utils.cell import coordinate_to_tuple

        try:
            row_idx, col_idx = coordinate_to_tuple(cell_ref)
        except Exception:
            return cell_ref

        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= row_idx <= merged_range.max_row and merged_range.min_col <= col_idx <= merged_range.max_col:
                return merged_range.start_cell.coordinate
        return cell_ref

    def _coerce_workbook_value(self, cell_data):
        formula_text = str(cell_data.get('formula_text') or '').strip()
        if formula_text:
            return formula_text

        value_type = str(cell_data.get('value_type') or '').strip().lower()
        normalized_value = cell_data.get('normalized_value')
        raw_value = cell_data.get('raw_value')
        candidate = normalized_value if normalized_value not in (None, '') else raw_value
        if candidate in (None, ''):
            return None

        if value_type in ('decimal', 'number', 'integer'):
            try:
                numeric = self._safe_float(candidate)
                if numeric is None:
                    return candidate
                if value_type == 'integer' and float(numeric).is_integer():
                    return int(round(numeric))
                return numeric
            except Exception:
                return candidate

        if value_type == 'date':
            try:
                return datetime.date.fromisoformat(str(candidate))
            except Exception:
                return candidate

        if value_type == 'datetime':
            try:
                return datetime.datetime.fromisoformat(str(candidate))
            except Exception:
                return candidate

        return candidate

    def _filled_workbook_bytes(self, file_bytes, parsed_rows, recalc=False):
        wb = load_workbook(BytesIO(file_bytes))
        for row_data in parsed_rows or []:
            sheet_name = row_data.get('sheet_name')
            if not sheet_name or sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for cell_data in row_data.get('cells', []):
                cell_ref = (cell_data.get('excel_address') or '').strip()
                if not cell_ref:
                    continue
                target_cell = self._resolve_target_cell(ws, cell_ref)
                try:
                    ws[target_cell] = self._coerce_workbook_value(cell_data)
                except Exception:
                    continue

        output = BytesIO()
        wb.save(output)
        workbook_bytes = output.getvalue()
        if recalc and ExcelRecalcService.is_available():
            workbook_bytes = ExcelRecalcService.recalc_xlsx_bytes(workbook_bytes)
        return workbook_bytes

    def _error_workbook_bytes(self, file_bytes, errors):
        wb = load_workbook(BytesIO(file_bytes))
        ws_errors = wb.create_sheet('Danh sách lỗi')
        ws_errors.append(['Sheet', 'Ô lỗi', 'Dòng', 'Cột', 'Trường', 'Mức độ', 'Nội dung lỗi'])
        grouped_comments = {}
        for error in errors:
            sheet_name = error.get('sheet_name')
            cell_address = error.get('cell_address')
            if sheet_name and cell_address and sheet_name in wb.sheetnames:
                key = (sheet_name, cell_address)
                grouped_comments.setdefault(key, []).append(error.get('error_message'))
                ws = wb[sheet_name]
                ws[cell_address].fill = self.ERROR_FILL
            ws_errors.append([
                sheet_name or '',
                cell_address or '',
                error.get('row_index') or '',
                error.get('column_index') or '',
                error.get('field_code') or '',
                error.get('severity') or '',
                error.get('error_message') or '',
            ])

        for (sheet_name, cell_address), messages in grouped_comments.items():
            wb[sheet_name][cell_address].comment = Comment('\n'.join(messages), 'System')

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def _sync_submission_rows(self, submission, parsed):
        ReportDataCell.query.filter(
            ReportDataCell.row_id.in_(
                db.session.query(ReportDataRow.id).filter(ReportDataRow.submission_id == submission.id)
            )
        ).delete(synchronize_session=False)
        ReportDataRow.query.filter_by(submission_id=submission.id).delete(synchronize_session=False)
        ReportValidationError.query.filter_by(submission_id=submission.id).delete(synchronize_session=False)

        for row_data in parsed['rows']:
            row = ReportDataRow(
                submission_id=submission.id,
                sheet_code=row_data['sheet_code'],
                section_code=row_data['section_code'],
                row_index=row_data['row_index'],
                status=row_data['status'],
                metadata_json=json.dumps({'sheet_name': row_data['sheet_name']}, ensure_ascii=False),
            )
            db.session.add(row)
            db.session.flush()
            for cell_data in row_data['cells']:
                db.session.add(ReportDataCell(
                    row_id=row.id,
                    field_code=cell_data['field_code'] or '',
                    excel_address=cell_data['excel_address'],
                    raw_value=cell_data['raw_value'],
                    normalized_value=cell_data['normalized_value'],
                    value_type=cell_data['value_type'],
                    formula_text=cell_data['formula_text'],
                ))

        for error in parsed['errors']:
            db.session.add(ReportValidationError(
                submission_id=submission.id,
                sheet_name=error.get('sheet_name'),
                section_code=error.get('section_code'),
                row_index=error.get('row_index'),
                column_index=error.get('column_index'),
                cell_address=error.get('cell_address'),
                field_code=error.get('field_code'),
                error_code=error.get('error_code'),
                error_message=error.get('error_message'),
                severity=error.get('severity') or 'ERROR',
            ))

    def _log_workflow(self, submission_id, from_status, to_status, action, actor_id, comment=None):
        db.session.add(ReportWorkflowHistory(
            submission_id=submission_id,
            from_status=from_status,
            to_status=to_status,
            action=action,
            comment=comment,
            actor_id=actor_id,
        ))

    def create_submission(self, template_id, period_id, user_id, reporting_unit, uploaded_file):
        template = db.session.get(FormTemplate, template_id)
        if not template or not template.is_active:
            raise ValueError('Biểu mẫu không tồn tại hoặc đã ngừng sử dụng.')

        period = db.session.get(ReportingPeriod, period_id)
        if not period:
            raise ValueError('Kỳ báo cáo không tồn tại.')
        if period.template_id != template_id:
            raise ValueError('Kỳ báo cáo không khớp với biểu mẫu.')
        if period.is_locked:
            raise ValueError('Kỳ báo cáo đã khóa.')

        version = self.get_published_version(template_id)
        config = self.ensure_version_config(template, version)

        original_bytes = uploaded_file.read()
        if not original_bytes:
            raise ValueError('File báo cáo trống.')

        submission = ReportSubmission(
            template_id=template_id,
            template_version_id=version.id,
            period_id=period_id,
            report_period=period.code,
            reporting_unit=reporting_unit,
            submitted_by=user_id,
            status=self.STATUS_DRAFT,
            original_filename=uploaded_file.filename,
        )
        db.session.add(submission)
        db.session.flush()

        file_token = f'{submission.id}_{uuid.uuid4().hex[:8]}'
        original_name = f'{file_token}_original.xlsx'
        original_path = self._write_bytes('originals', original_name, original_bytes)

        parsed = self.parse_workbook_bytes(original_bytes, config)
        filled_bytes = self._filled_workbook_bytes(
            original_bytes,
            parsed['rows'],
            recalc=ExcelRecalcService.is_available()
        )
        processed_name = f'{file_token}_processed.xlsx'
        processed_path = self._write_bytes('processed', processed_name, filled_bytes)

        error_path = None
        if parsed['errors']:
            error_name = f'{file_token}_errors.xlsx'
            error_bytes = self._error_workbook_bytes(filled_bytes, parsed['errors'])
            error_path = self._write_bytes('errors', error_name, error_bytes)

        submission.original_file_path = original_path
        submission.processed_file_path = processed_path
        submission.error_file_path = error_path
        submission.total_rows = parsed['total_rows']
        submission.valid_rows = parsed['valid_rows']
        submission.invalid_rows = parsed['invalid_rows']
        submission.warning_count = parsed['warning_count']
        submission.metadata_json = json.dumps({
            'recalc_mode': parsed['recalc_mode'],
            'metadata_values': parsed['metadata_values'],
            'config_version': version.version_number,
        }, ensure_ascii=False)
        submission.status = self.STATUS_IMPORT_FAILED if any(err['severity'] == 'ERROR' for err in parsed['errors']) else self.STATUS_DRAFT

        self._sync_submission_rows(submission, parsed)
        self._log_workflow(submission.id, None, submission.status, 'upload', user_id)
        db.session.commit()
        return submission

    def get_submission(self, submission_id):
        submission = db.session.get(ReportSubmission, submission_id)
        if not submission:
            raise ValueError('Lần nộp báo cáo không tồn tại.')
        return submission

    def get_submission_detail(self, submission_id):
        submission = self.get_submission(submission_id)
        version = db.session.get(FormVersion, submission.template_version_id)
        metadata = json.loads(submission.metadata_json) if submission.metadata_json else {}
        config = self.ensure_version_config(submission.template, version)

        rows = ReportDataRow.query.filter_by(submission_id=submission.id).order_by(ReportDataRow.row_index.asc()).all()
        cells = ReportDataCell.query.join(ReportDataRow, ReportDataRow.id == ReportDataCell.row_id).filter(
            ReportDataRow.submission_id == submission.id
        ).all()
        errors = ReportValidationError.query.filter_by(submission_id=submission.id).order_by(
            ReportValidationError.sheet_name.asc(),
            ReportValidationError.row_index.asc(),
            ReportValidationError.column_index.asc()
        ).all()
        history = ReportWorkflowHistory.query.filter_by(submission_id=submission.id).order_by(
            ReportWorkflowHistory.acted_at.desc()
        ).all()

        row_map = {}
        for row in rows:
            row_map[row.id] = {
                'id': row.id,
                'sheet_code': row.sheet_code,
                'section_code': row.section_code,
                'row_index': row.row_index,
                'status': row.status,
                'cells': [],
            }
        for cell in cells:
            row_map[cell.row_id]['cells'].append(cell)

        return {
            'submission': submission,
            'config': config,
            'metadata': metadata,
            'rows': list(row_map.values()),
            'errors': errors,
            'history': history,
        }

    def transition_submission(self, submission_id, action, actor_id, comment=None):
        submission = self.get_submission(submission_id)
        current_status = submission.status

        transitions = {
            'submit': ([self.STATUS_DRAFT, self.STATUS_RETURNED], self.STATUS_SUBMITTED),
            'review': ([self.STATUS_SUBMITTED], self.STATUS_UNDER_REVIEW),
            'approve': ([self.STATUS_SUBMITTED, self.STATUS_UNDER_REVIEW], self.STATUS_APPROVED),
            'return': ([self.STATUS_SUBMITTED, self.STATUS_UNDER_REVIEW], self.STATUS_RETURNED),
            'reject': ([self.STATUS_SUBMITTED, self.STATUS_UNDER_REVIEW], self.STATUS_REJECTED),
            'lock': ([self.STATUS_APPROVED], self.STATUS_LOCKED),
        }
        if action not in transitions:
            raise ValueError('Hành động workflow không hợp lệ.')

        allowed_statuses, next_status = transitions[action]
        if current_status not in allowed_statuses:
            raise ValueError(f'Không thể thực hiện thao tác {action} khi báo cáo đang ở trạng thái {current_status}.')

        if action == 'submit':
            blocking_errors = ReportValidationError.query.filter_by(
                submission_id=submission.id,
                severity='ERROR'
            ).count()
            if blocking_errors:
                raise ValueError('Báo cáo còn lỗi dữ liệu, không thể gửi.')

        submission.status = next_status
        submission.updated_at = datetime.datetime.now()
        if next_status == self.STATUS_SUBMITTED:
            submission.submitted_at = datetime.datetime.now()

        self._log_workflow(submission.id, current_status, next_status, action, actor_id, comment=comment)
        db.session.commit()
        return submission

    def read_submission_file_bytes(self, submission, prefer_processed=True):
        file_path = submission.processed_file_path if prefer_processed and submission.processed_file_path else submission.original_file_path
        if not file_path or not Path(file_path).exists():
            raise ValueError('Không tìm thấy file báo cáo trên máy chủ.')
        return Path(file_path).read_bytes()

    def build_preview_context(self, file_bytes, max_row=220, max_col=80):
        wb_formula = load_workbook(BytesIO(file_bytes), data_only=False)
        wb_values = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb_formula.active
        ws_values = wb_values[ws.title] if ws.title in wb_values.sheetnames else wb_values.active

        merged_map = {}
        covered_cells = set()
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col_idx, max_row_idx = merged_range.bounds
            top_left = (min_row, min_col)
            merged_map[top_left] = {
                'rowspan': max_row_idx - min_row + 1,
                'colspan': max_col_idx - min_col + 1,
            }
            for row_idx in range(min_row, max_row_idx + 1):
                for col_idx in range(min_col, max_col_idx + 1):
                    if (row_idx, col_idx) != top_left:
                        covered_cells.add((row_idx, col_idx))

        effective_max_row = min(ws.max_row or 1, max_row)
        effective_max_col = min(ws.max_column or 1, max_col)

        col_widths = []
        col_letters = []
        for col_idx in range(1, effective_max_col + 1):
            letter = get_column_letter(col_idx)
            col_letters.append(letter)
            width = ws.column_dimensions[letter].width
            col_widths.append(max(int((width or 8.43) * 7.5), 28))

        def _extract_color(color_obj):
            if color_obj and color_obj.type == 'rgb' and color_obj.rgb:
                rgb = str(color_obj.rgb)[-6:]
                if rgb != '000000':
                    return f'#{rgb}'
            return None

        def _border_css(border):
            if not border:
                return ''
            parts = []
            for side, css_side in [('left', 'border-left'), ('right', 'border-right'), ('top', 'border-top'), ('bottom', 'border-bottom')]:
                edge = getattr(border, side, None)
                if edge and edge.style and edge.style != 'none':
                    weight = {'thin': '1px', 'medium': '2px', 'thick': '3px', 'hair': '1px', 'dotted': '1px', 'dashed': '1px', 'double': '3px'}.get(edge.style, '1px')
                    style = 'double' if edge.style == 'double' else ('dotted' if edge.style in ('dotted', 'hair') else ('dashed' if edge.style == 'dashed' else 'solid'))
                    color = '#000'
                    if edge.color and edge.color.type == 'rgb' and edge.color.rgb:
                        color = f'#{str(edge.color.rgb)[-6:]}'
                    parts.append(f'{css_side}:{weight} {style} {color};')
            return ''.join(parts)

        rows = []
        for row_idx in range(1, effective_max_row + 1):
            cells = []
            for col_idx in range(1, effective_max_col + 1):
                if (row_idx, col_idx) in covered_cells:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                value_cell = ws_values.cell(row=row_idx, column=col_idx)
                merge_info = merged_map.get((row_idx, col_idx), {'rowspan': 1, 'colspan': 1})
                raw_value = value_cell.value
                display_value = '' if raw_value is None and isinstance(cell.value, str) and cell.value.startswith('=') else format_excel_number(raw_value, cell.number_format)
                fill = cell.fill
                font = cell.font
                alignment = cell.alignment
                cells.append({
                    'display_value': display_value,
                    'rowspan': merge_info['rowspan'],
                    'colspan': merge_info['colspan'],
                    'bold': bool(font and font.bold),
                    'italic': bool(font and font.italic),
                    'underline': bool(font and font.underline),
                    'font_size': int(font.sz) if font and font.sz else 11,
                    'font_color': _extract_color(font.color) if font and font.color else None,
                    'align': alignment.horizontal if alignment and alignment.horizontal else 'left',
                    'valign': alignment.vertical if alignment and alignment.vertical else 'center',
                    'bg_color': _extract_color(fill.fgColor) if fill and fill.fill_type and fill.fgColor else None,
                    'wrap_text': bool(alignment and alignment.wrap_text),
                    'border_style': _border_css(cell.border),
                })
            raw_height = ws.row_dimensions[row_idx].height
            rows.append({
                'height_px': max(int(raw_height * 1.0), 16) if raw_height else 20,
                'cells': cells,
            })

        return {
            'sheet_title': ws.title,
            'rows': rows,
            'col_widths': col_widths,
            'col_letters': col_letters,
        }
