# -*- coding: utf-8 -*-
from flask import Blueprint, render_template as flask_render_template, request, session, redirect, url_for, flash, send_file, jsonify
from models import db, ReportConfig, ReportData, User, ReportTemplateV2, ReportVersionV2, ReportSubmissionV2, ReportValueV2, AppRole
import json, io
from datetime import datetime
from utils import remove_accents, log_action, render_auto_template as render_template

forms_bp = Blueprint('forms_bp', __name__)

# ... (rest of existing routes kept the same) ...

@forms_bp.route('/stats/export', methods=['GET'])
def stats_export():
    """Export V1/V2 stats to Excel for Luckysheet rendering."""
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))
    
    rid = request.args.get('rid', type=int)
    is_v2 = request.args.get('v2')
    
    if not rid:
        return "Missing report ID", 400
    
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        if is_v2:
            # V2 Export logic
            template = db.session.get(ReportTemplateV2, rid)
            if not template:
                return "Template not found", 404
            
            version = ReportVersionV2.query.filter_by(template_id=rid, is_published=True).order_by(ReportVersionV2.created_at.desc()).first()
            if not version or not version.excel_file_blob:
                return "No published version", 400
            
            try:
                wb = openpyxl.load_workbook(io.BytesIO(version.excel_file_blob), data_only=False)
            except:
                wb = openpyxl.load_workbook(io.BytesIO(version.excel_file_blob))
            
            ws = wb.active
            
            # Fetch ALL submissions
            raw_subs = ReportSubmissionV2.query.filter_by(version_id=version.id).all()
            sub_ids = [s.id for s in raw_subs]
            
            if sub_ids:
                all_vals = ReportValueV2.query.filter(ReportValueV2.submission_id.in_(sub_ids)).all()
                
                # Build map: coord -> {unit: value}
                coord_values = {}
                for v in all_vals:
                    key = str(v.cell_key or '').strip()
                    if '!' in key:
                        _, coord = key.split('!', 1)
                    else:
                        coord = key
                    
                    if coord not in coord_values:
                        coord_values[coord] = {}
                    # Find unit for this submission
                    sub = next((s for s in raw_subs if s.id == v.submission_id), None)
                    if sub:
                        unit = sub.org_unit or 'Unknown'
                        coord_values[coord][unit] = v.value
                
                # Write data to sheet - simple row-by-row append
                row_idx = ws.max_row + 2
                ws.cell(row_idx, 1).value = "--- TỔNG HỢP DỮ LIỆU TỪ CÁC ĐƠN VỊ ---"
                row_idx += 1
                
                for sub in raw_subs:
                    row_idx += 1
                    unit = sub.org_unit or 'Unknown'
                    ws.cell(row_idx, 1).value = unit
                    ws.cell(row_idx, 2).value = sub.updated_at.strftime('%d/%m/%Y') if sub.updated_at else ''
                    
                    # Write values
                    sub_vals = [v for v in all_vals if v.submission_id == sub.id]
                    for v in sub_vals:
                        key = str(v.cell_key or '').strip()
                        if '!' in key:
                            _, coord = key.split('!', 1)
                        else:
                            coord = key
                        
                        try:
                            # Parse coord like "A5"
                            from openpyxl.utils import coordinate_to_tuple
                            col, row = coordinate_to_tuple(coord)
                            ws.cell(row, col).value = v.value
                        except:
                            pass
        else:
            # V1 Export - simpler logic
            config = db.session.get(ReportConfig, rid)
            if not config or not config.file_blob:
                return "Config not found", 400
            
            try:
                wb = openpyxl.load_workbook(io.BytesIO(config.file_blob), data_only=False)
            except:
                wb = openpyxl.load_workbook(io.BytesIO(config.file_blob))
            
            ws = wb.active
            
            # Add V1 data similarly...
        
        # Save and send
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"ThongKe_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        app.logger.error(f"Stats export error: {e}")
        return str(e), 500
