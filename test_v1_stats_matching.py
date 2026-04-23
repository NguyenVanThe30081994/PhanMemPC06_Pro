#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test V1 Stats - Kiểm tra render data rows với unit matching
"""

import io
import openpyxl
from excel_renderer import build_stats_table_html

# Mock ReportConfig
class MockReportConfig:
    def __init__(self):
        self.header_start = 1
        self.header_rows = 2
        self.config_json = '[{"idx": 2, "is_visible": true}, {"idx": 3, "is_visible": true}]'
        self.is_daily = False
        self.file_blob = None

# Tạo Excel file test với unit names
def create_test_excel_with_units():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    
    # Header rows
    ws['A1'] = "Đơn vị"
    ws['B1'] = "Số lượng"
    ws['C1'] = "Tỷ lệ"
    
    ws['A2'] = "Unit"
    ws['B2'] = "Qty"
    ws['C2'] = "Rate"
    
    # Data rows - với unit names
    ws['A3'] = "Đơn vị A"  # ← Unit name
    ws['B3'] = 0
    ws['C3'] = 0
    
    ws['A4'] = "Đơn vị B"  # ← Unit name
    ws['B4'] = 0
    ws['C4'] = 0
    
    # Format
    ws['B3'].number_format = '0'
    ws['C3'].number_format = '0.00'
    ws['B4'].number_format = '0'
    ws['C4'].number_format = '0.00'
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Test
print("=" * 80)
print("TEST: V1 Stats - Unit matching and data rendering")
print("=" * 80)

config = MockReportConfig()
config.file_blob = create_test_excel_with_units()

# Test Case 1: Submissions WITH values and matching units
print("\n✓ Test Case 1: Submissions WITH values and matching units")
submissions = [
    {
        'unit': 'Đơn vị A',
        'sender': 'Người A',
        'date': '23/04/2026',
        'values': {'2': '491', '3': '0.5'}
    },
    {
        'unit': 'Đơn vị B',
        'sender': 'Người B',
        'date': '23/04/2026',
        'values': {'2': '543', '3': '0.75'}
    }
]

try:
    html = build_stats_table_html(config.file_blob, config, submissions)
    
    # Check if data is rendered
    has_491 = '491' in html
    has_543 = '543' in html
    has_unit_a = 'Đơn vị A' in html
    has_unit_b = 'Đơn vị B' in html
    
    print(f"  Found '491': {has_491}")
    print(f"  Found '543': {has_543}")
    print(f"  Found 'Đơn vị A': {has_unit_a}")
    print(f"  Found 'Đơn vị B': {has_unit_b}")
    
    if has_491 and has_543:
        print("  ✅ PASS: Data rows rendered with values")
    else:
        print("  ❌ FAIL: Data not found in HTML")
        # Debug: print first 500 chars of HTML
        print(f"\n  HTML preview (first 500 chars):")
        print(f"  {html[:500]}")
        
except Exception as e:
    print(f"  ❌ FAIL: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 80)
print("Test completed")
print("=" * 80)
