# -*- coding: utf-8 -*-
from flask import Blueprint, render_template, request, session, redirect, url_for, flash, jsonify, g
from models import db, ReportTemplateV2, ReportVersionV2, ReportSubmissionV2, ReportValueV2, ReportAuditV2, User
from pc06_excel_engine import ExcelEngineV2
from utils import render_auto_template as _render_template
import json
import io
import os
from datetime import datetime

reports_v2_bp = Blueprint('reports_v2_bp', __name__)


@reports_v2_bp.route('/reports-v2/api/file/<int:vid>')
def get_version_file(vid):
    """API tra ve file Excel thu cho LuckyExcel"""
    if not session.get('uid'):
        return jsonify({"error": "Unauthorized"}), 401
    
    version = db.session.get(ReportVersionV2, vid)
    if not version or not version.excel_file_blob:
        return jsonify({"error": "Not found"}), 404
    
    from flask import send_file
    return send_file(
        io.BytesIO(version.excel_file_blob),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='template.xlsx'
    )


GLOBAL_UNITS = ['H\u1ec7 th\u1ed1ng', 'Admin', 'PC06']


def _is_global_user(is_admin, user_unit):
    """Check if user can access all units (admin or special global unit)"""
    return bool(is_admin) or (user_unit in GLOBAL_UNITS)


def _format_cell_value(val, number_format=None):
    """
    Format value cho hiển thị, ưu tiên nhận diện định dạng % và số nguyên của Excel
    """
    if val is None:
        return ''
        
    if isinstance(val, (int, float)):
        if number_format:
            fmt = str(number_format).lower()
            if '%' in fmt:
                decimals = 2 if '.0' in fmt else 0
                return f"{val * 100:.{decimals}f}%".replace('.', ',')
            if '0.0' in fmt:
                return f"{val:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            if '#' in fmt or '0' in fmt:
                return f"{int(round(val)):,}".replace(',', '.')
                
        if val == int(val):
            return str(int(val))
        
        rounded = round(val, 4)
        if rounded == int(rounded):
            return str(int(rounded))
        return str(rounded).rstrip('0').rstrip('.').replace('.', ',')
        
    return str(val)


def _get_cell_format(meta_data, sheet_name, coord):
    if not meta_data or not meta_data.get('sheets'):
        return None
    
    for sheet in meta_data['sheets']:
        if sheet.get('name') != sheet_name:
            continue
        for row in sheet.get('rows', []):
            for cell in row.get('cells', []):
                if cell.get('coord') == coord:
                    return cell.get('numberFormat')
    return None


def _normalize_v2_key(sheet_name, coord):
    return f"{sheet_name}!{coord}"


def _split_v2_key(raw_key):
    key = str(raw_key or '').strip()
    if '!' in key:
        sheet, coord = key.split('!', 1)
        return sheet.strip(), coord.strip().upper()
    return None, key.upper()


def _get_column_config(meta_data, coord):
    """
    Get per-column config from metadata['column_configs'] using Excel column letter.
    """
    try:
        from openpyxl.utils.cell import coordinate_from_string
        col_letter, _ = coordinate_from_string(coord)
    except Exception:
        col_letter = ''.join(ch for ch in str(coord or '') if ch.isalpha()).upper()
    return (meta_data or {}).get('column_configs', {}).get(col_letter, {}) or {}


def _get_effective_number_format(cell, meta_data, sheet_name, coord):
    """
    Priority:
    1. Real Excel cell.number_format
    2. Parsed metadata numberFormat
    3. Derived fallback from column_configs
    """
    fmt = None
    try:
        fmt = getattr(cell, 'number_format', None)
    except Exception:
        fmt = None

    if fmt and str(fmt).strip() and str(fmt).strip().lower() != 'general':
        return fmt

    fmt = _get_cell_format(meta_data, sheet_name, coord)
    if fmt:
        return fmt

    col_cfg = _get_column_config(meta_data, coord)
    if col_cfg.get('is_percent'):
        return '0.00%'
    if col_cfg.get('is_numeric'):
        return '#,##0.##'

    return None


def _coerce_numeric_for_display(val):
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return val
    try:
        raw = str(val).strip()
        if raw == '':
            return None
        raw = raw.replace(',', '')
        return float(raw)
    except Exception:
        return None


def _format_by_column_config(val, col_cfg):
    """
    Fallback formatter when Excel number format is missing.
    """
    if val is None or val == '':
        return ''

    num = _coerce_numeric_for_display(val)
    if num is None:
        return str(val)

    if col_cfg.get('is_percent'):
        pct = num * 100
        txt = f"{pct:.2f}".rstrip('0').rstrip('.')
        return f"{txt}%"

    if col_cfg.get('is_numeric'):
        try:
            if float(num).is_integer():
                return str(int(num))
            return f"{float(num):.2f}".rstrip('0').rstrip('.')
        except:
            return str(val)

    return str(val)


def _get_display_value(formula_cell, value_cell):
    """
    Return the value users should see on screen.
    - For formula cells, prefer cached value from data_only workbook.
    """
    formula_value = getattr(formula_cell, 'value', None) if formula_cell is not None else None
    value = getattr(value_cell, 'value', None) if value_cell is not None else None

    if isinstance(formula_value, str) and formula_value.startswith('='):
        return value

    if formula_value is not None:
        return formula_value

    return value


def _get_rendered_cell_text(formula_cell, value_cell, meta_data, sheet_name, coord, explicit_value=None):
    """
    Convert a cell value to final HTML display text.
    """
    raw = explicit_value if explicit_value is not None else _get_display_value(formula_cell, value_cell)
    if raw is None or raw == '':
        return ''

    fmt = _get_effective_number_format(formula_cell, meta_data, sheet_name, coord)
    col_cfg = _get_column_config(meta_data, coord)

    if fmt:
        try:
            return _format_cell_value(raw, fmt)
        except Exception:
            pass

    return _format_by_column_config(raw, col_cfg)


def _get_core_unit_id(name):
    """
    Trích xuất 'lõi' định danh duy nhất của đơn vị, loại bỏ mọi tiền tố hành chính và khoảng trắng.
    """
    if not name:
        return ""
    import unicodedata
    n = str(name).lower().strip()
    n = unicodedata.normalize('NFKD', n).encode('ascii', 'ignore').decode('utf-8')
    
    for char in ['-', '_', '.', ',', '/']:
        n = n.replace(char, ' ')
        
    if ' ' not in n:
        solid_prefixes = ['ubndthitran', 'ubndphuong', 'ubndxa', 'cathitran', 'caphuong', 
                          'congphuong', 'congxa', 'caxa', 'ubnd', 'ca', 'thitran', 'phuong', 'xa', 'tt']
        for p in solid_prefixes:
            if n.startswith(p):
                n = n[len(p):]
                break
    else:
        prefixes = [
            "uy ban nhan dan ", "cong an ", "thi tran ", "thanh pho ", "ubnd ", 
            "ca ", "xa ", "phuong ", "huyen ", "tt "
        ]
        changed = True
        while changed:
            changed = False
            for p in prefixes:
                if n.startswith(p):
                    n = n[len(p):].strip()
                    changed = True
                    
    n = n.replace(" ", "").strip()
    
    # --- BẢO VỆ TỪ KHÓA HEADER ---
    # Ngăn thuật toán nhận diện nhầm các từ ngữ trong Header thành tên đơn vị
    blocklist = [
        'stt', 'tendonvi', 'tendonvihanhchinh', 'hanhchinh', 'chitieu', 
        'ketqua', 'tyle', 'hoanthanh', 'tongso', 'toantinh', 'toanhuyen', 
        'baocao', 'danhsach', 'capmoi', 'capdoi', 'xaydung', 'csdl', 
        'ghichu', 'tong', 'cong', 'luoidiachinh', 'tongcong', 'ngay', 'thang', 'nam'
    ]
    if n in blocklist or len(n) < 2:
        return ""
        
    return n


def _find_smart_data_start_row(ws, min_row, max_row, min_col, max_col):
    """
    Thuật toán Auto-Detect ranh giới Header (dựa trên STT và Đánh số cột).
    Trả về dòng đầu tiên chứa dữ liệu (nằm ngay dưới Header).
    """
    scan_limit = min(max_row, min_row + 20)
    
    # Ưu tiên 1: Tìm dòng đánh số thứ tự ngang
    for r in range(min_row, scan_limit):
        num_count = 0
        for c in range(min_col, min(max_col + 1, min_col + 15)):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                try:
                    v = float(str(val).strip())
                    if v.is_integer() and 1 <= v <= 20:
                        num_count += 1
                except ValueError:
                    pass
        if num_count >= 3:
            return r + 1
            
    # Ưu tiên 2: Tìm dòng chứa chữ "STT" hoặc "Tên đơn vị"
    for r in range(min_row, scan_limit):
        for c in range(min_col, min(max_col + 1, min_col + 3)):
            val = ws.cell(row=r, column=c).value
            if val:
                txt = str(val).lower().replace(' ', '')
                if 'stt' in txt or 'tendonvi' in txt or 'hanhchinh' in txt:
                    return r + 1
                    
    return None


def _find_unit_rows_and_col(ws, min_row, max_row, min_col, max_col, user_identifiers):
    user_core_ids = set()
    for uid in user_identifiers:
        core_id = _get_core_unit_id(uid)
        if core_id and len(core_id) >= 2 and core_id not in ['pc06', 'admin', 'hethong']:
            user_core_ids.add(core_id)
            
    matched_rows = []
    matched_col = None
    
    if not user_core_ids:
        return matched_rows, matched_col

    for r in range(min_row, max_row + 1):
        for c in range(min_col, min(max_col + 1, min_col + 3)):
            cell_val = ws.cell(row=r, column=c).value
            if not cell_val:
                continue
                
            cell_core_id = _get_core_unit_id(str(cell_val))
            if not cell_core_id:
                continue
                
            if cell_core_id in user_core_ids:
                matched_rows.append(r)
                if matched_col is None:
                    matched_col = c
                break 
                
    return matched_rows, matched_col


def _is_editable_by_row_context(cell, col_idx, unit_col):
    if unit_col is not None and col_idx <= unit_col:
        return False
    if isinstance(cell.value, str):
        raw = cell.value.strip()
        if raw == '':
            return True
        if raw.startswith('='):
            return False
        return False
    if cell.value is None:
        return True
    if isinstance(cell.value, (int, float)):
        return True
    return False


def _get_sheet_region(meta_data, ws, wb):
    sheet_meta = next((s for s in meta_data.get('sheets', []) if s.get('name') == ws.title), None)
    if sheet_meta:
        region = sheet_meta.get('activeRenderRegion', {})
        min_row = region.get('r1', 1)
        min_col = region.get('c1', 1)
        max_row = region.get('r2', ws.max_row)
        max_col = region.get('c2', ws.max_column)
    else:
        max_row, max_col = ExcelEngineV2._get_true_max_row_col(wb, ws)
        min_row, min_col = 1, 1
    return min_row, min_col, max_row, max_col


def _collect_allowed_input_keys(wb, meta_data, user_identifiers, is_admin):
    from excel_renderer import is_input_cell

    is_global = _is_global_user(is_admin, session.get('unit', ''))
    allowed_keys = set()

    for ws in wb.worksheets:
        min_row, min_col, max_row, max_col = _get_sheet_region(meta_data, ws, wb)
        
        # Áp dụng thuật toán nhận diện ranh giới vào cả bảo mật input
        smart_start_row = _find_smart_data_start_row(ws, min_row, max_row, min_col, max_col)
        actual_data_start = smart_start_row if smart_start_row else meta_data.get('data_start_row', 4)
        
        unit_rows, unit_col = _find_unit_rows_and_col(ws, actual_data_start, max_row, min_col, max_col, user_identifiers)

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                coord = cell.coordinate
                key = _normalize_v2_key(ws.title, coord)

                if is_global:
                    if is_input_cell(cell):
                        allowed_keys.add(key)
                    continue

                if r in unit_rows and _is_editable_by_row_context(cell, c, unit_col):
                    allowed_keys.add(key)

    return allowed_keys


@reports_v2_bp.route('/reports-v2')
def dashboard():
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))
    
    templates = ReportTemplateV2.query.order_by(ReportTemplateV2.created_at.desc()).all()
    is_admin = session.get('is_admin', False)
    
    if request.args.get('mobile') == '1':
        return render_template('reports_v2_dashboard_mobile.html', templates=templates, is_admin=is_admin)
    
    return _render_template('reports_v2_dashboard.html', templates=templates, is_admin=is_admin)


@reports_v2_bp.route('/reports-v2/upload', methods=['POST'])
def upload_template():
    if not session.get('is_admin'):
        return jsonify({"error": "Unauthorized"}), 403

    file = request.files.get('template_excel')
    name = request.form.get('name', 'Báo cáo Mới')
    is_daily = request.form.get('is_daily') == 'true' or 'is_daily' in request.form

    if not file or not file.filename:
        return jsonify({"error": "No file uploaded"}), 400

    try:
        file_content = file.read()
        
        import uuid
        temp_path = os.path.join("tmp", f"{uuid.uuid4().hex}.xlsx")
        os.makedirs("tmp", exist_ok=True)
        with open(temp_path, "wb") as f:
            f.write(file_content)

        metadata = ExcelEngineV2.parse_template(temp_path)

        template = ReportTemplateV2.query.filter_by(name=name).first()
        is_new = False
        if not template:
            template = ReportTemplateV2(
                name=name,
                created_by=session.get('fullname'),
                is_daily=is_daily
            )
            db.session.add(template)
            db.session.flush()
            is_new = True
        else:
            template.is_daily = is_daily

        new_version = ReportVersionV2(
            template_id=template.id,
            version_tag=datetime.now().strftime("%Y%m%d%H%M"),
            metadata_json=json.dumps(metadata, ensure_ascii=False),
            excel_file_blob=file_content,
            is_published=True
        )
        db.session.add(new_version)

        ExcelEngineV2.save_logic_to_source(name, metadata)

        db.session.commit()

        if is_new:
            from utils import push_global_notif
            push_global_notif("Biểu mẫu mới", f"Vừa có biểu mẫu mới: {name}", f"/reports-v2/render/{template.id}", exclude_uid=session['uid'])

        if os.path.exists(temp_path):
            os.remove(temp_path)

        return jsonify({"success": True, "template_id": template.id, "version_id": new_version.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@reports_v2_bp.route('/reports-v2/edit/<int:tid>', methods=['GET', 'POST'])
def edit_template(tid):
    if not session.get('is_admin'):
        return redirect(url_for('auth_bp.login'))

    template = db.session.get(ReportTemplateV2, tid)
    if not template:
        flash('Không tìm thấy biểu mẫu V2!', 'danger')
        return redirect(url_for('reports_v2_bp.dashboard'))

    if request.method == 'POST':
        template.name = request.form.get('name', template.name)
        template.description = request.form.get('description', template.description)
        template.is_daily = 'is_daily' in request.form

        file = request.files.get('template_excel')
        if file and file.filename:
            try:
                file_content = file.read()
                
                import uuid
                temp_path = os.path.join("tmp", f"{uuid.uuid4().hex}.xlsx")
                os.makedirs("tmp", exist_ok=True)
                with open(temp_path, "wb") as f:
                    f.write(file_content)

                metadata = ExcelEngineV2.parse_template(temp_path)
                ReportVersionV2.query.filter_by(template_id=tid, is_published=True).update({'is_published': False})

                new_version = ReportVersionV2(
                    template_id=tid,
                    version_tag=datetime.now().strftime("%Y%m%d%H%M"),
                    metadata_json=json.dumps(metadata, ensure_ascii=False),
                    excel_file_blob=file_content,
                    is_published=True
                )
                db.session.add(new_version)
                ExcelEngineV2.save_logic_to_source(template.name, metadata)

                if os.path.exists(temp_path):
                    os.remove(temp_path)

                flash('Đã cập nhật file Excel và tạo phiên bản mới!', 'info')
            except Exception as e:
                db.session.rollback()
                flash(f'Lỗi xử lý file: {str(e)}', 'danger')
                return redirect(url_for('reports_v2_bp.edit_template', tid=tid))

        db.session.commit()
        flash('Đã cập nhật biểu mẫu V2!', 'success')
        return redirect(url_for('reports_v2_bp.dashboard'))

    versions = ReportVersionV2.query.filter_by(template_id=tid).order_by(ReportVersionV2.created_at.desc()).all()
    
    user_agent = request.headers.get('User-Agent', '').lower()
    is_mobile_request = 'mobile' in user_agent or 'android' in user_agent or 'iphone' in user_agent
    
    if is_mobile_request:
        flash('Vui lòng sử dụng máy tính để chỉnh sửa biểu mẫu V2.', 'warning')
        return redirect(url_for('reports_v2_bp.dashboard'))
    
    return _render_template('reports_v2_edit.html', template=template, versions=versions)


@reports_v2_bp.route('/reports-v2/config/<int:tid>')
def config_template(tid):
    if not session.get('is_admin'):
        return redirect(url_for('auth_bp.login'))
    
    template = db.session.get(ReportTemplateV2, tid)
    if not template:
        flash('Không tìm thấy biểu mẫu!', 'danger')
        return redirect(url_for('reports_v2_bp.dashboard'))
    
    version = ReportVersionV2.query.filter_by(template_id=tid, is_published=True).order_by(ReportVersionV2.created_at.desc()).first()
    if not version or not version.excel_file_blob:
        flash('Chưa có file Excel! Vui lòng upload trước.', 'warning')
        return redirect(url_for('reports_v2_bp.edit_template', tid=tid))
    
    from pc06_excel_scanner import scan_excel_structure
    detected = scan_excel_structure(version.excel_file_blob)
    
    try:
        existing_config = json.loads(version.metadata_json or '{}')
    except:
        existing_config = {}
    
    column_configs = existing_config.get('column_configs', {})
    header_groups = existing_config.get('header_groups', [])
    data_start_row = existing_config.get('data_start_row', detected.get('data_start_row', 4))
    unit_column = existing_config.get('unit_column', 'B')
    header_row = existing_config.get('header_row', 3)
    header_column = existing_config.get('header_column', 'A')
    
    return _render_template('reports_v2_config.html',
                          template=template,
                          detected=detected,
                          column_configs=column_configs,
                          header_groups=header_groups,
                          data_start_row=data_start_row,
                          unit_column=unit_column,
                          header_row=header_row,
                          header_column=header_column)


@reports_v2_bp.route('/reports-v2/config/<int:tid>', methods=['POST'])
def save_config(tid):
    if not session.get('is_admin'):
        return redirect(url_for('auth_bp.login'))
    
    template = db.session.get(ReportTemplateV2, tid)
    if not template:
        flash('Không tìm thấy biểu mẫu!', 'danger')
        return redirect(url_for('reports_v2_bp.dashboard'))
    
    version = ReportVersionV2.query.filter_by(template_id=tid, is_published=True).order_by(ReportVersionV2.created_at.desc()).first()
    if not version:
        flash('Không có phiên bản!', 'danger')
        return redirect(url_for('reports_v2_bp.dashboard'))
    
    try:
        try:
            metadata = json.loads(version.metadata_json or '{}')
        except:
            metadata = {}
        
        column_configs = {}
        all_columns = request.form.getlist('columns') if request.form.get('columns') else []
        
        if not all_columns:
            for key in request.form.keys():
                if key.startswith('is_numeric_') or key.startswith('is_percent_') or key.startswith('is_formula_') or key.startswith('is_sortable_'):
                    col = key.split('_', 2)[-1]
                    if col not in column_configs:
                        column_configs[col] = {}
        
        for col in column_configs:
            column_configs[col] = {
                'is_numeric': f'is_numeric_{col}' in request.form,
                'is_percent': f'is_percent_{col}' in request.form,
                'is_formula': f'is_formula_{col}' in request.form,
                'is_sortable': f'is_sortable_{col}' in request.form,
            }
        
        metadata['column_configs'] = column_configs
        metadata['data_start_row'] = int(request.form.get('data_start_row', 4))
        metadata['unit_column'] = request.form.get('unit_column', 'B')
        metadata['header_row'] = int(request.form.get('header_row', 3))
        metadata['header_column'] = request.form.get('header_column', 'A')
        metadata['config_version'] = '2.0'
        
        version.metadata_json = json.dumps(metadata, ensure_ascii=False)
        db.session.commit()
        
        flash('Đã lưu cấu hình V2!', 'success')
        return redirect(url_for('reports_v2_bp.dashboard'))
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi: {str(e)}', 'danger')
        return redirect(url_for('reports_v2_bp.config_template', tid=tid))


@reports_v2_bp.route('/reports-v2/delete/<int:tid>', methods=['POST'])
def delete_template(tid):
    if not session.get('is_admin'):
        return jsonify({"error": "Unauthorized"}), 403

    template = db.session.get(ReportTemplateV2, tid)
    if not template:
        return jsonify({"error": "Not found"}), 404

    try:
        versions = ReportVersionV2.query.filter_by(template_id=tid).all()
        for v in versions:
            for sub in ReportSubmissionV2.query.filter_by(version_id=v.id).all():
                ReportValueV2.query.filter_by(submission_id=sub.id).delete()
                ReportAuditV2.query.filter_by(submission_id=sub.id).delete()
                db.session.delete(sub)
            db.session.delete(v)
        db.session.delete(template)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@reports_v2_bp.route('/reports-v2/render/<int:tid>')
def render_report(tid):
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))

    template = db.session.get(ReportTemplateV2, tid)
    if not template:
        return "Template Not Found", 404

    version = ReportVersionV2.query.filter_by(template_id=tid, is_published=True).order_by(ReportVersionV2.created_at.desc()).first()
    if not version:
        return "No published version found", 404

    if not version.excel_file_blob:
        return "No Excel file stored for this version", 400

    submission = ReportSubmissionV2.query.filter_by(version_id=version.id, user_id=session['uid'], status='draft').first()
    existing_values = {}
    if submission:
        for val in submission.values:
            existing_values[val.cell_key] = val.value

    is_admin = session.get('is_admin', False)
    user_unit = session.get('unit_area', session.get('unit', ''))
    is_global = _is_global_user(is_admin, user_unit)
    
    user_identifiers = [
        session.get('username', ''),
        session.get('fullname', ''),
        session.get('unit_area', ''),
        session.get('unit', '')
    ]
    user_identifiers = [str(u) for u in user_identifiers if u]

    from excel_renderer import _build_merge_lookup, _col_widths_px, _row_height_px, _cell_css, is_input_cell
    import openpyxl as _opx

    try:
        # Load workbook 2 lần: 
        # wb_formula: để lấy công thức, style, layout (data_only=False)
        # wb_values: để lấy giá trị cached thực tế của công thức (data_only=True)
        try:
            wb_formula = _opx.load_workbook(io.BytesIO(version.excel_file_blob), data_only=False)
            wb_values = _opx.load_workbook(io.BytesIO(version.excel_file_blob), data_only=True)
        except:
            wb_formula = _opx.load_workbook(io.BytesIO(version.excel_file_blob))
            wb_values = wb_formula
    except Exception as e:
        return f"Error loading Excel: {e}", 500

    wb = wb_formula # Dùng wb_formula làm workbook chính cho style/layout

    meta_data = {}
    try:
        meta_data = json.loads(version.metadata_json or '{}')
    except Exception:
        meta_data = {}

    sheets_html = []
    for ws in wb.worksheets:
        ws_values = wb_values[ws.title] if ws.title in wb_values.sheetnames else ws
        spans, shadows = _build_merge_lookup(ws)
        col_widths = _col_widths_px(ws)
        colgroup = '<colgroup>' + ''.join(f'<col style="width:{w}px">' for w in col_widths) + '</colgroup>'

        min_row, min_col, max_row, max_col = _get_sheet_region(meta_data, ws, wb)
        
        # BẢN VÁ: Tìm data_start_row bằng thuật toán nhận diện STT
        smart_start_row = _find_smart_data_start_row(ws, min_row, max_row, min_col, max_col)
        actual_data_start = smart_start_row if smart_start_row else meta_data.get('data_start_row', 4)

        unit_rows, unit_col = _find_unit_rows_and_col(ws, actual_data_start, max_row, min_col, max_col, user_identifiers)
        
        all_unit_keys = set()
        for r in range(actual_data_start, max_row + 1):
            for c in range(min_col, min(max_col + 1, min_col + 3)):
                cell_val = ws.cell(row=r, column=c).value
                if cell_val:
                    cell_core_id = _get_core_unit_id(str(cell_val))
                    if cell_core_id:
                        all_unit_keys.add((r, cell_core_id))
                        break
        
        first_unit_row = min([r for r, k in all_unit_keys]) if all_unit_keys else None
        user_rows_set = set(unit_rows)
        other_unit_rows = {r for r, k in all_unit_keys if r not in user_rows_set}
        
        header_row_cfg = meta_data.get('header_row', 3)
        should_filter = not is_global and unit_rows and header_row_cfg

        user_data_end = max_row
        if unit_rows:
            user_first = min(unit_rows)
            for r, key in sorted(all_unit_keys):
                if r > user_first and r not in unit_rows:
                    user_data_end = r - 1
                    break

        rows_html = []
        for r in range(min_row, max_row + 1):
            if ws.row_dimensions[r].hidden:
                continue

            if not should_filter:
                pass
            elif r < actual_data_start:  # BẢO VỆ HEADER
                pass
            elif r in unit_rows:
                pass
            elif r in other_unit_rows:
                continue
            elif r > user_first and r <= user_data_end:
                pass
            else:
                continue

            rh = _row_height_px(ws, r)
            row_parts = [f'<tr style="height:{rh}px">']
            for c in range(min_col, max_col + 1):
                if (r, c) in shadows:
                    continue

                cell = ws.cell(row=r, column=c)
                rowspan, colspan = spans.get((r, c), (1, 1))
                css = _cell_css(cell)

                if is_global:
                    is_input = is_input_cell(cell)
                else:
                    is_input = (r in unit_rows) and _is_editable_by_row_context(cell, c, unit_col)

                coord = cell.coordinate
                key = _normalize_v2_key(ws.title, coord)
                rs_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ''
                cs_attr = f' colspan="{colspan}"' if colspan > 1 else ''
                base = 'padding:3px 6px;border:1px solid #d1d5db;overflow:hidden;box-sizing:border-box;'
                if is_input:
                    base += 'background-color:#e0f2fe;'
                full_css = base + css
                td = f'<td{rs_attr}{cs_attr} style="{full_css}">'

                if is_input:
                    raw_val = existing_values.get(key, existing_values.get(coord, ''))
                    value_cell = ws_values.cell(row=r, column=c)
                    if raw_val in (None, ''):
                        # Nếu chưa có dữ liệu nhập, hiển thị giá trị mặc định từ template
                        val = _get_rendered_cell_text(cell, value_cell, meta_data, ws.title, coord)
                    else:
                        # Nếu đã có dữ liệu nhập, format theo cấu hình
                        fmt = _get_effective_number_format(cell, meta_data, ws.title, coord)
                        col_cfg = _get_column_config(meta_data, coord)
                        if fmt:
                            try:
                                val = _format_cell_value(raw_val, fmt)
                            except Exception:
                                val = _format_by_column_config(raw_val, col_cfg)
                        else:
                            val = _format_by_column_config(raw_val, col_cfg)

                    safe_val = str(val).replace('"', '&quot;')
                    inner = (
                        f'<input type="text" class="grid-input" '
                        f'data-key="{key}" data-coord="{coord}" '
                        f'value="{safe_val}" onchange="markDirty()" '
                        f'style="width:100%;height:100%;border:none;background:transparent;padding:2px;font-size:inherit;">'
                    )
                else:
                    value_cell = ws_values.cell(row=r, column=c)
                    inner = _get_rendered_cell_text(cell, value_cell, meta_data, ws.title, coord)

                row_parts.append(f'{td}{inner}</td>')
            row_parts.append('</tr>')
            rows_html.append(''.join(row_parts))

        table_html = (
            f'<table class="excel-render-table" '
            f'style="border-collapse:collapse;font-size:12px;width:100%;table-layout:fixed;'
            f'font-family:Calibri,Arial,sans-serif;min-width:1000px;">'
            f'{colgroup}<tbody>{"".join(rows_html)}</tbody></table>'
        )

        from markupsafe import Markup
        sheets_html.append({'name': ws.title, 'html': Markup(table_html)})

    from models import ReportConfig
    v2_templates = ReportTemplateV2.query.all()
    v1_configs = ReportConfig.query.all()
    return _render_template(
        'reports_v2_render.html',
        template=template,
        version=version,
        sheets_html=sheets_html,
        is_admin=is_admin,
        user_unit=user_unit,
        v2_templates=v2_templates,
        v1_configs=v1_configs,
        form_type='v2'
    )


@reports_v2_bp.route('/reports-v2/input-lucky/<int:tid>')
def input_lucky(tid):
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))
    
    template = db.session.get(ReportTemplateV2, tid)
    if not template:
        return "Template Not Found", 404
    
    version = ReportVersionV2.query.filter_by(template_id=tid, is_published=True).order_by(ReportVersionV2.created_at.desc()).first()
    if not version:
        return "No published version found", 404
    
    if not version.excel_file_blob:
        return "No Excel file stored", 400
    
    config_json = '{}'
    try:
        meta = json.loads(version.metadata_json or '{}')
        config_json = json.dumps({
            'unit_column': meta.get('unit_column', 'B'),
            'input_range': meta.get('input_range', [])
        })
    except:
        config_json = json.dumps({'unit_column': 'B', 'input_range': []})
    
    user_unit = session.get('unit_area', session.get('unit', ''))
    
    return render_template('reports_v2_input_lucky.html',
        template=template,
        version=version,
        user_unit=user_unit,
        config_json=config_json
    )


@reports_v2_bp.route('/reports-v2/submit', methods=['POST'])
def submit_data():
    if not session.get('uid'):
        return jsonify({"error": "Unauthorized"}), 403

    data = request.json or {}
    version_id = data.get('version_id')
    values = data.get('values', {})

    if not version_id:
        return jsonify({'success': False, 'message': 'Thiếu thông tin phiên bản mẫu.'}), 400

    version = db.session.get(ReportVersionV2, version_id)
    if not version or not version.excel_file_blob:
        return jsonify({'success': False, 'message': 'Phiên bản mẫu không hợp lệ.'}), 404

    try:
        import openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(version.excel_file_blob), rich_text=True, data_only=True)
        except:
            wb = openpyxl.load_workbook(io.BytesIO(version.excel_file_blob), data_only=True)

        meta_data = {}
        try:
            meta_data = json.loads(version.metadata_json or '{}')
        except Exception:
            meta_data = {}

        is_admin = session.get('is_admin', False)
        user_identifiers = [
            session.get('username', ''),
            session.get('fullname', ''),
            session.get('unit_area', ''),
            session.get('unit', '')
        ]
        user_identifiers = [str(u) for u in user_identifiers if u]

        allowed_prefixed_keys = _collect_allowed_input_keys(wb, meta_data, user_identifiers, is_admin)

        normalized_payload = {}
        for raw_key, val in values.items():
            sheet_name, coord = _split_v2_key(raw_key)
            resolved_key = None

            if sheet_name:
                candidate = _normalize_v2_key(sheet_name, coord)
                if candidate in allowed_prefixed_keys:
                    resolved_key = candidate
            else:
                matched = [k for k in allowed_prefixed_keys if k.endswith(f'!{coord}')]
                if len(matched) == 1:
                    resolved_key = matched[0]

            if not resolved_key:
                return jsonify({'success': False, 'message': f'Lỗi bảo mật: Ô {raw_key} không thuộc quyền quản lý của đơn vị bạn!'}), 403

            normalized_payload[resolved_key] = val

        submission = ReportSubmissionV2.query.filter_by(version_id=version_id, user_id=session['uid'], status='draft').first()

        if not submission:
            submission = ReportSubmissionV2(
                version_id=version_id,
                user_id=session['uid'],
                org_unit=session.get('unit_area', session.get('unit', 'PC06'))
            )
            db.session.add(submission)
            db.session.flush()

        existing_rows = ReportValueV2.query.filter_by(submission_id=submission.id).all()
        existing_map = {row.cell_key: row for row in existing_rows}

        for key, val in normalized_payload.items():
            existing_val = existing_map.get(key)
            old_val = existing_val.value if existing_val else None

            if str(old_val) != str(val):
                db.session.add(ReportAuditV2(
                    submission_id=submission.id,
                    user_id=session['uid'],
                    cell_key=key,
                    old_value=old_val,
                    new_value=val
                ))

                if existing_val:
                    existing_val.value = val
                else:
                    db.session.add(ReportValueV2(submission_id=submission.id, cell_key=key, value=val))

        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@reports_v2_bp.route('/reports-v2/export/<int:sid>')
def export_submission(sid):
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))

    submission = db.session.get(ReportSubmissionV2, sid)
    if not submission:
        return "Submission Not Found", 404

    version = submission.version
    if not version.excel_file_blob:
        return "No original template file found", 400

    try:
        import openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(version.excel_file_blob), rich_text=True)
        except:
            wb = openpyxl.load_workbook(io.BytesIO(version.excel_file_blob))

        ws_by_title = {ws.title: ws for ws in wb.worksheets}
        for val in submission.values:
            sheet_name, coord = _split_v2_key(val.cell_key)
            if sheet_name and sheet_name in ws_by_title:
                try:
                    ws_by_title[sheet_name][coord].value = val.value
                except Exception:
                    pass
            else:
                for ws in wb.worksheets:
                    try:
                        ws[coord].value = val.value
                    except Exception:
                        pass

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Report_{submission.org_unit}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        from flask import send_file
        return send_file(output, as_attachment=True, download_name=filename)
    except Exception as e:
        return str(e), 500


@reports_v2_bp.route('/reports-v2/submission/<int:sub_id>')
def review_submission(sub_id):
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))

    submission = db.session.get(ReportSubmissionV2, sub_id)
    if not submission:
        flash('Không tìm thấy bản nộp!', 'danger')
        return redirect(url_for('forms_bp.stats'))

    version = submission.version
    template = version.template

    existing_values = {val.cell_key: val.value for val in submission.values}

    from excel_renderer import _build_merge_lookup, _row_height_px, _cell_css, is_input_cell
    import openpyxl as _opx

    try:
        try:
            wb_formula = _opx.load_workbook(io.BytesIO(version.excel_file_blob), data_only=False)
            wb_values = _opx.load_workbook(io.BytesIO(version.excel_file_blob), data_only=True)
        except:
            wb_formula = _opx.load_workbook(io.BytesIO(version.excel_file_blob))
            wb_values = wb_formula
    except Exception as e:
        return f"Error loading Excel: {e}", 500

    wb = wb_formula

    sheets_html = []
    meta_data = {}
    try:
        meta_data = json.loads(version.metadata_json or '{}')
    except Exception:
        meta_data = {}

    for ws in wb.worksheets:
        min_row, min_col, max_row, max_col = _get_sheet_region(meta_data, ws, wb)
        spans, shadows = _build_merge_lookup(ws)

        col_widths = []
        for i in range(min_col, max_col + 1):
            letter = _opx.utils.get_column_letter(i)
            w = ws.column_dimensions[letter].width or 8.43
            col_widths.append(max(int(w * 7), 45))

        colgroup = '<colgroup>' + ''.join(f'<col style="width:{w}px">' for w in col_widths) + '</colgroup>'
        rows_html = []

        for r in range(min_row, max_row + 1):
            if ws.row_dimensions[r].hidden:
                continue
            rh = _row_height_px(ws, r)
            row_parts = [f'<tr style="height:{rh}px">']

            for c in range(min_col, max_col + 1):
                if (r, c) in shadows:
                    continue

                cell = ws.cell(row=r, column=c)
                rowspan, colspan = spans.get((r, c), (1, 1))
                css = _cell_css(cell)

                coord = cell.coordinate
                key = _normalize_v2_key(ws.title, coord)
                saved_val = existing_values.get(key, existing_values.get(coord, None))
                value_cell = ws_values.cell(row=r, column=c)

                if saved_val not in (None, ''):
                    fmt = _get_effective_number_format(cell, meta_data, ws.title, coord)
                    col_cfg = _get_column_config(meta_data, coord)
                    if fmt:
                        try:
                            val = _format_cell_value(saved_val, fmt)
                        except Exception:
                            val = _format_by_column_config(saved_val, col_cfg)
                    else:
                        val = _format_by_column_config(saved_val, col_cfg)
                else:
                    val = _get_rendered_cell_text(cell, value_cell, meta_data, ws.title, coord)

                rs_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ''
                cs_attr = f' colspan="{colspan}"' if colspan > 1 else ''
                bg = 'background-color:#f0f9ff;' if is_input_cell(cell) else ''
                td = f'<td{rs_attr}{cs_attr} style="padding:3px 6px;border:1px solid #d1d5db;{bg}{css}">'
                row_parts.append(f'{td}{val}</td>')

            row_parts.append('</tr>')
            rows_html.append(''.join(row_parts))

        sheets_html.append({
            'name': ws.title,
            'html': f'<table class="excel-render-table" style="border-collapse:collapse;font-size:12px;width:100%;table-layout:fixed;font-family:Calibri,Arial,sans-serif;min-width:1000px;">{colgroup}<tbody>{"".join(rows_html)}</tbody></table>'
        })

    return _render_template('reports_v2_review.html', submission=submission, template=template, sheets=sheets_html)