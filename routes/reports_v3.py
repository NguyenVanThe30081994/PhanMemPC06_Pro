# -*- coding: utf-8 -*-
"""
V3 Metadata-Driven Reporting Architecture
Theo V3Excel.md - 5 Module:
1. Template Importer - Doc Excel, xu li merges
2. Schema Builder - Sinh fields tu multi-row headers
3. Report API - CRUD, validate, audit
4. Permission Layer - RBAC theo don vi
5. Excel Renderer - Export giu layout goc
"""
from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, session, current_app, send_file
from models import db, ReportTemplateV3, ReportTemplateFieldV3, ReportVersionV3, ReportSubmissionV3, ReportValueV3, ReportAuditV3, User
from datetime import datetime
from werkzeug.utils import secure_filename
import json
import os
import io
import re
import unicodedata

reports_v3_bp = Blueprint('reports_v3', __name__)


# ==================== MODULE 1: TEMPLATE IMPORTER ====================
def read_excel_matrix(ws):
    """
    Doc worksheet thanh ma tran 2D + merges info.
    Xu li merged cells bang cach doc !merges.
    """
    merges = []
    for mr in ws.merged_cells.ranges:
        merges.append({
            'min_row': mr.min_row,
            'max_row': mr.max_row,
            'min_col': mr.min_col,
            'max_col': mr.max_col
        })
    
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    
    matrix = []
    for r in range(1, max_row + 1):
        row_data = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            row_data.append({
                'value': cell.value,
                'col': c,
                'row': r,
                'data_type': cell.data_type
            })
        matrix.append(row_data)
    
    return {'matrix': matrix, 'merges': merges, 'max_row': max_row, 'max_col': max_col}


def propagate_merged_headers(matrix, merges, header_row_count):
    """
    Propagate gia tri o goc trai-tren sang cac o phu trong header.
    Khi header duoc merge, chi o goc (min_row, min_col) co gia tri.
    """
    merge_map = {}
    for m in merges:
        key = f"{m['min_row']},{m['min_col']}"
        top_left_value = matrix[m['min_row'] - 1][m['min_col'] - 1]['value']
        for r in range(m['min_row'] - 1, m['max_row']):
            for c in range(m['min_col'] - 1, m['max_col']):
                merge_map[(r, c)] = top_left_value
    
    for r in range(header_row_count):
        for c in range(len(matrix[r])):
            if (r, c) in merge_map:
                matrix[r][c]['value'] = merge_map[(r, c)]
    
    return matrix


# ==================== MODULE 2: SCHEMA BUILDER ====================
def generate_field_code(header_path):
    """
    Sinh field_code on dinh tu header path.
    VD: ["Doanh thu", "Quy 1", "Thuc hien"] -> doanh_thu__quy_1__thuc_hien
    """
    path_str = '__'.join(str(h) for h in header_path if h)
    
    # Unicode normalize
    normalized = unicodedata.normalize('NFKD', path_str)
    ascii_str = normalized.encode('ascii', 'ignore').decode('ascii')
    
    # Lower, replace spaces _, remove special chars
    field_code = re.sub(r'[^a-z0-9_]', '_', ascii_str.lower())
    field_code = re.sub(r'_+', '_', field_code).strip('_')
    
    return field_code


def build_schema(ws, header_row_count=3):
    """
    Build schema JSON tu Excel template.
    Tra ve:
    - header_row_count: so dong header
    - fields: danh sach field definitions
    - grid_header_tree: cay header cho UI
    """
    data = read_excel_matrix(ws)
    matrix = propagate_merged_headers(data['matrix'], data['merges'], header_row_count)
    
    fields = []
    grid_header_tree = []
    
    # Xay grid_header_tree
    for col_idx in range(len(matrix[0])):
        col_headers = []
        for row_idx in range(header_row_count):
            val = matrix[row_idx][col_idx]['value']
            col_headers.append(str(val) if val else '')
        grid_header_tree.append(col_headers)
    
    # Xay fields cho data columns (sau header rows)
    data_start_row = header_row_count
    
    for col_idx in range(len(matrix[0])):
        header_path = [matrix[r][col_idx]['value'] for r in range(header_row_count)]
        
        # Bo cot trong
        if not any(str(h) for h in header_path if h):
            continue
        
        field_code = generate_field_code(header_path)
        
        # Detect data type
        data_type = 'text'
        if data_start_row < len(matrix):
            first_value = matrix[data_start_row][col_idx]['value']
            if first_value is not None:
                if isinstance(first_value, (int, float)):
                    data_type = 'number'
                elif isinstance(first_value, str):
                    clean_val = first_value.replace('.', '').replace('-', '').replace(',', '')
                    if clean_val.isdigit():
                        data_type = 'number'
        
        fields.append({
            'field_code': field_code,
            'field_name': ' - '.join(str(h) for h in header_path if h),
            'header_path': [str(h) for h in header_path],
            'column_index': col_idx + 1,  # 1-indexed
            'data_type': data_type,
            'editable': True,
            'required': False
        })
    
    return {
        'schema_version': '1.0',
        'header_row_count': header_row_count,
        'fields': fields,
        'grid_header_tree': grid_header_tree,
        'total_rows': data['max_row'],
        'total_cols': data['max_col']
    }


# ==================== MODULE 3: REPORT API ====================
def _render(name, **kwargs):
    kwargs['is_admin'] = session.get('is_admin')
    kwargs['fullname'] = session.get('fullname')
    kwargs['unit_area'] = session.get('unit_area')
    return render_template(name, **kwargs)


@reports_v3_bp.route('/reports-v3/')
def index():
    """Trang chu V3"""
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))
    return _render('reports_v3_index.html')


@reports_v3_bp.route('/reports-v3/templates')
def list_templates():
    """Danh sach template V3"""
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))
    
    templates = ReportTemplateV3.query.filter_by(is_active=True).order_by(ReportTemplateV3.created_at.desc()).all()
    
    data = []
    for t in templates:
        pub_versions = [v for v in t.versions if v.is_published]
        data.append({
            'id': t.id,
            'name': t.name,
            'description': t.description,
            'template_code': t.template_code,
            'version_count': len(t.versions),
            'has_published': len(pub_versions) > 0,
            'created_at': t.created_at.strftime('%d/%m/%Y') if t.created_at else ''
        })
    
    return jsonify({'success': True, 'templates': data})


@reports_v3_bp.route('/reports-v3/templates/upload', methods=['POST'])
def upload_template():
    """
    Upload Excel template.
    - Parse header, sinh schema
    - Luu template + version
    """
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400
    
    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'message': 'No file selected'}), 400
    
    name = request.form.get('name', '').strip()
    description = request.form.get('description', '').strip()
    header_count = int(request.form.get('header_row_count', 3))
    
    if not name:
        return jsonify({'success': False, 'message': 'Name is required'}), 400
    
    # Read Excel
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file.read()))
    ws = wb.active
    
    # Build schema
    schema = build_schema(ws, header_count)
    
    # Generate template_code
    base_code = re.sub(r'[^a-z0-9]', '_', name.lower())
    template_code = re.sub(r'_+', '_', base_code).strip('_')
    
    # Check uniqueness
    existing = ReportTemplateV3.query.filter_by(template_code=template_code).first()
    if existing:
        template_code = f"{template_code}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # Create template + fields
    template = ReportTemplateV3(
        name=name,
        description=description,
        template_code=template_code,
        created_by=session.get('fullname')
    )
    db.session.add(template)
    db.session.flush()
    
    # Create field definitions
    for field_def in schema['fields']:
        field = ReportTemplateFieldV3(
            template_id=template.id,
            field_code=field_def['field_code'],
            field_name=field_def['field_name'],
            header_path=json.dumps(field_def['header_path']),
            column_index=field_def['column_index'],
            data_type=field_def['data_type'],
            editable=field_def['editable'],
            required=field_def['required']
        )
        db.session.add(field)
    
    # Create version with schema + excel blob
    version = ReportVersionV3(
        template_id=template.id,
        version_tag='v1.0',
        schema_json=json.dumps(schema, ensure_ascii=False),
        excel_blob=file.read(),
        header_row_count=header_count,
        created_by=session.get('fullname'),
        is_published=True
    )
    db.session.add(version)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f'Upload thanh cong! {len(schema["fields"])} fields',
        'template_id': template.id,
        'version_id': version.id
    })


@reports_v3_bp.route('/reports-v3/templates/<int:tid>/schema')
def get_schema(tid):
    """Get schema JSON for a template"""
    if not session.get('uid'):
        return jsonify({'success': False}), 403
    
    template = db.session.get(ReportTemplateV3, tid)
    if not template:
        return jsonify({'success': False, 'message': 'Not found'}), 404
    
    version = ReportVersionV3.query.filter_by(template_id=tid, is_published=True).first()
    if not version or not version.schema_json:
        return jsonify({'success': False, 'message': 'No published version'}), 404
    
    return jsonify({
        'success': True,
        'schema': json.loads(version.schema_json)
    })


@reports_v3_bp.route('/reports-v3/reports')
def list_reports():
    """
    List reports cho don vi hien tai.
    Permission: chi xem bao cao cua don vi minh.
    """
    if not session.get('uid'):
        return jsonify({'success': False}), 403
    
    unit_id = session.get('unit_area', '')
    period_id = request.args.get('period_id')
    
    query = ReportSubmissionV3.query.filter_by(unit_id=unit_id)
    if period_id:
        query = query.filter_by(period_id=period_id)
    
    submissions = query.order_by(ReportSubmissionV3.created_at.desc()).all()
    
    data = []
    for s in submissions:
        version = s.version
        template = version.template if version else None
        data.append({
            'id': s.id,
            'template_name': template.name if template else 'N/A',
            'period_name': s.period_name,
            'status': s.status,
            'created_at': s.created_at.strftime('%d/%m/%Y %H:%M') if s.created_at else '',
            'updated_at': s.updated_at.strftime('%d/%m/%Y %H:%M') if s.updated_at else ''
        })
    
    return jsonify({'success': True, 'reports': data})


@reports_v3_bp.route('/reports-v3/reports/create', methods=['POST'])
def create_report():
    """
    Tao report instance moi.
    Permission: user thuoc don vi nao tao report cho don vi do.
    """
    if not session.get('uid'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    data = request.json or {}
    template_id = data.get('template_id')
    period_id = data.get('period_id')
    period_name = data.get('period_name', '')
    
    if not template_id:
        return jsonify({'success': False, 'message': 'Thieu template_id'}), 400
    
    template = db.session.get(ReportTemplateV3, template_id)
    if not template:
        return jsonify({'success': False, 'message': 'Template not found'}), 404
    
    version = ReportVersionV3.query.filter_by(template_id=template_id, is_published=True).first()
    if not version:
        return jsonify({'success': False, 'message': 'No published version'}), 404
    
    unit_id = session.get('unit_area', '')
    
    # Check existing
    existing = ReportSubmissionV3.query.filter_by(
        version_id=version.id,
        unit_id=unit_id,
        period_id=period_id
    ).first()
    
    if existing:
        return jsonify({'success': True, 'report_id': existing.id, 'message': 'Report da ton tai'})
    
    submission = ReportSubmissionV3(
        version_id=version.id,
        user_id=session.get('uid'),
        unit_id=unit_id,
        period_id=period_id,
        period_name=period_name,
        status='draft'
    )
    db.session.add(submission)
    db.session.commit()
    
    return jsonify({'success': True, 'report_id': submission.id})


@reports_v3_bp.route('/reports-v3/reports/<int:rid>/values')
def get_values(rid):
    """
    Get values cua mot report.
    Permission: chi don vi so huu moi doc duoc.
    """
    if not session.get('uid'):
        return jsonify({'success': False}), 403
    
    submission = db.session.get(ReportSubmissionV3, rid)
    if not submission:
        return jsonify({'success': False, 'message': 'Not found'}), 404
    
    # Permission check
    if submission.unit_id != session.get('unit_area'):
        return jsonify({'success': False, 'message': 'Khong co quyen truy cap'}), 403
    
    values = {v.field_code: v.value for v in submission.values}
    
    return jsonify({
        'success': True,
        'values': values,
        'status': submission.status
    })


@reports_v3_bp.route('/reports-v3/reports/<int:rid>/values', methods=['PUT'])
def save_values(rid):
    """
    Save cell values voi audit log.
    Permission: chi don vi so huu moi ghi duoc.
    """
    if not session.get('uid'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    submission = db.session.get(ReportSubmissionV3, rid)
    if not submission:
        return jsonify({'success': False, 'message': 'Not found'}), 404
    
    # Permission check
    if submission.unit_id != session.get('unit_area'):
        return jsonify({'success': False, 'message': 'Khong co quyen ghi'}), 403
    
    data = request.json or {}
    values = data.get('values', {})  # {field_code: value}
    
    # Get old values for audit
    old_values = {v.field_code: v.value for v in submission.values}
    
    # Upsert values
    for field_code, value in values.items():
        existing = ReportValueV3.query.filter_by(
            submission_id=submission.id,
            field_code=field_code
        ).first()
        
        old_val = old_values.get(field_code, '')
        
        if existing:
            if existing.value != str(value):
                # Audit log
                audit = ReportAuditV3(
                    submission_id=submission.id,
                    user_id=session.get('uid'),
                    field_code=field_code,
                    old_value=old_val,
                    new_value=str(value) if value else ''
                )
                db.session.add(audit)
                existing.value = str(value) if value else ''
        else:
            # Find field_id
            field = ReportTemplateFieldV3.query.filter_by(
                template_id=submission.version.template_id,
                field_code=field_code
            ).first()
            
            val = ReportValueV3(
                submission_id=submission.id,
                field_id=field.id if field else None,
                field_code=field_code,
                value=str(value) if value else ''
            )
            db.session.add(val)
            
            if old_val != str(value):
                audit = ReportAuditV3(
                    submission_id=submission.id,
                    user_id=session.get('uid'),
                    field_code=field_code,
                    old_value=old_val,
                    new_value=str(value) if value else ''
                )
                db.session.add(audit)
    
    submission.updated_at = datetime.now()
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Da luu'})


@reports_v3_bp.route('/reports-v3/reports/<int:rid>/submit', methods=['POST'])
def submit_report(rid):
    """
    Submit report - danh dau da hoan thanh.
    Permission: chi don vi so huu moi submit duoc.
    """
    if not session.get('uid'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    submission = db.session.get(ReportSubmissionV3, rid)
    if not submission:
        return jsonify({'success': False, 'message': 'Not found'}), 404
    
    if submission.unit_id != session.get('unit_area'):
        return jsonify({'success': False, 'message': 'Khong co quyen'}), 403
    
    submission.status = 'submitted'
    submission.submitted_at = datetime.now()
    submission.updated_at = datetime.now()
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Da submit'})


# ==================== MODULE 5: EXCEL RENDERER ====================
@reports_v3_bp.route('/reports-v3/reports/<int:rid>/export')
def export_report(rid):
    """
    Export Excel giu layout goc voi du lieu.
    Permission: chi don vi so huu moi export duoc.
    """
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))
    
    submission = db.session.get(ReportSubmissionV3, rid)
    if not submission:
        return "Not found", 404
    
    if submission.unit_id != session.get('unit_area'):
        return "Khong co quyen", 403
    
    version = submission.version
    if not version or not version.excel_blob:
        return "No template", 404
    
    # Load workbook
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(version.excel_blob))
    ws = wb.active
    
    # Get values
    values_map = {v.field_code: v.value for v in submission.values}
    
    # Get schema to map field_code -> column_index
    schema = {}
    if version.schema_json:
        schema = json.loads(version.schema_json)
    
    field_map = {}
    for f in schema.get('fields', []):
        field_map[f['column_index']] = f['field_code']
    
    header_count = schema.get('header_row_count', 3)
    
    # Fill data
    for col_idx, field_code in field_map.items():
        value = values_map.get(field_code, '')
        if value:
            col_letter_idx = col_idx - 1  # 0-indexed
            for row_idx in range(header_count, ws.max_row + 1):
                cell = ws.cell(row=row_idx + 1, column=col_letter_idx + 1)
                if cell.value is None:
                    cell.value = value
                    break
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"BaoCao_{submission.period_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@reports_v3_bp.route('/reports-v3/templates/<int:tid>/delete', methods=['POST'])
def delete_template(tid):
    """Xoa template (admin only)"""
    if not session.get('is_admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    template = db.session.get(ReportTemplateV3, tid)
    if not template:
        return jsonify({'success': False, 'message': 'Not found'}), 404
    
    # Cascade delete
    for version in template.versions:
        for sub in version.submissions:
            for audit in sub.audits:
                db.session.delete(audit)
            for val in sub.values:
                db.session.delete(val)
            db.session.delete(sub)
        db.session.delete(version)
    
    for field in template.fields:
        db.session.delete(field)
    
    db.session.delete(template)
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'Da xoa template'})


@reports_v3_bp.route('/reports-v3/input/<int:tid>')
def input_report(tid):
    """Man hinh nhap lieu cho mot template"""
    if not session.get('uid'):
        return redirect(url_for('auth_bp.login'))
    
    template = db.session.get(ReportTemplateV3, tid)
    if not template:
        flash('Khong tim thay template!', 'danger')
        return redirect(url_for('dashboard_bp.dashboard'))
    
    version = ReportVersionV3.query.filter_by(template_id=tid, is_published=True).order_by(ReportVersionV3.created_at.desc()).first()
    if not version:
        flash('Khong co phien ban nao!', 'danger')
        return redirect(url_for('dashboard_bp.dashboard'))
    
    # Parse schema
    schema = {}
    if version.schema_json:
        try:
            schema = json.loads(version.schema_json)
        except:
            pass
    
    # Get or create submission for current unit
    unit_id = session.get('unit_area', '')
    submission = ReportSubmissionV3.query.filter_by(
        version_id=version.id,
        unit_id=unit_id
    ).first()
    
    report_id = submission.id if submission else None
    
    return _render('reports_v3_input.html', 
                 template=template, 
                 version=version, 
                 config=json.dumps(schema),
                 report_id=report_id)
