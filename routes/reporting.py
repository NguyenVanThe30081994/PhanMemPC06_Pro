# -*- coding: utf-8 -*-
import calendar
import json
import os
from datetime import date, datetime, time, timedelta

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from flask import Blueprint, current_app, flash, g, jsonify, redirect, request, session, send_file, url_for
from sqlalchemy import func, or_
from werkzeug.utils import secure_filename

from models import (
    AppRole,
    db,
    ReportAuditLog,
    ReportCycle,
    ReportExportJob,
    ReportInstance,
    ReportSubmission,
    ReportSubmissionCell,
    ReportSubmissionValue,
    ReportTemplate,
    ReportTemplateField,
    ReportTemplateSheet,
    ReportTemplateVersion,
    ReportType,
    ReportUnit,
    ReportValidationLog,
    ReportingPeriod,
    User,
)
from category_helpers import get_category_items, get_module_field_items
from report_engine import build_preview_workbook, normalize_code, parse_workbook, render_sheet_html, safe_filename, write_workbook_copy
from utils import apply_migrations, extract_unit_key, is_unit_match, log_action, render_auto_template as render_template

reporting_bp = Blueprint("reporting_bp", __name__)

PERIODIC_CYCLE_OPTIONS = [
    ("monthly", "Hàng tháng"),
    ("quarterly", "Hàng quý"),
    ("half_year", "6 tháng"),
    ("yearly", "1 năm"),
]

PERIODIC_CYCLE_LABELS = {code: label for code, label in PERIODIC_CYCLE_OPTIONS}


def _ensure_report_schema():
    apply_migrations(current_app)


def _is_admin():
    return bool(session.get("is_admin"))


def _is_pc06_role(user):
    if not user or not getattr(user, "role_id", None):
        return False
    role = getattr(user, "role", None) or db.session.get(AppRole, user.role_id)
    role_name = normalize_code(getattr(role, "name", ""))
    return "pc06" in role_name


def _require_login():
    return bool(session.get("uid"))


def _ensure_default_types():
    defaults = [
        ("daily", "Hằng ngày", "daily", "Báo cáo hằng ngày"),
        ("periodic", "Định kỳ", "periodic", "Báo cáo định kỳ"),
        ("ad_hoc", "Đột xuất", "ad_hoc", "Báo cáo đột xuất"),
    ]
    changed = False
    for code, name, freq, desc in defaults:
        if not ReportType.query.filter_by(code=code).first():
            db.session.add(ReportType(code=code, name=name, frequency=freq, description=desc))
            changed = True
    if changed:
        db.session.commit()


def _looks_like_org_unit(value):
    text = (value or "").strip().lower()
    if not text:
        return False
    markers = [
        "cong an",
        "ubnd",
        "doi ",
        "đội ",
        "phong ",
        "phòng ",
        "ban ",
        "xa ",
        "xã ",
        "phuong ",
        "phường ",
        "thi tran",
        "thị trấn",
        "quan ",
        "huyen ",
    ]
    return any(marker in text for marker in markers)


def _is_generic_unit_key(value):
    return (value or "").strip().lower() in {"xa", "phuong", "huyen", "quan", "tp", "thi", "tran"}


def _preferred_unit_identity(user):
    unit_area = (getattr(user, "unit_area", None) or "").strip()
    fullname = (getattr(user, "fullname", None) or "").strip()
    username = (getattr(user, "username", None) or "").strip()

    unit_key = extract_unit_key(unit_area) if unit_area else ""
    if unit_area and unit_key and not _is_generic_unit_key(unit_key):
        return unit_area, unit_key

    if fullname and _looks_like_org_unit(fullname):
        fullname_key = extract_unit_key(fullname) or normalize_code(fullname)
        if fullname_key and not _is_generic_unit_key(fullname_key):
            return fullname, fullname_key

    if unit_area:
        return unit_area, unit_key or normalize_code(unit_area)

    if fullname:
        return fullname, extract_unit_key(fullname) or normalize_code(fullname)

    return username, normalize_code(username)


def _sync_units_from_users():
    users = (
        User.query.filter(User.is_active.is_(True))
        .order_by(User.unit_area.asc(), User.fullname.asc())
        .all()
    )
    seen = set()
    changed = False
    for user in users:
        name, key = _preferred_unit_identity(user)
        key = normalize_code(key)
        if not key or not name or key in seen:
            continue
        seen.add(key)
        unit = ReportUnit.query.filter_by(code=key).first()
        if not unit:
            db.session.add(ReportUnit(code=key, name=name, source="user", is_active=True))
            changed = True
        elif unit.name != name:
            unit.name = name
            unit.source = "user"
            changed = True
    if changed:
        db.session.commit()


def _current_user_report_unit():
    if not session.get("uid"):
        return None
    user = db.session.get(User, session.get("uid"))
    if not user:
        return None
    preferred_name, preferred_key = _preferred_unit_identity(user)
    key = normalize_code(preferred_key)
    if key:
        unit = ReportUnit.query.filter_by(code=key, is_active=True).first()
        if unit:
            return unit
    if preferred_name:
        key = normalize_code(preferred_name)
        unit = ReportUnit.query.filter_by(code=key, is_active=True).first()
        if unit:
            return unit
    return None


def _template_version(template):
    version = (
        ReportTemplateVersion.query.filter_by(template_id=template.id, is_current=True)
        .order_by(ReportTemplateVersion.version_no.desc())
        .first()
    )
    if version:
        return version
    return (
        ReportTemplateVersion.query.filter_by(template_id=template.id)
        .order_by(ReportTemplateVersion.version_no.desc())
        .first()
    )


def _template_file_path(version):
    return version.source_path if version else None


def _version_usage_count(version_id):
    return ReportCycle.query.filter_by(template_version_id=version_id).count()


def _template_version_ids(template_id):
    return [row.id for row in ReportTemplateVersion.query.filter_by(template_id=template_id).all()]


def _professional_unit_options():
    return get_module_field_items("tasks", "domain") or get_category_items("Đội nghiệp vụ")


def _template_professional_unit(template):
    return (getattr(template, "professional_unit", None) or "").strip() or "Chưa phân đội"


def _unit_group_options():
    return get_module_field_items("contacts", "unit_name") or get_category_items("Đơn vị")


def _empty_scope_payload():
    return {
        "mode": "all",
        "unit_ids": [],
        "role_ids": [],
        "unit_groups": [],
    }


def _parse_scope_payload(raw_value):
    payload = _empty_scope_payload()
    if not raw_value:
        return payload
    try:
        data = json.loads(raw_value)
    except Exception:
        data = []

    if isinstance(data, list):
        payload["unit_ids"] = sorted({int(v) for v in data if str(v).isdigit()})
        if payload["unit_ids"]:
            payload["mode"] = "targeted"
        return payload

    if not isinstance(data, dict):
        return payload

    payload["unit_ids"] = sorted({int(v) for v in data.get("unit_ids", []) if str(v).isdigit()})
    payload["role_ids"] = sorted({int(v) for v in data.get("role_ids", []) if str(v).isdigit()})
    payload["unit_groups"] = sorted(
        {
            str(v).strip()
            for v in data.get("unit_groups", [])
            if str(v).strip()
        }
    )
    mode = (data.get("mode") or "").strip().lower()
    payload["mode"] = "targeted" if (payload["unit_ids"] or payload["role_ids"] or payload["unit_groups"]) else "all"
    if mode == "all" and not (payload["unit_ids"] or payload["role_ids"] or payload["unit_groups"]):
        payload["mode"] = "all"
    return payload


def _scope_payload_from_form(form):
    payload = _empty_scope_payload()
    payload["unit_ids"] = sorted({int(v) for v in form.getlist("unit_ids") if str(v).isdigit()})
    payload["role_ids"] = sorted({int(v) for v in form.getlist("role_ids") if str(v).isdigit()})
    payload["unit_groups"] = sorted({str(v).strip() for v in form.getlist("unit_groups") if str(v).strip()})
    if payload["unit_ids"] or payload["role_ids"] or payload["unit_groups"]:
        payload["mode"] = "targeted"
    return payload


def _find_report_unit_for_identity(all_units, unit_name, unit_key):
    key = normalize_code(unit_key)
    if key:
        for unit in all_units:
            if normalize_code(unit.code) == key:
                return unit
    name = (unit_name or "").strip()
    if name:
        for unit in all_units:
            if is_unit_match(unit.name, name):
                return unit
    return None


def _unit_matches_group(unit, group_name):
    group_value = (group_name or "").strip()
    if not group_value:
        return False
    if is_unit_match(unit.name, group_value) or is_unit_match(unit.code, group_value):
        return True
    normalized_group = normalize_code(group_value)
    normalized_name = normalize_code(unit.name)
    normalized_code = normalize_code(unit.code)
    if not normalized_group:
        return False
    return (
        normalized_name == normalized_group
        or normalized_code == normalized_group
        or normalized_name.startswith(f"{normalized_group}_")
        or normalized_code.startswith(f"{normalized_group}_")
    )


def _resolve_scope_units(scope_payload):
    all_units = ReportUnit.query.filter_by(is_active=True).order_by(ReportUnit.name.asc()).all()
    has_explicit_scope = bool(
        scope_payload.get("unit_ids") or scope_payload.get("role_ids") or scope_payload.get("unit_groups")
    )
    if scope_payload.get("mode") == "all" and not has_explicit_scope:
        return all_units

    selected_ids = set(int(v) for v in scope_payload.get("unit_ids", []) if str(v).isdigit())

    role_ids = [int(v) for v in scope_payload.get("role_ids", []) if str(v).isdigit()]
    if role_ids:
        users = User.query.filter(User.role_id.in_(role_ids), User.is_active.is_(True)).all()
        for user in users:
            unit_name, unit_key = _preferred_unit_identity(user)
            matched_unit = _find_report_unit_for_identity(all_units, unit_name, unit_key)
            if matched_unit:
                selected_ids.add(matched_unit.id)

    unit_groups = [group for group in scope_payload.get("unit_groups", []) if str(group).strip()]
    if unit_groups:
        for unit in all_units:
            if any(_unit_matches_group(unit, group_name) for group_name in unit_groups):
                selected_ids.add(unit.id)

    if not selected_ids and not has_explicit_scope:
        return all_units
    if not selected_ids:
        return []
    return [unit for unit in all_units if unit.id in selected_ids]


def _template_current_cycle(template):
    version_ids = _template_version_ids(template.id)
    if not version_ids:
        return None
    cycles = (
        ReportCycle.query.filter(ReportCycle.template_version_id.in_(version_ids))
        .order_by(ReportCycle.created_at.desc())
        .all()
    )
    for cycle in cycles:
        if cycle.status != "closed":
            return cycle
    return cycles[0] if cycles else None


def _save_directive_file(file_storage, template_code):
    if not file_storage or not file_storage.filename:
        return "", ""
    original_name = secure_filename(file_storage.filename)
    if not original_name:
        return "", ""
    directive_dir = os.path.join(current_app.root_path, "report_templates", "directives")
    os.makedirs(directive_dir, exist_ok=True)
    filename = safe_filename(
        f"{normalize_code(template_code) or 'report'}_directive_{datetime.now().strftime('%Y%m%d%H%M%S')}_{original_name}"
    )
    storage_path = os.path.join(directive_dir, filename)
    file_storage.save(storage_path)
    return original_name, storage_path


def _template_range_defaults(version):
    defaults = {
        "header_start_row": 1,
        "header_end_row": 2,
        "unit_start_row": 3,
        "unit_end_row": 20,
        "total_start_row": 21,
        "total_end_row": 21,
        "start_column": "A",
        "end_column": "Z",
    }
    if not version:
        return defaults
    metadata = json.loads(version.metadata_json or "{}")
    first_sheet = (metadata.get("sheets") or [{}])[0]
    defaults.update(
        {
            "header_start_row": first_sheet.get("header_start_row") or defaults["header_start_row"],
            "header_end_row": first_sheet.get("header_end_row") or defaults["header_end_row"],
            "unit_start_row": first_sheet.get("unit_start_row") or first_sheet.get("data_start_row") or defaults["unit_start_row"],
            "unit_end_row": first_sheet.get("unit_end_row") or first_sheet.get("data_end_row") or defaults["unit_end_row"],
            "total_start_row": first_sheet.get("total_start_row") or defaults["total_start_row"],
            "total_end_row": first_sheet.get("total_end_row") or defaults["total_end_row"],
            "start_column": first_sheet.get("start_column") or defaults["start_column"],
            "end_column": first_sheet.get("end_column") or defaults["end_column"],
        }
    )
    return defaults


def _workbook_sheet_names(source_path):
    if not source_path or not os.path.exists(source_path):
        return []
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _template_sheet_range_defaults(version):
    defaults = _template_range_defaults(version)
    metadata = json.loads((version.metadata_json if version else "") or "{}")
    metadata_sheets = metadata.get("sheets") or []
    configured = []
    configured_names = set()

    for sheet in metadata_sheets:
        sheet_name = (sheet.get("sheet_name") or "").strip()
        if not sheet_name or sheet_name in configured_names:
            continue
        configured_names.add(sheet_name)
        configured.append(
            {
                "sheet_name": sheet_name,
                "header_start_row": sheet.get("header_start_row") or defaults["header_start_row"],
                "header_end_row": sheet.get("header_end_row") or defaults["header_end_row"],
                "unit_start_row": sheet.get("unit_start_row") or sheet.get("data_start_row") or defaults["unit_start_row"],
                "unit_end_row": sheet.get("unit_end_row") or sheet.get("data_end_row") or defaults["unit_end_row"],
                "total_start_row": sheet.get("total_start_row") or defaults["total_start_row"],
                "total_end_row": sheet.get("total_end_row") or defaults["total_end_row"],
                "start_column": sheet.get("start_column") or defaults["start_column"],
                "end_column": sheet.get("end_column") or defaults["end_column"],
            }
        )

    for sheet_name in _workbook_sheet_names(version.source_path if version else ""):
        if sheet_name in configured_names:
            continue
        configured.append({"sheet_name": sheet_name, **defaults})

    return configured


def _build_parse_options(form, defaults=None):
    defaults = defaults or {}
    header_start_row = _parse_positive_int(form.get("header_start_row"), defaults.get("header_start_row", 1))
    header_end_row = _parse_positive_int(form.get("header_end_row"), max(header_start_row, defaults.get("header_end_row", 2)))
    if header_end_row < header_start_row:
        header_end_row = header_start_row
    unit_start_row = _parse_positive_int(form.get("unit_start_row"), max(header_end_row + 1, defaults.get("unit_start_row", 3)))
    unit_end_row = _parse_positive_int(form.get("unit_end_row"), max(unit_start_row, defaults.get("unit_end_row", unit_start_row)))
    if unit_end_row < unit_start_row:
        unit_end_row = unit_start_row
    total_start_row = _parse_positive_int(form.get("total_start_row"), max(unit_end_row + 1, defaults.get("total_start_row", unit_end_row + 1)))
    total_end_row = _parse_positive_int(form.get("total_end_row"), max(total_start_row, defaults.get("total_end_row", total_start_row)))
    if total_end_row < total_start_row:
        total_end_row = total_start_row
    start_column = (form.get("start_column") or defaults.get("start_column") or "A").strip().upper()
    end_column = (form.get("end_column") or defaults.get("end_column") or "Z").strip().upper()
    return {
        "header_start_row": header_start_row,
        "header_end_row": header_end_row,
        "header_rows": max(1, header_end_row - header_start_row + 1),
        "data_start_row": unit_start_row,
        "data_end_row": unit_end_row,
        "total_start_row": total_start_row,
        "total_end_row": total_end_row,
        "start_column": start_column,
        "end_column": end_column,
    }


def _build_sheet_parse_options(form, defaults=None, workbook_sheet_names=None, global_defaults=None):
    defaults = defaults or []
    workbook_sheet_names = workbook_sheet_names or []
    global_defaults = global_defaults or {}
    defaults_by_name = {item.get("sheet_name"): item for item in defaults if item.get("sheet_name")}

    names = form.getlist("sheet_name")
    header_start_rows = form.getlist("sheet_header_start_row")
    header_end_rows = form.getlist("sheet_header_end_row")
    unit_start_rows = form.getlist("sheet_unit_start_row")
    unit_end_rows = form.getlist("sheet_unit_end_row")
    total_start_rows = form.getlist("sheet_total_start_row")
    total_end_rows = form.getlist("sheet_total_end_row")
    start_columns = form.getlist("sheet_start_column")
    end_columns = form.getlist("sheet_end_column")

    def _value_at(items, index, fallback=""):
        return items[index] if index < len(items) else fallback

    def _coerce_for_sheet(sheet_name, overrides=None):
        overrides = overrides or {}
        base = defaults_by_name.get(sheet_name, {})
        header_start_row = _parse_positive_int(
            overrides.get("header_start_row"),
            base.get("header_start_row", global_defaults.get("header_start_row", 1)),
        )
        header_end_row = _parse_positive_int(
            overrides.get("header_end_row"),
            max(header_start_row, base.get("header_end_row", global_defaults.get("header_end_row", header_start_row))),
        )
        if header_end_row < header_start_row:
            header_end_row = header_start_row
        unit_start_row = _parse_positive_int(
            overrides.get("unit_start_row"),
            max(header_end_row + 1, base.get("unit_start_row", global_defaults.get("unit_start_row", header_end_row + 1))),
        )
        unit_end_row = _parse_positive_int(
            overrides.get("unit_end_row"),
            max(unit_start_row, base.get("unit_end_row", global_defaults.get("unit_end_row", unit_start_row))),
        )
        if unit_end_row < unit_start_row:
            unit_end_row = unit_start_row
        total_start_row = _parse_positive_int(
            overrides.get("total_start_row"),
            max(unit_end_row + 1, base.get("total_start_row", global_defaults.get("total_start_row", unit_end_row + 1))),
        )
        total_end_row = _parse_positive_int(
            overrides.get("total_end_row"),
            max(total_start_row, base.get("total_end_row", global_defaults.get("total_end_row", total_start_row))),
        )
        if total_end_row < total_start_row:
            total_end_row = total_start_row
        start_column = (
            overrides.get("start_column")
            or base.get("start_column")
            or global_defaults.get("start_column")
            or "A"
        ).strip().upper()
        end_column = (
            overrides.get("end_column")
            or base.get("end_column")
            or global_defaults.get("end_column")
            or start_column
        ).strip().upper()
        return {
            "header_start_row": header_start_row,
            "header_end_row": header_end_row,
            "header_rows": max(1, header_end_row - header_start_row + 1),
            "data_start_row": unit_start_row,
            "data_end_row": unit_end_row,
            "total_start_row": total_start_row,
            "total_end_row": total_end_row,
            "start_column": start_column,
            "end_column": end_column,
        }

    options = {}
    for index, raw_name in enumerate(names):
        sheet_name = (raw_name or "").strip()
        if not sheet_name:
            continue
        options[sheet_name] = _coerce_for_sheet(
            sheet_name,
            overrides={
                "header_start_row": _value_at(header_start_rows, index),
                "header_end_row": _value_at(header_end_rows, index),
                "unit_start_row": _value_at(unit_start_rows, index),
                "unit_end_row": _value_at(unit_end_rows, index),
                "total_start_row": _value_at(total_start_rows, index),
                "total_end_row": _value_at(total_end_rows, index),
                "start_column": _value_at(start_columns, index),
                "end_column": _value_at(end_columns, index),
            },
        )

    for sheet_name in workbook_sheet_names:
        if sheet_name in options:
            continue
        options[sheet_name] = _coerce_for_sheet(sheet_name)

    return options


def _save_template_periodic_config(template, report_type, form):
    if not template:
        return None
    if not report_type or report_type.code != "periodic":
        template.periodic_cycle_type = None
        template.periodic_due_day = None
        template.periodic_due_month = None
        return None

    cycle_type = (form.get("periodic_cycle_type") or "").strip().lower()
    if cycle_type not in PERIODIC_CYCLE_LABELS:
        return "Bạn cần chọn loại chu kỳ cho báo cáo định kỳ."

    due_day = _parse_bounded_int(form.get("periodic_due_day"), 1, 31, None)
    if due_day is None:
        return "Bạn cần nhập ngày hạn hợp lệ cho báo cáo định kỳ."

    due_month = None
    if cycle_type in {"half_year", "yearly"}:
        due_month = _parse_bounded_int(form.get("periodic_due_month"), 1, 12, None)
        if due_month is None:
            return "Bạn cần nhập tháng hạn hợp lệ cho báo cáo định kỳ."

    template.periodic_cycle_type = cycle_type
    template.periodic_due_day = due_day
    template.periodic_due_month = due_month
    return None


def _apply_version_structure(version, source_path, source_filename, parse_options, notes="", sheet_parse_options=None):
    metadata = parse_workbook(
        source_path,
        header_rows=parse_options["header_rows"],
        data_start_row=parse_options["data_start_row"],
        header_start_row=parse_options["header_start_row"],
        header_end_row=parse_options["header_end_row"],
        data_end_row=parse_options["data_end_row"],
        total_start_row=parse_options["total_start_row"],
        total_end_row=parse_options["total_end_row"],
        start_column=parse_options["start_column"],
        end_column=parse_options["end_column"],
        sheet_options=sheet_parse_options,
    )

    version.source_filename = source_filename
    version.source_path = source_path
    version.metadata_json = json.dumps(metadata, ensure_ascii=False)
    version.notes = notes

    ReportTemplateField.query.filter_by(version_id=version.id).delete(synchronize_session=False)
    ReportTemplateSheet.query.filter_by(version_id=version.id).delete(synchronize_session=False)

    for sheet in metadata.get("sheets", []):
        db.session.add(
            ReportTemplateSheet(
                version_id=version.id,
                sheet_name=sheet["sheet_name"],
                order_index=sheet["order_index"],
                header_start_row=sheet.get("header_start_row") or sheet.get("min_row") or 1,
                header_end_row=sheet.get("header_end_row") or ((sheet.get("header_start_row") or 1) + max(int(sheet.get("header_rows") or 1) - 1, 0)),
                header_rows=sheet["header_rows"],
                data_start_row=sheet["data_start_row"],
                data_end_row=sheet["data_end_row"],
                unit_key_column="A",
                can_input=True,
                visible_in_preview=True,
                summary_json=json.dumps(
                    {
                        "fields": len(sheet.get("fields", [])),
                        "input_cells": len(sheet.get("input_cells", [])),
                        "unit_start_row": sheet.get("unit_start_row"),
                        "unit_end_row": sheet.get("unit_end_row"),
                        "total_start_row": sheet.get("total_start_row"),
                        "total_end_row": sheet.get("total_end_row"),
                        "start_column": sheet.get("start_column"),
                        "end_column": sheet.get("end_column"),
                    },
                    ensure_ascii=False,
                ),
            )
        )
        for field in sheet.get("fields", []):
            db.session.add(
                ReportTemplateField(
                    version_id=version.id,
                    sheet_name=sheet["sheet_name"],
                    field_code=field["field_code"],
                    field_name=field["field_name"],
                    display_name=field["field_name"],
                    column_index=field["column_index"],
                    column_letter=field["column_letter"],
                    data_type=field["data_type"],
                    input_mode=field["input_mode"],
                    is_required=field["is_required"],
                    is_visible=field["is_visible"],
                    is_editable=field["is_editable"],
                    default_value=field["default_value"],
                    validation_rule=field["validation_rule"],
                    dictionary_source=field["dictionary_source"],
                    formula_expression=field["formula_expression"],
                    aggregation_type=field["aggregation_type"],
                    display_order=field["display_order"],
                    path_code=field["path_code"],
                )
            )
    return metadata


def _preview_workbook(file_path):
    return load_workbook(file_path, data_only=True)


def _purge_cycle(cycle):
    instances = ReportInstance.query.filter_by(cycle_id=cycle.id).all()
    for instance in instances:
        submissions = ReportSubmission.query.filter_by(instance_id=instance.id).all()
        for submission in submissions:
            _purge_submission(submission)
        db.session.delete(instance)
    ReportExportJob.query.filter_by(cycle_id=cycle.id).delete(synchronize_session=False)
    db.session.delete(cycle)


def _scope_unit_ids(cycle):
    return {unit.id for unit in _resolve_scope_units(_parse_scope_payload(cycle.scope_json))}


def _cycle_accessible(cycle, unit_id, is_admin=False):
    if is_admin:
        return True
    if not unit_id:
        return False
    scope_payload = _parse_scope_payload(cycle.scope_json)
    has_explicit_scope = bool(
        scope_payload.get("unit_ids") or scope_payload.get("role_ids") or scope_payload.get("unit_groups")
    )
    scope = {unit.id for unit in _resolve_scope_units(scope_payload)}
    if not has_explicit_scope:
        return True
    return unit_id in scope


def _get_cycle_instance(cycle, unit, user):
    instance = ReportInstance.query.filter_by(
        cycle_id=cycle.id,
        report_unit_id=unit.id if unit else None,
    ).first()
    if not instance and unit:
        template_version = db.session.get(ReportTemplateVersion, cycle.template_version_id)
        period = _ensure_reporting_period(cycle)
        instance = ReportInstance(
            template_id=template_version.template_id if template_version else None,
            version_id=template_version.id if template_version else None,
            period_id=period.id if period else None,
            user_id=user.id if user else None,
            org_unit=unit.name,
            cycle_id=cycle.id,
            report_unit_id=unit.id,
            assigned_user_id=user.id if user else None,
            status="draft",
        )
        db.session.add(instance)
        db.session.commit()
    return instance


def _latest_submission(instance_id):
    return (
        ReportSubmission.query.filter_by(instance_id=instance_id)
        .order_by(ReportSubmission.version_no.desc(), ReportSubmission.created_at.desc())
        .first()
    )


def _submission_metadata(submission):
    if not submission or not submission.metadata_json:
        return {}
    try:
        payload = json.loads(submission.metadata_json or "{}")
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _submission_business_date(submission):
    metadata = _submission_metadata(submission)
    report_date = _parse_date(metadata.get("report_date", "")) if metadata else None
    if report_date:
        return report_date
    submission_time = submission.submitted_at or submission.created_at
    return submission_time.date() if submission_time else None


def _ordered_submissions(instance_id, descending=False):
    if descending:
        return (
            ReportSubmission.query.filter_by(instance_id=instance_id)
            .order_by(ReportSubmission.version_no.desc(), ReportSubmission.created_at.desc(), ReportSubmission.id.desc())
            .all()
        )
    return (
        ReportSubmission.query.filter_by(instance_id=instance_id)
        .order_by(ReportSubmission.version_no.asc(), ReportSubmission.created_at.asc(), ReportSubmission.id.asc())
        .all()
    )


def _latest_submission_for_date(instance_id, report_date):
    matches = [submission for submission in _ordered_submissions(instance_id) if _submission_business_date(submission) == report_date]
    return matches[-1] if matches else None


def _latest_submission_through_date(instance_id, report_date):
    matches = [
        submission
        for submission in _ordered_submissions(instance_id)
        if _submission_business_date(submission) and _submission_business_date(submission) <= report_date
    ]
    return matches[-1] if matches else None


def _submission_history(instance_id, report_date=None):
    submissions = _ordered_submissions(instance_id, descending=True)
    if report_date:
        submissions = [
            submission
            for submission in submissions
            if _submission_business_date(submission) == report_date
        ]
    return submissions


def _submission_history_through_date(instance_id, report_date):
    return [
        submission
        for submission in _ordered_submissions(instance_id, descending=True)
        if _submission_business_date(submission) and _submission_business_date(submission) <= report_date
    ]


def _submissions_through_date(instance_id, report_date):
    return [
        submission
        for submission in _ordered_submissions(instance_id)
        if _submission_business_date(submission) and _submission_business_date(submission) <= report_date
    ]


def _daily_snapshot_submissions_through_date(instance_id, report_date):
    latest_by_day = {}
    for submission in _submissions_through_date(instance_id, report_date):
        business_date = _submission_business_date(submission)
        if not business_date:
            continue
        latest_by_day[business_date] = submission
    return [latest_by_day[day] for day in sorted(latest_by_day.keys())]


def _parse_numeric_value(raw_value):
    if raw_value is None:
        return None
    text = str(raw_value).strip()
    if not text:
        return None
    normalized = text.replace(",", "")
    try:
        return float(normalized)
    except Exception:
        return None


def _cumulative_submission_cell_values(submissions, template_version_id):
    if not submissions or not template_version_id:
        return {}

    fields = ReportTemplateField.query.filter_by(version_id=template_version_id).all()
    field_map = {(field.sheet_name, field.column_letter.upper()): field for field in fields}
    numeric_types = {"number", "float", "decimal", "int", "integer"}
    values = {}
    numeric_totals = {}

    for submission in submissions:
        rows = (
            ReportSubmissionCell.query.filter_by(submission_id=submission.id)
            .order_by(ReportSubmissionCell.id.asc())
            .all()
        )
        for row in rows:
            cell_address = (row.cell_address or "").strip().upper()
            if not cell_address:
                continue
            col_letters = "".join(ch for ch in cell_address if ch.isalpha()).upper()
            field = field_map.get((row.sheet_name, col_letters))
            sheet_values = values.setdefault(row.sheet_name, {})
            parsed = _parse_numeric_value(row.raw_value)
            field_label_code = normalize_code(_field_display_name(field)) if field else ""
            is_identity_like = any(
                marker in field_label_code
                for marker in {"don_vi", "ten_don_vi", "ma_don_vi", "stt", "so_thu_tu", "ma_so", "ky_hieu"}
            )
            is_numeric_field = bool(field and (field.data_type or "").lower() in numeric_types)
            is_editable_numeric_input = bool(field and field.is_editable and parsed is not None and not is_identity_like)
            if is_numeric_field or is_editable_numeric_input:
                if parsed is None:
                    sheet_values[cell_address] = row.raw_value or ""
                    continue
                total = numeric_totals.get((row.sheet_name, cell_address), 0.0) + parsed
                numeric_totals[(row.sheet_name, cell_address)] = total
                sheet_values[cell_address] = int(total) if float(total).is_integer() else total
            else:
                sheet_values[cell_address] = row.raw_value or ""

    return values


def _merged_submission_cell_values(submissions):
    values = {}
    for submission in submissions:
        for sheet_name, cells in _submission_cell_values(submission.id).items():
            values.setdefault(sheet_name, {}).update(cells)
    return values


def _resolve_working_submission_state(instance, template_version, report_type=None, report_date=None):
    if not instance or not template_version:
        return {
            "latest_submission": None,
            "history": [],
            "existing_values": {},
            "daily_submissions": [],
        }

    if report_type and report_type.code == "daily" and report_date:
        daily_submissions = _daily_snapshot_submissions_through_date(instance.id, report_date)
        latest_submission = daily_submissions[-1] if daily_submissions else None
        history = _submission_history_through_date(instance.id, report_date)
        existing_values = _cumulative_submission_cell_values(daily_submissions, template_version.id)
        return {
            "latest_submission": latest_submission,
            "history": history,
            "existing_values": existing_values,
            "daily_submissions": daily_submissions,
        }

    latest_submission = _latest_submission(instance.id)
    history = _submission_history(instance.id)
    existing_values = _submission_cell_values(latest_submission.id) if latest_submission else {}
    return {
        "latest_submission": latest_submission,
        "history": history,
        "existing_values": existing_values,
        "daily_submissions": [],
    }


def _resolve_entry_submission_state(instance, template_version, report_type=None, report_date=None):
    if not instance or not template_version:
        return {
            "latest_submission": None,
            "history": [],
            "existing_values": {},
        }

    if report_type and report_type.code == "daily" and report_date:
        latest_submission = _latest_submission_for_date(instance.id, report_date)
        history = _submission_history(instance.id, report_date=report_date)
        existing_values = _submission_cell_values(latest_submission.id) if latest_submission else {}
        return {
            "latest_submission": latest_submission,
            "history": history,
            "existing_values": existing_values,
        }

    latest_submission = _latest_submission(instance.id)
    history = _submission_history(instance.id)
    existing_values = _submission_cell_values(latest_submission.id) if latest_submission else {}
    return {
        "latest_submission": latest_submission,
        "history": history,
        "existing_values": existing_values,
    }


def _effective_daily_cutoff_date(cycle, instance_id=None):
    if cycle and cycle.close_at:
        return cycle.close_at.date()
    if cycle and cycle.due_at:
        return cycle.due_at.date()
    if instance_id:
        latest = _latest_submission(instance_id)
        latest_date = _submission_business_date(latest) if latest else None
        if latest_date:
            return latest_date
    return date.today()


def _resolve_effective_instance_state(instance, cycle, template_version, report_type=None):
    if not instance or not cycle or not template_version:
        return {
            "latest_submission": None,
            "history": [],
            "existing_values": {},
            "daily_submissions": [],
            "report_date": None,
            "mode": "empty",
        }

    if report_type and report_type.code == "daily":
        cutoff_date = _effective_daily_cutoff_date(cycle, instance.id)
        state = _resolve_working_submission_state(
            instance,
            template_version,
            report_type=report_type,
            report_date=cutoff_date,
        )
        state["report_date"] = cutoff_date
        state["mode"] = "daily_cumulative"
        return state

    state = _resolve_working_submission_state(
        instance,
        template_version,
        report_type=report_type,
        report_date=None,
    )
    state["report_date"] = None
    state["mode"] = "latest_snapshot"
    return state


def _export_effective_values(cycle, instance, template_version, effective_values, suffix="effective"):
    output_name = safe_filename(
        f"{template_version.template_id}_cycle_{cycle.id}_instance_{instance.id}_{suffix}.xlsx"
    )
    output_path = os.path.join(current_app.root_path, "report_exports", output_name)
    write_workbook_copy(template_version.source_path, output_path, effective_values)
    return output_path


def _materialize_effective_submission_export(
    instance,
    cycle,
    template_version,
    report_type=None,
    overwrite=False,
    apply=False,
):
    state = _resolve_effective_instance_state(
        instance,
        cycle,
        template_version,
        report_type=report_type,
    )
    latest_submission = state["latest_submission"]
    result = {
        "instance": instance,
        "state": state,
        "latest_submission": latest_submission,
        "has_data": bool(latest_submission),
        "has_existing_export": False,
        "needs_export": False,
        "exported": False,
        "export_path": "",
        "effective_time": (latest_submission.submitted_at or latest_submission.created_at) if latest_submission else None,
        "error": "",
    }
    if not latest_submission:
        return result

    processed_path = (latest_submission.processed_file_path or "").strip()
    has_existing_export = bool(processed_path and os.path.exists(processed_path))
    needs_export = overwrite or not has_existing_export
    result["has_existing_export"] = has_existing_export
    result["needs_export"] = needs_export
    result["export_path"] = processed_path if has_existing_export else ""

    if apply and needs_export:
        try:
            export_path = _export_effective_values(
                cycle,
                instance,
                template_version,
                state["existing_values"] or {},
                suffix=state["mode"],
            )
            latest_submission.processed_file_path = export_path
            result["exported"] = True
            result["export_path"] = export_path
        except Exception as exc:
            result["error"] = str(exc)
            return result

    if apply:
        effective_time = result["effective_time"]
        instance.status = latest_submission.status or instance.status or "draft"
        instance.submitted_at = effective_time
        if cycle and (cycle.is_locked or cycle.status == "closed"):
            instance.locked_at = cycle.close_at or instance.locked_at or effective_time or datetime.now()

    return result


def _materialize_cycle_effective_exports(cycle, overwrite=False, apply=False, include_open=False):
    summary = {
        "cycle_id": cycle.id if cycle else None,
        "instances_total": 0,
        "instances_with_data": 0,
        "instances_without_data": 0,
        "exports_generated": 0,
        "exports_reused": 0,
        "errors": [],
        "results": [],
    }
    if not cycle:
        return summary
    if not include_open and not (cycle.is_locked or cycle.status == "closed"):
        return summary

    template_version = db.session.get(ReportTemplateVersion, cycle.template_version_id)
    if not template_version:
        summary["errors"].append(f"cycle={cycle.id}: missing template version")
        return summary

    report_type = _report_type(cycle)
    instances = ReportInstance.query.filter_by(cycle_id=cycle.id).order_by(ReportInstance.id.asc()).all()
    summary["instances_total"] = len(instances)
    for instance in instances:
        result = _materialize_effective_submission_export(
            instance,
            cycle,
            template_version,
            report_type=report_type,
            overwrite=overwrite,
            apply=apply,
        )
        summary["results"].append(result)
        if result["has_data"]:
            summary["instances_with_data"] += 1
        else:
            summary["instances_without_data"] += 1
        if result["error"]:
            summary["errors"].append(
                f"cycle={cycle.id} instance={instance.id}: {result['error']}"
            )
            continue
        if result["exported"]:
            summary["exports_generated"] += 1
        elif result["has_existing_export"]:
            summary["exports_reused"] += 1

    return summary


def _submission_download_path(submission):
    if not submission:
        return ""
    processed_path = (submission.processed_file_path or "").strip()
    if processed_path and os.path.exists(processed_path):
        return processed_path
    file_path = (submission.file_path or "").strip()
    if file_path and os.path.exists(file_path):
        return file_path
    return ""


def _submission_cell_values(submission_id):
    values = {}
    rows = (
        ReportSubmissionCell.query.filter_by(submission_id=submission_id)
        .order_by(ReportSubmissionCell.id.asc())
        .all()
    )
    for row in rows:
        values.setdefault(row.sheet_name, {})[row.cell_address] = row.raw_value or ""
    return values


def _audit(action, object_type, object_id, details=""):
    try:
        db.session.add(
            ReportAuditLog(
                actor_user_id=session.get("uid"),
                action=action,
                module="report",
                object_type=object_type,
                object_id=object_id,
                details=details,
            )
        )
        db.session.commit()
    except Exception:
        db.session.rollback()
    try:
        log_action(session.get("uid"), session.get("fullname", ""), action, "Báo cáo", details)
    except Exception:
        pass


def _parse_dt(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except Exception:
        for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M"):
            try:
                return datetime.strptime(value, fmt)
            except Exception:
                continue
    return None


def _parse_date(value):
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except Exception:
            continue
    return None


def _start_of_day(value):
    return datetime.combine(value, time(0, 0, 0))


def _end_of_day(value):
    return datetime.combine(value, time(23, 59, 59))


def _next_day_start(value):
    return _start_of_day(value + timedelta(days=1))


def _submission_time_expr():
    return func.coalesce(ReportSubmission.submitted_at, ReportSubmission.created_at)


def _parse_positive_int(value, default=1):
    try:
        parsed = int(value)
    except Exception:
        parsed = default
    return parsed if parsed > 0 else default


def _parse_bounded_int(value, minimum, maximum, default=None):
    try:
        parsed = int(value)
    except Exception:
        return default
    if parsed < minimum or parsed > maximum:
        return default
    return parsed


def _template_for_cycle(cycle):
    template_version = db.session.get(ReportTemplateVersion, cycle.template_version_id) if cycle else None
    return db.session.get(ReportTemplate, template_version.template_id) if template_version else None


def _template_periodic_config(template):
    kind = (getattr(template, "periodic_cycle_type", "") or "").strip().lower()
    if kind not in PERIODIC_CYCLE_LABELS:
        kind = ""
    due_day = _parse_bounded_int(getattr(template, "periodic_due_day", None), 1, 31, None)
    due_month = _parse_bounded_int(getattr(template, "periodic_due_month", None), 1, 12, None)
    return {
        "type": kind,
        "label": PERIODIC_CYCLE_LABELS.get(kind, ""),
        "due_day": due_day,
        "due_month": due_month,
    }


def _periodic_schedule_rule_text(template):
    config = _template_periodic_config(template)
    kind = config["type"]
    due_day = config["due_day"]
    due_month = config["due_month"]
    if not kind or not due_day:
        return ""
    if kind == "monthly":
        return f"Định kỳ hằng tháng, hạn trước ngày {due_day}"
    if kind == "quarterly":
        return f"Định kỳ hằng quý, hạn trước ngày {due_day} của tháng cuối quý"
    if kind == "half_year" and due_month:
        return f"Định kỳ 6 tháng, hạn ngày {due_day} tháng {due_month}"
    if kind == "yearly" and due_month:
        return f"Định kỳ hằng năm, hạn ngày {due_day} tháng {due_month}"
    return ""


def _safe_calendar_date(year, month, day):
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, min(max(1, day), last_day))


def _next_month(year, month):
    if month == 12:
        return year + 1, 1
    return year, month + 1


def _next_quarter_end(year, month):
    quarter_end_month = ((month - 1) // 3 + 1) * 3
    return year, quarter_end_month


def _default_periodic_due_date(template, reference_date=None):
    config = _template_periodic_config(template)
    kind = config["type"]
    due_day = config["due_day"]
    due_month = config["due_month"]
    if not kind or not due_day:
        return None

    reference_date = reference_date or date.today()
    year = reference_date.year

    if kind == "monthly":
        candidate = _safe_calendar_date(year, reference_date.month, due_day)
        if candidate >= reference_date:
            return candidate
        next_year, next_month = _next_month(year, reference_date.month)
        return _safe_calendar_date(next_year, next_month, due_day)

    if kind == "quarterly":
        candidate_year, candidate_month = _next_quarter_end(year, reference_date.month)
        candidate = _safe_calendar_date(candidate_year, candidate_month, due_day)
        if candidate >= reference_date:
            return candidate
        next_month = candidate_month + 3
        next_year = candidate_year
        if next_month > 12:
            next_month -= 12
            next_year += 1
        return _safe_calendar_date(next_year, next_month, due_day)

    if kind in {"half_year", "yearly"} and due_month:
        candidate = _safe_calendar_date(year, due_month, due_day)
        if candidate >= reference_date:
            return candidate
        return _safe_calendar_date(year + 1, due_month, due_day)

    return None


def _periodic_form_defaults(template):
    config = _template_periodic_config(template)
    return {
        "type": config["type"] or "monthly",
        "due_day": config["due_day"] or "",
        "due_month": config["due_month"] or "",
        "rule_text": _periodic_schedule_rule_text(template),
    }


def _field_display_name(field):
    return (field.display_name or field.field_name or field.field_code or "").strip()


def _field_levels(field):
    parts = [part.strip() for part in str(field.path_code or field.field_name or field.field_code).split(" > ") if part.strip()]
    if parts:
        parts[-1] = _field_display_name(field)
    else:
        parts = [_field_display_name(field)]
    return parts


def _row_context_label(ws, sheet_fields, existing_values, sheet_name, row_index):
    preferred = []
    fallback = []
    for field in sheet_fields:
        coord = f"{field.column_letter}{row_index}"
        value = existing_values.get(sheet_name, {}).get(coord, ws[coord].value)
        display_value = _cell_display_value(value)
        if not display_value:
            continue
        label_code = normalize_code(_field_display_name(field))
        if any(marker in label_code for marker in {"don_vi", "ten_don_vi", "ma_don_vi", "stt", "so_thu_tu"}):
            preferred.append(display_value)
        elif not field.is_editable:
            fallback.append(display_value)
    if preferred:
        return " - ".join(preferred[:2])
    if fallback:
        return fallback[0]
    return f"Dòng {row_index}"


def _row_matches_unit(ws, sheet_fields, existing_values, sheet_name, row_index, unit):
    if not unit:
        return False
    unit_name = (getattr(unit, "name", None) or "").strip()
    unit_code = (getattr(unit, "code", None) or "").strip()
    if not unit_name and not unit_code:
        return False

    candidates = []
    for field in sheet_fields:
        coord = f"{field.column_letter}{row_index}"
        value = existing_values.get(sheet_name, {}).get(coord, ws[coord].value)
        display_value = _cell_display_value(value)
        if not display_value:
            continue
        label_code = normalize_code(_field_display_name(field))
        if any(marker in label_code for marker in {"don_vi", "ten_don_vi", "ma_don_vi", "stt", "so_thu_tu"}):
            candidates.append(display_value)
        elif not field.is_editable:
            candidates.append(display_value)

    for candidate in candidates:
        if unit_name and is_unit_match(candidate, unit_name):
            return True
        if unit_code and is_unit_match(candidate, unit_code):
            return True
    return False


def _cell_display_value(value):
    if value is None:
        return ""
    return str(value).strip()


def _form_checked(name):
    return "1" in request.form.getlist(name)


def _sheet_header_range(sheet_or_meta):
    if isinstance(sheet_or_meta, dict):
        start_row = int(sheet_or_meta.get("header_start_row") or sheet_or_meta.get("min_row") or 1)
        end_row = int(
            sheet_or_meta.get("header_end_row")
            or (start_row + max(int(sheet_or_meta.get("header_rows") or 1) - 1, 0))
        )
    else:
        start_row = int(getattr(sheet_or_meta, "header_start_row", 0) or 1)
        end_row = int(
            getattr(sheet_or_meta, "header_end_row", 0)
            or (start_row + max(int(getattr(sheet_or_meta, "header_rows", 1) or 1) - 1, 0))
        )
    if end_row < start_row:
        end_row = start_row
    return start_row, end_row


def _preferred_sticky_column(sheet_meta, ws, editable_values=None):
    editable_values = editable_values or {}
    fields = sheet_meta.get("fields") or []

    def _field_label(field):
        return normalize_code(
            field.get("field_name")
            or field.get("path_code")
            or field.get("field_code")
            or ""
        )

    for field in fields:
        label_code = _field_label(field)
        if any(marker in label_code for marker in {"don_vi", "ten_don_vi", "ma_don_vi"}):
            return int(field.get("column_index") or 0) or None

    unit_start_row = int(sheet_meta.get("unit_start_row") or sheet_meta.get("data_start_row") or 1)
    unit_end_row = int(sheet_meta.get("unit_end_row") or sheet_meta.get("data_end_row") or unit_start_row)
    start_col = column_index_from_string(sheet_meta.get("start_column") or "A")
    end_col = column_index_from_string(sheet_meta.get("end_column") or sheet_meta.get("start_column") or "A")

    for col_idx in range(start_col, end_col + 1):
        header_code = ""
        for field in fields:
            if int(field.get("column_index") or 0) == col_idx:
                header_code = _field_label(field)
                break
        if any(marker in header_code for marker in {"stt", "so_thu_tu", "thu_tu"}):
            continue
        text_hits = 0
        numeric_like = 0
        for row_idx in range(unit_start_row, unit_end_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = editable_values.get(cell.coordinate, cell.value)
            display = _cell_display_value(value)
            if not display:
                continue
            if display.isdigit():
                numeric_like += 1
                continue
            text_hits += 1
        if text_hits > 0 and text_hits >= numeric_like:
            return col_idx

    return start_col


def _cycle_units(cycle):
    return _resolve_scope_units(_parse_scope_payload(cycle.scope_json))


def _resolve_cycle_unit(cycle):
    if _is_admin():
        requested_unit_id = int(request.values.get("unit_id") or 0)
        if requested_unit_id:
            unit = db.session.get(ReportUnit, requested_unit_id)
            if unit and _cycle_accessible(cycle, unit.id, True):
                return unit
        scoped_units = _cycle_units(cycle)
        if scoped_units:
            return scoped_units[0]
    return _current_user_report_unit()


def _purge_submission(submission):
    if submission.file_path and os.path.exists(submission.file_path):
        try:
            os.remove(submission.file_path)
        except Exception:
            pass
    ReportValidationLog.query.filter_by(submission_id=submission.id).delete(synchronize_session=False)
    ReportExportJob.query.filter_by(submission_id=submission.id).delete(synchronize_session=False)
    ReportSubmissionValue.query.filter_by(submission_id=submission.id).delete(synchronize_session=False)
    ReportSubmissionCell.query.filter_by(submission_id=submission.id).delete(synchronize_session=False)
    db.session.delete(submission)


def _refresh_instance_status(instance):
    latest = _latest_submission(instance.id)
    if latest:
        instance.status = latest.status or "draft"
        instance.submitted_at = latest.submitted_at
        instance.locked_at = latest.submitted_at if latest.status == "submitted" else None
    else:
        instance.status = "draft"
        instance.submitted_at = None
        instance.locked_at = None


def _report_type(cycle):
    return db.session.get(ReportType, cycle.report_type_id) if cycle and cycle.report_type_id else None


def _submission_timeliness(cycle, submission, report_type=None):
    if not submission:
        return "Chưa nộp"
    report_type = report_type or _report_type(cycle)
    submitted_at = submission.submitted_at or submission.created_at
    if not submitted_at:
        return "Đã lưu"
    report_code = report_type.code if report_type else ""
    if report_code == "daily":
        business_date = _submission_business_date(submission)
        if cycle:
            report_day = (cycle.open_at or cycle.created_at or submitted_at).date()
        else:
            report_day = business_date or submitted_at.date()
        if business_date == report_day:
            return "Đúng ngày"
        if business_date:
            return f"Báo cáo ngày {business_date.strftime('%d/%m/%Y')}"
        return f"Báo cáo ngày {submitted_at.strftime('%d/%m/%Y')}"
    if cycle and cycle.due_at:
        return "Đúng hạn" if submitted_at <= cycle.due_at else "Quá hạn"
    return "Đã báo cáo"


def _report_schedule_text(cycle, report_type=None, report_date=None):
    report_type = report_type or _report_type(cycle)
    if not cycle:
        return ""
    report_code = report_type.code if report_type else ""
    if report_code == "daily":
        target_day = report_date or (cycle.open_at or cycle.due_at or cycle.created_at)
        return f"Báo cáo ngày {target_day.strftime('%d/%m/%Y')}" if target_day else "Báo cáo hằng ngày"
    if report_code == "periodic":
        template = _template_for_cycle(cycle)
        rule_text = _periodic_schedule_rule_text(template)
        if rule_text and cycle.due_at:
            return f"{rule_text} • Hạn kỳ hiện tại: {cycle.due_at.strftime('%d/%m/%Y')}"
        if rule_text:
            return rule_text
    if cycle.due_at:
        return f"Hạn báo cáo: {cycle.due_at.strftime('%d/%m/%Y')}"
    return "Không đặt hạn cụ thể"


def _period_dates_for_cycle(cycle, report_type=None):
    report_type = report_type or _report_type(cycle)
    start_date = (cycle.open_at or cycle.created_at or datetime.now()).date()
    end_date = (cycle.due_at or cycle.open_at or cycle.created_at or datetime.now()).date()
    if report_type and report_type.code == "daily":
        end_date = start_date
    return start_date, end_date


def _ensure_reporting_period(cycle, report_type=None):
    report_type = report_type or _report_type(cycle)
    period = db.session.get(ReportingPeriod, cycle.legacy_period_id) if cycle.legacy_period_id else None
    start_date, end_date = _period_dates_for_cycle(cycle, report_type=report_type)
    period_type = report_type.code if report_type else "periodic"
    if not period:
        code = f"report_{cycle.id}_{start_date.strftime('%Y%m%d')}"
        period = ReportingPeriod(
            template_id=db.session.get(ReportTemplateVersion, cycle.template_version_id).template_id,
            code=code,
            name=cycle.name[:100],
            period_type=period_type,
            is_adhoc=(period_type == "ad_hoc"),
            start_date=start_date,
            end_date=end_date,
            deadline=cycle.due_at,
            is_locked=cycle.is_locked,
            created_by=session.get("uid"),
        )
        db.session.add(period)
        db.session.flush()
        cycle.legacy_period_id = period.id
        return period
    period.template_id = db.session.get(ReportTemplateVersion, cycle.template_version_id).template_id
    period.name = cycle.name[:100]
    period.period_type = period_type
    period.is_adhoc = period_type == "ad_hoc"
    period.start_date = start_date
    period.end_date = end_date
    period.deadline = cycle.due_at
    period.is_locked = cycle.is_locked
    return period


def _build_cycle_name(template, report_type):
    today = datetime.now().strftime("%d/%m/%Y")
    if report_type and report_type.code == "daily":
        return f"{template.name} - ngày {today}"
    return template.name


def _ensure_cycle_instances(cycle):
    unit_ids = [unit.id for unit in _cycle_units(cycle)]
    if not unit_ids:
        return
    existing_unit_ids = {instance.report_unit_id for instance in ReportInstance.query.filter_by(cycle_id=cycle.id).all()}
    template_version = db.session.get(ReportTemplateVersion, cycle.template_version_id)
    period = _ensure_reporting_period(cycle)
    for unit_id in unit_ids:
        if unit_id in existing_unit_ids:
            continue
        unit = db.session.get(ReportUnit, unit_id)
        user = None
        if unit:
            user = (
                User.query.filter(
                    or_(User.unit_key == unit.code, User.unit_area == unit.name, User.fullname == unit.name)
                )
                .order_by(User.id.asc())
                .first()
            )
        db.session.add(
            ReportInstance(
                template_id=template_version.template_id if template_version else None,
                version_id=template_version.id if template_version else None,
                period_id=period.id if period else None,
                user_id=user.id if user else None,
                org_unit=unit.name if unit else "",
                cycle_id=cycle.id,
                report_unit_id=unit_id,
                assigned_user_id=user.id if user else None,
                status="draft",
            )
        )


def _submission_status_label(submission):
    if not submission:
        return ""
    if submission.status == "submitted":
        return "Đã gửi báo cáo"
    if submission.status == "draft":
        return "Lưu nháp"
    return submission.status or ""


def _dashboard_cycle_progress(cycle, report_type=None, report_date=None):
    if not cycle:
        return {"reported": 0, "total": 0}
    report_type = report_type or _report_type(cycle)
    report_date = report_date or date.today()
    instances = ReportInstance.query.filter_by(cycle_id=cycle.id).order_by(ReportInstance.id.asc()).all()
    total_units = len(instances) or len(_cycle_units(cycle))
    reported = 0
    for instance in instances:
        if report_type and report_type.code == "daily":
            if _latest_submission_for_date(instance.id, report_date):
                reported += 1
        else:
            if _latest_submission(instance.id):
                reported += 1
    return {"reported": reported, "total": total_units}


def _dashboard_cycle_status(cycle, instance=None, report_type=None, report_date=None, progress=None):
    report_type = report_type or _report_type(cycle)
    report_date = report_date or date.today()
    is_daily = bool(report_type and report_type.code == "daily")
    if instance:
        submission = _latest_submission_for_date(instance.id, report_date) if is_daily else _latest_submission(instance.id)
        if is_daily:
            return {"label": "Hôm nay đã báo cáo" if submission else "Hôm nay chưa báo cáo", "done": bool(submission)}
        return {"label": "Đã báo cáo" if submission else "Chưa báo cáo", "done": bool(submission)}

    progress = progress or _dashboard_cycle_progress(cycle, report_type=report_type, report_date=report_date)
    has_any_report = bool(progress["reported"])
    if is_daily:
        return {"label": "Hôm nay đã có báo cáo" if has_any_report else "Hôm nay chưa báo cáo", "done": has_any_report}
    return {"label": "Đã có báo cáo" if has_any_report else "Chưa báo cáo", "done": has_any_report}


def _dashboard_hero_stats(cycles, cycle_status_map=None):
    now = datetime.now()
    active_total = 0
    overdue_total = 0
    locked_total = 0
    reported_total = 0

    for cycle in cycles or []:
        is_locked = bool(cycle and (cycle.is_locked or cycle.status == "closed"))
        if is_locked:
            locked_total += 1
        else:
            active_total += 1
            if cycle and cycle.due_at and cycle.due_at < now:
                overdue_total += 1

        status_info = (cycle_status_map or {}).get(getattr(cycle, "id", None))
        if status_info and status_info.get("done"):
            reported_total += 1

    return {
        "active_total": active_total,
        "overdue_total": overdue_total,
        "locked_total": locked_total,
        "reported_total": reported_total,
    }


def _build_submission_summary(submission, template_version, workbook=None, field_lookup=None, sheet_fields=None, limit=8):
    if not submission or not template_version:
        return {"text": "", "count": 0}

    if workbook is None:
        workbook = load_workbook(template_version.source_path, data_only=False)

    if field_lookup is None or sheet_fields is None:
        fields = ReportTemplateField.query.filter_by(version_id=template_version.id).order_by(ReportTemplateField.display_order.asc()).all()
        field_lookup = {(field.sheet_name, field.field_code): field for field in fields}
        sheet_fields = {}
        for field in fields:
            sheet_fields.setdefault(field.sheet_name, []).append(field)

    existing_values = _submission_cell_values(submission.id)
    value_rows = (
        ReportSubmissionValue.query.filter_by(submission_id=submission.id)
        .order_by(ReportSubmissionValue.id.asc())
        .all()
    )
    if not value_rows:
        return {"text": "", "count": 0}

    row_title_cache = {}
    summary_entries = []
    for value_row in value_rows:
        value_text = (value_row.value_text or "").strip()
        if not value_text and value_row.value_number is None:
            continue
        cell_address = (value_row.cell_address or "").strip().upper()
        row_digits = "".join(ch for ch in cell_address if ch.isdigit())
        row_index = int(row_digits) if row_digits.isdigit() else 0
        sheet_name = value_row.sheet_name
        field = field_lookup.get((sheet_name, value_row.field_code))
        field_label = _field_display_name(field) if field else (value_row.field_code or cell_address)
        row_title = ""
        if row_index and sheet_name in workbook.sheetnames:
            cache_key = (sheet_name, row_index)
            if cache_key not in row_title_cache:
                row_title_cache[cache_key] = _row_context_label(
                    workbook[sheet_name],
                    sheet_fields.get(sheet_name, []),
                    existing_values,
                    sheet_name,
                    row_index,
                )
            row_title = row_title_cache[cache_key]
        label = f"{row_title} - {field_label}" if row_title else field_label
        summary_entries.append(f"{label}: {value_text or value_row.value_number}")

    if not summary_entries:
        return {"text": "", "count": 0}

    visible_entries = summary_entries[:limit]
    text = "; ".join(visible_entries)
    if len(summary_entries) > limit:
        text += f"; ... (+{len(summary_entries) - limit})"
    return {"text": text, "count": len(summary_entries)}


def _history_rows_for_submissions(submissions, template_version, report_type, include_unit=False, include_actor=False, cycle=None):
    if not submissions or not template_version:
        return []

    workbook = load_workbook(template_version.source_path, data_only=False)
    fields = (
        ReportTemplateField.query.filter_by(version_id=template_version.id)
        .order_by(ReportTemplateField.display_order.asc())
        .all()
    )
    field_lookup = {(field.sheet_name, field.field_code): field for field in fields}
    sheet_fields = {}
    for field in fields:
        sheet_fields.setdefault(field.sheet_name, []).append(field)

    rows = []
    user_ids = {submission.submitted_by for submission in submissions if submission.submitted_by}
    user_map = {user.id: user for user in User.query.filter(User.id.in_(user_ids)).all()} if user_ids else {}
    instance_ids = {submission.instance_id for submission in submissions if submission.instance_id}
    instance_map = {instance.id: instance for instance in ReportInstance.query.filter(ReportInstance.id.in_(instance_ids)).all()} if instance_ids else {}
    unit_ids = {instance.report_unit_id for instance in instance_map.values() if instance.report_unit_id}
    unit_map = {unit.id: unit for unit in ReportUnit.query.filter(ReportUnit.id.in_(unit_ids)).all()} if unit_ids else {}

    for submission in submissions:
        summary = _build_submission_summary(
            submission,
            template_version,
            workbook=workbook,
            field_lookup=field_lookup,
            sheet_fields=sheet_fields,
        )
        row = {
            "submission": submission,
            "submitted_at": submission.submitted_at or submission.created_at,
            "content_text": summary["text"],
            "content_count": summary["count"],
            "status_label": _submission_status_label(submission),
            "timeliness": _submission_timeliness(cycle, submission, report_type=report_type),
        }
        if include_actor:
            row["actor"] = user_map.get(submission.submitted_by)
        if include_unit:
            instance = instance_map.get(submission.instance_id)
            row["unit"] = unit_map.get(instance.report_unit_id) if instance and instance.report_unit_id else None
            row["instance"] = instance
        rows.append(row)
    return rows


def _ensure_active_cycle_for_template(template, version, report_type):
    scope_json = template.assignment_scope_json or json.dumps(_empty_scope_payload(), ensure_ascii=False)
    version_ids = _template_version_ids(template.id)
    open_cycles = ReportCycle.query.filter(
        ReportCycle.template_version_id.in_(version_ids),
        ReportCycle.status != "closed",
    ).order_by(ReportCycle.created_at.desc()).all() if version_ids else []

    selected_cycle = None
    for cycle in open_cycles:
        has_submissions = (
            db.session.query(ReportSubmission.id)
            .join(ReportInstance, ReportSubmission.instance_id == ReportInstance.id)
            .filter(ReportInstance.cycle_id == cycle.id)
            .first()
        )
        if not has_submissions:
            selected_cycle = cycle
            break

    if not selected_cycle and open_cycles:
        selected_cycle = open_cycles[0]

    if not selected_cycle:
        open_at = datetime.now()
        if report_type and report_type.code == "daily":
            due_at = _end_of_day(datetime.now().date())
        elif report_type and report_type.code == "periodic":
            periodic_due_date = _default_periodic_due_date(template)
            due_at = _end_of_day(periodic_due_date) if periodic_due_date else None
        else:
            due_at = None
        selected_cycle = ReportCycle(
            template_version_id=version.id,
            report_type_id=report_type.id,
            name=_build_cycle_name(template, report_type),
            open_at=open_at,
            close_at=None,
            due_at=due_at,
            auto_lock_at=None,
            status="open",
            scope_json=scope_json,
            is_locked=False,
            note="",
        )
        db.session.add(selected_cycle)
        db.session.flush()
    else:
        selected_cycle.template_version_id = version.id
        selected_cycle.report_type_id = report_type.id
        selected_cycle.scope_json = scope_json
        if report_type and report_type.code == "daily":
            selected_cycle.open_at = _start_of_day(datetime.now().date())
            selected_cycle.due_at = _end_of_day(datetime.now().date())
            selected_cycle.name = _build_cycle_name(template, report_type)
        elif report_type and report_type.code == "periodic":
            periodic_due_date = _default_periodic_due_date(template)
            if periodic_due_date:
                selected_cycle.due_at = _end_of_day(periodic_due_date)
            selected_cycle.name = template.name
        else:
            selected_cycle.name = template.name

    _ensure_reporting_period(selected_cycle, report_type=report_type)
    _ensure_cycle_instances(selected_cycle)
    db.session.commit()
    return selected_cycle


def _ensure_active_cycles_for_ready_templates():
    templates = (
        ReportTemplate.query.filter_by(status="active")
        .order_by(ReportTemplate.updated_at.desc(), ReportTemplate.id.desc())
        .all()
    )
    for template in templates:
        version = _template_version(template)
        if not version or not template.report_type_id:
            continue
        report_type = db.session.get(ReportType, template.report_type_id)
        if not report_type:
            continue
        _ensure_active_cycle_for_template(template, version, report_type)


def _group_admin_templates(templates):
    groups = {}
    for template in templates:
        group_name = _template_professional_unit(template)
        groups.setdefault(group_name, []).append(
            {
                "template": template,
                "current_version": _template_version(template),
                "current_cycle": _template_current_cycle(template),
                "report_type": db.session.get(ReportType, template.report_type_id) if template.report_type_id else None,
            }
        )
    return [
        {"name": name, "entries": groups[name]}
        for name in sorted(groups.keys(), key=lambda value: (value == "Chưa phân đội", value.lower()))
    ]


def _group_cycles_by_professional_unit(cycles):
    groups = {}
    for cycle in cycles:
        template_version = db.session.get(ReportTemplateVersion, cycle.template_version_id)
        template = db.session.get(ReportTemplate, template_version.template_id) if template_version else None
        group_name = _template_professional_unit(template)
        groups.setdefault(group_name, []).append(
            {
                "cycle": cycle,
                "template": template,
                "template_version": template_version,
                "report_type": _report_type(cycle),
            }
        )
    return [
        {"name": name, "entries": groups[name]}
        for name in sorted(groups.keys(), key=lambda value: (value == "Chưa phân đội", value.lower()))
    ]


def _save_submission(instance, payload, final_submit=False, report_date=None):
    cycle = db.session.get(ReportCycle, instance.cycle_id)
    template_version = db.session.get(ReportTemplateVersion, cycle.template_version_id)
    report_type = _report_type(cycle)
    period = _ensure_reporting_period(cycle, report_type=report_type)
    report_unit = db.session.get(ReportUnit, instance.report_unit_id) if instance.report_unit_id else None
    sheet_meta = json.loads(template_version.metadata_json or "{}").get("sheets", [])
    sheets = {sheet["sheet_name"]: sheet for sheet in sheet_meta}

    payload_sheets = payload.get("sheets", {}) if isinstance(payload, dict) else {}
    errors = []
    normalized = {}

    existing_version = (
        ReportSubmission.query.filter_by(instance_id=instance.id)
        .count()
    )
    metadata_payload = {"note": payload.get("note", "") if isinstance(payload, dict) else ""}
    if report_type and report_type.code == "daily" and report_date:
        metadata_payload["report_date"] = report_date.strftime("%Y-%m-%d")

    submission = ReportSubmission(
        template_id=template_version.template_id,
        template_version_id=template_version.id,
        period_id=period.id if period else cycle.id,
        report_period=cycle.name,
        reporting_unit=report_unit.name if report_unit else (session.get("unit_area") or ""),
        submitted_by=session.get("uid") or instance.assigned_user_id or instance.user_id or 0,
        instance_id=instance.id,
        version_no=existing_version + 1,
        status="submitted" if final_submit else "draft",
        original_filename=template_version.source_filename,
        original_file_path=template_version.source_path,
        processed_file_path="",
        error_file_path="",
        total_rows=0,
        valid_rows=0,
        invalid_rows=0,
        warning_count=0,
        metadata_json=json.dumps(metadata_payload, ensure_ascii=False),
        note=payload.get("note", "") if isinstance(payload, dict) else "",
        submitted_at=datetime.now() if final_submit else None,
    )
    db.session.add(submission)
    db.session.flush()

    for sheet_name, cells in payload_sheets.items():
        if sheet_name not in sheets:
            continue
        sheet_fields = ReportTemplateField.query.filter_by(
            version_id=template_version.id,
            sheet_name=sheet_name,
        ).all()
        field_by_col = {f.column_index: f for f in sheet_fields}
        sheet_values = cells if isinstance(cells, dict) else {}
        for cell_address, raw_value in sheet_values.items():
            col_letters = "".join(ch for ch in cell_address if ch.isalpha())
            row_digits = "".join(ch for ch in cell_address if ch.isdigit())
            if not col_letters or not row_digits:
                continue
            column_index = column_index_from_string(col_letters)
            field = field_by_col.get(column_index)
            field_code = field.field_code if field else ""
            text_value = "" if raw_value is None else str(raw_value).strip()
            number_value = None
            if field and field.data_type in {"number", "float", "decimal"}:
                try:
                    number_value = float(text_value.replace(",", ""))
                except Exception:
                    number_value = None
            db.session.add(
                ReportSubmissionCell(
                    submission_id=submission.id,
                    sheet_name=sheet_name,
                    cell_address=cell_address,
                    raw_value=text_value,
                    is_formula=False,
                    formula_text="",
                )
            )
            if field_code:
                db.session.add(
                    ReportSubmissionValue(
                        submission_id=submission.id,
                        sheet_name=sheet_name,
                        field_code=field_code,
                        cell_address=cell_address,
                        value_text=text_value,
                        value_number=number_value,
                        value_json=json.dumps({"cell": cell_address, "value": text_value}, ensure_ascii=False),
                    )
                )
        for field in sheet_fields:
            if field.is_required:
                has_value = any(v.field_code == field.field_code and (v.value_text or v.value_number is not None) for v in ReportSubmissionValue.query.filter_by(submission_id=submission.id).all())
                if not has_value:
                    field_label = _field_display_name(field)
                    errors.append((sheet_name, field_label, "Trường bắt buộc chưa có dữ liệu"))
                    db.session.add(
                        ReportValidationLog(
                            submission_id=submission.id,
                            sheet_name=sheet_name,
                            field_code=field.field_code,
                            cell_address="",
                            severity="error",
                            message=f"Trường '{field_label}' chưa có dữ liệu",
                        )
                    )

    if errors and final_submit:
        submission.status = "draft"
        instance.status = "draft"
        db.session.commit()
        return submission, errors

    if final_submit:
        instance.status = "submitted"
        instance.submitted_at = datetime.now()
        instance.locked_at = datetime.now()
    else:
        instance.status = "draft"
        instance.updated_at = datetime.now()

    db.session.commit()
    return submission, errors


def _export_submission(submission):
    instance = db.session.get(ReportInstance, submission.instance_id)
    cycle = db.session.get(ReportCycle, instance.cycle_id)
    version = db.session.get(ReportTemplateVersion, cycle.template_version_id)
    output_name = safe_filename(f"{version.template_id}_cycle_{cycle.id}_submission_{submission.id}.xlsx")
    output_path = os.path.join(current_app.root_path, "report_exports", output_name)
    values = _submission_cell_values(submission.id)
    write_workbook_copy(version.source_path, output_path, values)
    job = ReportExportJob(
        cycle_id=cycle.id,
        submission_id=submission.id,
        status="done",
        output_path=output_path,
        finished_at=datetime.now(),
    )
    db.session.add(job)
    db.session.commit()
    return output_path


@reporting_bp.route("/admin/reports")
def admin_dashboard():
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    _ensure_default_types()
    _sync_units_from_users()
    _ensure_active_cycles_for_ready_templates()

    templates = ReportTemplate.query.order_by(ReportTemplate.updated_at.desc()).all()
    versions = ReportTemplateVersion.query.order_by(ReportTemplateVersion.created_at.desc()).all()
    cycles = ReportCycle.query.order_by(ReportCycle.created_at.desc()).all()
    units = ReportUnit.query.filter_by(is_active=True).order_by(ReportUnit.name.asc()).all()
    report_types = ReportType.query.order_by(ReportType.name.asc()).all()
    recent_submissions = ReportSubmission.query.order_by(ReportSubmission.created_at.desc()).limit(20).all()

    current_versions = {}
    for template in templates:
        current_version = _template_version(template)
        current_versions[template.id] = current_version
    template_report_types = {
        template.id: (db.session.get(ReportType, template.report_type_id) if template.report_type_id else None)
        for template in templates
    }
    cycle_status_map = {}
    cycle_progress_map = {}
    for cycle in cycles:
        report_type = _report_type(cycle)
        progress = _dashboard_cycle_progress(cycle, report_type=report_type, report_date=date.today())
        cycle_progress_map[cycle.id] = progress
        cycle_status_map[cycle.id] = _dashboard_cycle_status(
            cycle,
            instance=None,
            report_type=report_type,
            report_date=date.today(),
            progress=progress,
        )
    hero_stats = _dashboard_hero_stats(cycles, cycle_status_map=cycle_status_map)

    return render_template(
        "reporting_dashboard.html",
        templates=templates,
        template_groups=_group_admin_templates(templates),
        versions=versions,
        cycles=cycles,
        units=units,
        report_types=report_types,
        professional_units=_professional_unit_options(),
        recent_submissions=recent_submissions,
        current_versions=current_versions,
        template_report_types=template_report_types,
        cycle_status_map=cycle_status_map,
        cycle_progress_map=cycle_progress_map,
        hero_stats=hero_stats,
        can_view_cycle_progress=True,
        is_admin=True,
    )


@reporting_bp.route("/admin/reports/upload", methods=["POST"])
def upload_template():
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()

    file = request.files.get("template_file")
    if not file or not file.filename.lower().endswith(".xlsx"):
        flash("Chỉ nhận file .xlsx", "danger")
        return redirect(url_for("reporting_bp.admin_dashboard"))

    name = (request.form.get("name") or file.filename.rsplit(".", 1)[0]).strip()
    code = normalize_code(name or file.filename.rsplit(".", 1)[0]) or f"template_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    description = (request.form.get("description") or "").strip()
    professional_unit = (request.form.get("professional_unit") or "").strip()
    report_type_id = int(request.form.get("report_type_id") or 0)
    report_type = db.session.get(ReportType, report_type_id)
    if not report_type:
        flash("Bạn cần chọn loại báo cáo.", "danger")
        return redirect(url_for("reporting_bp.admin_dashboard"))
    parse_options = _build_parse_options(request.form)
    notes = description

    template = ReportTemplate.query.filter_by(code=code).first()
    if not template:
        template = ReportTemplate(
            code=code,
            name=name,
            description=description,
            status="draft",
            assignment_scope_json=json.dumps(_empty_scope_payload(), ensure_ascii=False),
        )
        db.session.add(template)
        db.session.flush()
    else:
        template.name = name
        template.description = description
        template.status = "draft"
        db.session.flush()

    template.report_type_id = report_type.id
    if professional_unit:
        template.professional_unit = professional_unit

    directive_file = request.files.get("directive_file")
    directive_filename, directive_path = _save_directive_file(directive_file, code)
    if directive_filename and directive_path:
        if template.directive_path and os.path.exists(template.directive_path):
            try:
                os.remove(template.directive_path)
            except Exception:
                pass
        template.directive_filename = directive_filename
        template.directive_path = directive_path

    existing_versions = ReportTemplateVersion.query.filter_by(template_id=template.id).count()
    version_no = existing_versions + 1
    filename = safe_filename(f"{code}_v{version_no}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
    storage_path = os.path.join(current_app.root_path, "report_templates", filename)
    file.save(storage_path)
    for version in ReportTemplateVersion.query.filter_by(template_id=template.id).all():
        version.is_current = False

    version = ReportTemplateVersion(
        template_id=template.id,
        version_no=version_no,
        source_filename=secure_filename(file.filename),
        source_path=storage_path,
        metadata_json="{}",
        notes=notes,
        is_current=True,
    )
    db.session.add(version)
    db.session.flush()

    metadata = _apply_version_structure(
        version,
        storage_path,
        secure_filename(file.filename),
        parse_options,
        notes=notes,
    )

    db.session.commit()
    _audit("upload_template", "report_template", template.id, name)
    flash("Đã tạo biểu mẫu. Tiếp theo cấu hình chi tiết cho từng sheet.", "success")
    return redirect(url_for("reporting_bp.template_settings", template_id=template.id))


@reporting_bp.route("/admin/reports/templates/<int:template_id>")
def template_detail(template_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    template = db.session.get(ReportTemplate, template_id)
    if not template:
        return "Not Found", 404
    current_version = _template_version(template)
    fields = []
    field_rows = []
    header_level_count = 1
    if current_version:
        fields = ReportTemplateField.query.filter_by(version_id=current_version.id).order_by(ReportTemplateField.sheet_name.asc(), ReportTemplateField.display_order.asc()).all()
        for field in fields:
            raw_levels = [part.strip() for part in str(field.path_code or field.field_name or "").split(" > ") if part.strip()]
            if not raw_levels:
                raw_levels = [field.field_name or field.field_code]
            if field.display_name:
                if raw_levels:
                    raw_levels[-1] = field.display_name
                else:
                    raw_levels = [field.display_name]
            header_level_count = max(header_level_count, len(raw_levels))
            field_rows.append({
                "field": field,
                "levels": raw_levels,
            })
    return render_template(
        "reporting_template_detail.html",
        template=template,
        current_version=current_version,
        fields=fields,
        field_rows=field_rows,
        header_level_count=header_level_count,
    )


@reporting_bp.route("/admin/reports/templates/<int:template_id>/settings")
def template_settings(template_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    template = db.session.get(ReportTemplate, template_id)
    if not template:
        return "Not Found", 404
    current_version = _template_version(template)
    current_report_type = db.session.get(ReportType, template.report_type_id) if template and template.report_type_id else None
    return render_template(
        "reporting_template_settings.html",
        template=template,
        current_version=current_version,
        current_report_type=current_report_type,
        report_types=ReportType.query.order_by(ReportType.name.asc()).all(),
        professional_units=_professional_unit_options(),
        roles=AppRole.query.order_by(AppRole.name.asc()).all(),
        units=ReportUnit.query.filter_by(is_active=True).order_by(ReportUnit.name.asc()).all(),
        unit_groups=_unit_group_options(),
        scope_payload=_parse_scope_payload(template.assignment_scope_json),
        range_defaults=_template_range_defaults(current_version),
        sheet_range_defaults=_template_sheet_range_defaults(current_version),
        periodic_config=_periodic_form_defaults(template),
        periodic_cycle_options=PERIODIC_CYCLE_OPTIONS,
    )


@reporting_bp.route("/admin/reports/templates/<int:template_id>/settings", methods=["POST"])
def save_template_settings(template_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    template = db.session.get(ReportTemplate, template_id)
    if not template:
        return "Not Found", 404
    version = _template_version(template)
    if not version:
        flash("Mẫu này chưa có file cấu hình.", "warning")
        return redirect(url_for("reporting_bp.admin_dashboard"))

    report_type_id = int(request.form.get("report_type_id") or 0)
    report_type = db.session.get(ReportType, report_type_id)
    if not report_type:
        flash("Bạn cần chọn loại báo cáo.", "danger")
        return redirect(url_for("reporting_bp.template_settings", template_id=template.id))

    professional_unit = (request.form.get("professional_unit") or "").strip()
    if not professional_unit:
        flash("Bạn cần chọn đội nghiệp vụ.", "danger")
        return redirect(url_for("reporting_bp.template_settings", template_id=template.id))

    template.description = (request.form.get("description") or "").strip()
    template.report_type_id = report_type.id
    template.professional_unit = professional_unit
    template.assignment_scope_json = json.dumps(_scope_payload_from_form(request.form), ensure_ascii=False)
    periodic_error = _save_template_periodic_config(template, report_type, request.form)
    if periodic_error:
        flash(periodic_error, "danger")
        return redirect(url_for("reporting_bp.template_settings", template_id=template.id))

    directive_file = request.files.get("directive_file")
    directive_filename, directive_path = _save_directive_file(directive_file, template.code or template.name or "report")
    if directive_filename and directive_path:
        if template.directive_path and os.path.exists(template.directive_path):
            try:
                os.remove(template.directive_path)
            except Exception:
                pass
        template.directive_filename = directive_filename
        template.directive_path = directive_path

    old_source_path = version.source_path
    new_upload = request.files.get("template_file")
    source_path = version.source_path
    source_filename = version.source_filename
    if new_upload and new_upload.filename:
        if not new_upload.filename.lower().endswith(".xlsx"):
            flash("Chỉ nhận file .xlsx", "danger")
            return redirect(url_for("reporting_bp.template_settings", template_id=template.id))
        filename = safe_filename(
            f"{template.code or normalize_code(template.name) or 'report'}_current_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        )
        source_path = os.path.join(current_app.root_path, "report_templates", filename)
        new_upload.save(source_path)
        source_filename = secure_filename(new_upload.filename)

    defaults = _template_range_defaults(version)
    parse_options = _build_parse_options(request.form, defaults=defaults)
    source_sheet_names = _workbook_sheet_names(source_path)
    sheet_parse_options = _build_sheet_parse_options(
        request.form,
        defaults=_template_sheet_range_defaults(version),
        workbook_sheet_names=source_sheet_names,
        global_defaults=parse_options,
    )
    _apply_version_structure(
        version,
        source_path,
        source_filename,
        parse_options,
        notes=template.description or "",
        sheet_parse_options=sheet_parse_options,
    )

    if source_path != old_source_path and old_source_path and os.path.exists(old_source_path):
        try:
            os.remove(old_source_path)
        except Exception:
            pass

    template.status = "active"
    db.session.commit()
    _sync_units_from_users()
    active_cycle = _ensure_active_cycle_for_template(template, version, report_type)
    _audit("save_template_settings", "report_template", template.id, template.name)
    flash("Đã lưu cấu hình biểu mẫu.", "success")
    return redirect(url_for("reporting_bp.admin_cycle_detail", cycle_id=active_cycle.id))


@reporting_bp.route("/admin/reports/templates/<int:template_id>/save-config", methods=["POST"])
def save_template_config(template_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    template = db.session.get(ReportTemplate, template_id)
    if not template:
        return "Not Found", 404
    version = _template_version(template)
    if not version:
        flash("Mẫu này chưa có file cấu hình.", "warning")
        return redirect(url_for("reporting_bp.template_detail", template_id=template_id))

    fields = ReportTemplateField.query.filter_by(version_id=version.id).order_by(ReportTemplateField.sheet_name.asc(), ReportTemplateField.display_order.asc()).all()
    for field in fields:
        levels = []
        for idx in range(1, 21):
            value = (request.form.get(f"field_{field.id}_level_{idx}") or "").strip()
            if value:
                levels.append(value)
        if levels:
            field.display_name = levels[-1]
            field.path_code = " > ".join(levels)
        else:
            field.display_name = field.display_name or field.field_name or field.field_code
        field.is_visible = _form_checked(f"field_{field.id}_visible")
        field.is_editable = field.is_visible and _form_checked(f"field_{field.id}_editable")

    db.session.commit()
    _audit("save_template_config", "report_template", template.id, template.name)
    flash("Đã lưu cấu hình trường dữ liệu.", "success")
    return redirect(url_for("reporting_bp.template_detail", template_id=template.id))


@reporting_bp.route("/admin/reports/templates/<int:template_id>/fields/<int:field_id>/display-name", methods=["POST"])
def update_field_display_name(template_id, field_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    field = db.session.get(ReportTemplateField, field_id)
    if not field:
        return "Not Found", 404
    version = db.session.get(ReportTemplateVersion, field.version_id)
    if not version or version.template_id != template_id:
        return "Forbidden", 403
    display_name = (request.form.get("display_name") or "").strip()
    levels = []
    for idx in range(1, 21):
        value = (request.form.get(f"level_{idx}") or "").strip()
        if value:
            levels.append(value)
    if levels:
        display_name = levels[-1]
    if not display_name:
        display_name = field.field_name or field.field_code
    if levels:
        levels[-1] = display_name
        field.path_code = " > ".join(levels)
    field.display_name = display_name
    db.session.commit()
    _audit("update_field_display_name", "report_template_field", field.id, display_name)
    flash("Đã cập nhật tên hiển thị trường dữ liệu.", "success")
    return redirect(url_for("reporting_bp.template_detail", template_id=template_id))


@reporting_bp.route("/admin/reports/templates/<int:template_id>/versions/<int:version_id>/activate", methods=["POST"])
def activate_version(template_id, version_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    ReportTemplateVersion.query.filter_by(template_id=template_id).update({"is_current": False})
    version = db.session.get(ReportTemplateVersion, version_id)
    if version:
        version.is_current = True
        db.session.commit()
        _audit("activate_version", "report_template_version", version.id, f"template={template_id}")
        flash("Đã chuyển sang bản mẫu này.", "success")
    return redirect(url_for("reporting_bp.template_detail", template_id=template_id))


@reporting_bp.route("/admin/reports/templates/<int:template_id>/view")
def admin_template_preview(template_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    template = db.session.get(ReportTemplate, template_id)
    if not template:
        return "Not Found", 404
    template_version = _template_version(template)
    if not template_version:
        flash("Biểu mẫu này chưa có file để xem.", "warning")
        return redirect(url_for("reporting_bp.template_detail", template_id=template.id))
    metadata = json.loads(template_version.metadata_json or "{}")
    workbook, formula_values = build_preview_workbook(template_version.source_path, {})
    preview_sheets = []
    for sheet_meta in metadata.get("sheets", []):
        ws = workbook[sheet_meta["sheet_name"]]
        start_row, header_end_row = _sheet_header_range(sheet_meta)
        unit_end_row = int(sheet_meta.get("unit_end_row") or sheet_meta.get("data_end_row") or header_end_row + 1)
        total_end_row = int(sheet_meta.get("total_end_row") or 0)
        start_col = column_index_from_string(sheet_meta.get("start_column") or "A")
        end_col = column_index_from_string(sheet_meta.get("end_column") or sheet_meta.get("start_column") or "A")
        sticky_col = _preferred_sticky_column(sheet_meta, ws, formula_values.get(sheet_meta["sheet_name"], {}))
        preview_sheets.append(
            {
                "sheet_name": sheet_meta["sheet_name"],
                "html": render_sheet_html(
                    ws,
                    editable_values=formula_values.get(sheet_meta["sheet_name"], {}),
                    field_lookup={},
                    editable=False,
                    start_row=start_row,
                    end_row=max(unit_end_row, total_end_row or 0),
                    min_col=start_col,
                    max_col=end_col,
                    header_end_row=header_end_row,
                    sticky_first_col=sticky_col,
                ),
            }
        )
    return render_template(
        "report_template_preview.html",
        template=template,
        template_version=template_version,
        preview_sheets=preview_sheets,
        download_url=url_for("reporting_bp.download_template_source", template_id=template.id),
    )


@reporting_bp.route("/admin/reports/templates/<int:template_id>/download")
def download_template_source(template_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    template = db.session.get(ReportTemplate, template_id)
    version = _template_version(template) if template else None
    if not version or not version.source_path or not os.path.exists(version.source_path):
        return "Not Found", 404
    return send_file(version.source_path, as_attachment=True, download_name=version.source_filename or os.path.basename(version.source_path))


@reporting_bp.route("/admin/reports/templates/<int:template_id>/directive")
def download_template_directive(template_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    template = db.session.get(ReportTemplate, template_id)
    if not template or not template.directive_path or not os.path.exists(template.directive_path):
        return "Not Found", 404
    return send_file(template.directive_path, as_attachment=True, download_name=template.directive_filename or os.path.basename(template.directive_path))


@reporting_bp.route("/admin/reports/templates/<int:template_id>/versions/<int:version_id>/delete", methods=["POST"])
def delete_version(template_id, version_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    version = db.session.get(ReportTemplateVersion, version_id)
    if not version or version.template_id != template_id:
        return "Not Found", 404
    if _version_usage_count(version.id):
        flash("Không thể xóa bản mẫu này vì đã có báo cáo sử dụng.", "warning")
        return redirect(url_for("reporting_bp.template_detail", template_id=template_id))

    remaining_versions = ReportTemplateVersion.query.filter(
        ReportTemplateVersion.template_id == template_id,
        ReportTemplateVersion.id != version.id,
    ).order_by(ReportTemplateVersion.version_no.desc()).all()
    was_current = bool(version.is_current)

    if version.source_path and os.path.exists(version.source_path):
        try:
            os.remove(version.source_path)
        except Exception:
            pass

    ReportTemplateField.query.filter_by(version_id=version.id).delete(synchronize_session=False)
    ReportTemplateSheet.query.filter_by(version_id=version.id).delete(synchronize_session=False)
    db.session.delete(version)

    if remaining_versions:
        if was_current:
            for item in remaining_versions:
                item.is_current = False
            remaining_versions[0].is_current = True
        message = "Đã xóa bản mẫu."
    else:
        template = db.session.get(ReportTemplate, template_id)
        if template:
            db.session.delete(template)
        message = "Đã xóa mẫu báo cáo vì không còn bản nào."

    db.session.commit()
    _audit("delete_version", "report_template_version", version_id, f"template={template_id}")
    flash(message, "success")
    if remaining_versions:
        return redirect(url_for("reporting_bp.template_detail", template_id=template_id))
    return redirect(url_for("reporting_bp.admin_dashboard"))


@reporting_bp.route("/admin/reports/templates/<int:template_id>/delete", methods=["POST"])
def delete_template(template_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    template = db.session.get(ReportTemplate, template_id)
    if not template:
        return "Not Found", 404

    versions = ReportTemplateVersion.query.filter_by(template_id=template.id).all()
    version_ids = [version.id for version in versions]
    if version_ids:
        cycles = ReportCycle.query.filter(ReportCycle.template_version_id.in_(version_ids)).all()
        for cycle in cycles:
            _purge_cycle(cycle)
    ReportingPeriod.query.filter_by(template_id=template.id).delete(synchronize_session=False)

    for version in versions:
        if version.source_path and os.path.exists(version.source_path):
            try:
                os.remove(version.source_path)
            except Exception:
                pass
        ReportTemplateField.query.filter_by(version_id=version.id).delete(synchronize_session=False)
        ReportTemplateSheet.query.filter_by(version_id=version.id).delete(synchronize_session=False)
        db.session.delete(version)
    if template.directive_path and os.path.exists(template.directive_path):
        try:
            os.remove(template.directive_path)
        except Exception:
            pass
    db.session.delete(template)
    db.session.commit()
    _audit("delete_template", "report_template", template_id, template.name)
    flash("Đã xóa mẫu báo cáo.", "success")
    return redirect(url_for("reporting_bp.admin_dashboard"))


@reporting_bp.route("/admin/reports/units/sync", methods=["POST"])
def sync_units():
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    _sync_units_from_users()
    flash("Đã đồng bộ đơn vị từ tài khoản.", "success")
    return redirect(url_for("reporting_bp.admin_dashboard"))


@reporting_bp.route("/admin/reports/cycles/<int:cycle_id>")
def admin_cycle_detail(cycle_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    cycle = db.session.get(ReportCycle, cycle_id)
    if not cycle:
        return "Not Found", 404
    template_version = db.session.get(ReportTemplateVersion, cycle.template_version_id)
    template = db.session.get(ReportTemplate, template_version.template_id) if template_version else None
    report_type = _report_type(cycle)
    is_daily = report_type and report_type.code == "daily"
    scoped_units = _cycle_units(cycle)
    instance_list = ReportInstance.query.filter_by(cycle_id=cycle.id).order_by(ReportInstance.report_unit_id.asc()).all()
    instance_map = {instance.report_unit_id: instance for instance in instance_list if instance.report_unit_id}

    # --- Date filter for daily reports ---
    report_date_str = request.args.get("report_date", "")
    filter_date = _parse_date(report_date_str) if report_date_str else None
    if is_daily and not filter_date:
        filter_date = date.today()

    # --- Build available dates list (for daily reports) ---
    available_dates = []
    if is_daily:
        all_submission_dates = set()
        all_instances = ReportInstance.query.filter_by(cycle_id=cycle.id).all()
        for inst in all_instances:
            subs = ReportSubmission.query.filter_by(instance_id=inst.id).all()
            for s in subs:
                business_date = _submission_business_date(s)
                if business_date:
                    all_submission_dates.add(business_date)
            # Also include the cycle open_at date
            if inst.opened_at:
                all_submission_dates.add(inst.opened_at.date())
        # Include the cycle's open date and today
        if cycle.open_at:
            all_submission_dates.add(cycle.open_at.date())
        all_submission_dates.add(date.today())
        available_dates = sorted(all_submission_dates, reverse=True)

    unit_rows = []
    latest_submissions = []
    submission_total = 0
    submitted_total = 0
    reported_total = 0
    not_reported_total = 0
    on_time_total = 0
    late_total = 0
    filter_start = _start_of_day(filter_date) if is_daily and filter_date else None
    filter_end = _next_day_start(filter_date) if is_daily and filter_date else None

    for unit in scoped_units:
        instance = instance_map.get(unit.id)

        working_state = _resolve_working_submission_state(
            instance,
            template_version,
            report_type=report_type,
            report_date=filter_date if is_daily else None,
        ) if instance else {"latest_submission": None, "history": []}
        latest_submission = working_state["latest_submission"]
        submission_count = len(working_state["history"])

        submission_total += submission_count
        timeliness = _submission_timeliness(cycle, latest_submission, report_type=report_type)

        # For daily reports: count units with data available through the selected date.
        if is_daily:
            if latest_submission:
                reported_total += 1
                latest_submissions.append(latest_submission)
                if latest_submission.status == "submitted":
                    on_time_total += 1
            else:
                not_reported_total += 1
        else:
            if latest_submission and instance and instance.status == "submitted":
                submitted_total += 1
            if latest_submission:
                latest_submissions.append(latest_submission)
                if timeliness in {"Đúng ngày", "Đúng hạn", "Đã báo cáo"}:
                    on_time_total += 1
                elif timeliness in {"Quá ngày", "Quá hạn"}:
                    late_total += 1

        unit_rows.append(
            {
                "unit": unit,
                "instance": instance,
                "latest_submission": latest_submission,
                "submission_count": submission_count,
                "timeliness": timeliness,
            }
        )

    instance_ids = [row["instance"].id for row in unit_rows if row["instance"]]
    submissions = (
        ReportSubmission.query.filter(
            ReportSubmission.instance_id.in_(instance_ids),
            *(
                [
                    _submission_time_expr() < filter_end,
                ]
                if is_daily and filter_start and filter_end
                else []
            ),
        )
        .order_by(ReportSubmission.created_at.desc(), ReportSubmission.id.desc())
        .all()
        if instance_ids
        else []
    )
    history_rows = _history_rows_for_submissions(
        submissions,
        template_version,
        report_type,
        include_unit=True,
        include_actor=True,
        cycle=cycle,
    )
    return render_template(
        "reporting_cycle_detail.html",
        cycle=cycle,
        template=template,
        template_version=template_version,
        report_type=report_type,
        is_daily=is_daily,
        scoped_units=scoped_units,
        scope_count=len(scoped_units),
        scope_payload=_parse_scope_payload(cycle.scope_json),
        unit_rows=unit_rows,
        history_rows=history_rows,
        submission_total=submission_total,
        submitted_total=submitted_total,
        reported_total=reported_total,
        not_reported_total=not_reported_total,
        on_time_total=on_time_total,
        late_total=late_total,
        filter_date=filter_date,
        filter_date_str=filter_date.strftime("%d/%m/%Y") if filter_date else "",
        available_dates=available_dates,
        today=date.today(),
    )


@reporting_bp.route("/admin/reports/cycles/<int:cycle_id>/update", methods=["POST"])
def update_cycle(cycle_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    cycle = db.session.get(ReportCycle, cycle_id)
    if not cycle:
        return "Not Found", 404
    template = _template_for_cycle(cycle)
    name = (request.form.get("name") or cycle.name or "").strip()
    if not name:
        flash("Tên báo cáo không được để trống.", "danger")
        return redirect(url_for("reporting_bp.admin_cycle_detail", cycle_id=cycle.id))
    report_type_id = int(request.form.get("report_type_id") or cycle.report_type_id or 0)
    report_type = db.session.get(ReportType, report_type_id)
    if not report_type:
        flash("Loại báo cáo không hợp lệ.", "danger")
        return redirect(url_for("reporting_bp.admin_cycle_detail", cycle_id=cycle.id))
    cycle.name = name
    cycle.report_type_id = report_type.id
    cycle.note = (request.form.get("note") or "").strip()
    if report_type.code == "daily":
        report_date = _parse_date(request.form.get("report_date")) or datetime.now().date()
        cycle.open_at = _start_of_day(report_date)
        cycle.due_at = _end_of_day(report_date)
    elif report_type.code == "periodic":
        due_date = _parse_date(request.form.get("due_date"))
        if not due_date and template:
            due_date = _default_periodic_due_date(template)
        cycle.open_at = cycle.open_at or datetime.now()
        cycle.due_at = _end_of_day(due_date) if due_date else None
    else:
        due_date = _parse_date(request.form.get("due_date"))
        cycle.open_at = cycle.open_at or datetime.now()
        cycle.due_at = _end_of_day(due_date) if due_date else None
    cycle.auto_lock_at = None
    _ensure_reporting_period(cycle, report_type=report_type)
    db.session.commit()
    _audit("update_cycle", "report_cycle", cycle.id, cycle.name)
    flash("Đã cập nhật báo cáo.", "success")
    return redirect(url_for("reporting_bp.admin_cycle_detail", cycle_id=cycle.id))


@reporting_bp.route("/admin/reports/cycles/<int:cycle_id>/close", methods=["POST"])
def close_cycle(cycle_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    cycle = db.session.get(ReportCycle, cycle_id)
    if not cycle:
        return "Not Found", 404
    cycle.status = "closed"
    cycle.is_locked = True
    cycle.close_at = datetime.now()
    _ensure_reporting_period(cycle)
    for instance in ReportInstance.query.filter_by(cycle_id=cycle.id).all():
        instance.locked_at = cycle.close_at
    materialized = _materialize_cycle_effective_exports(
        cycle,
        overwrite=True,
        apply=True,
        include_open=True,
    )
    db.session.commit()
    _audit("close_cycle", "report_cycle", cycle.id, cycle.name)
    if materialized["errors"]:
        flash(
            f"Đã đóng báo cáo, nhưng có {len(materialized['errors'])} đơn vị chưa tạo được tệp dữ liệu chốt.",
            "warning",
        )
    flash("Đã đóng báo cáo.", "success")
    return redirect(url_for("reporting_bp.admin_cycle_detail", cycle_id=cycle.id))


@reporting_bp.route("/admin/reports/cycles/<int:cycle_id>/delete", methods=["POST"])
def delete_cycle(cycle_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    cycle = db.session.get(ReportCycle, cycle_id)
    if not cycle:
        return "Not Found", 404
    cycle_name = cycle.name
    _purge_cycle(cycle)
    db.session.commit()
    _audit("delete_cycle", "report_cycle", cycle_id, cycle_name)
    flash("Đã xóa báo cáo.", "success")
    return redirect(url_for("reporting_bp.admin_dashboard"))


@reporting_bp.route("/admin/reports/submissions/<int:submission_id>/delete", methods=["POST"])
def delete_submission(submission_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    submission = db.session.get(ReportSubmission, submission_id)
    if not submission:
        return "Not Found", 404
    instance = db.session.get(ReportInstance, submission.instance_id)
    cycle_id = instance.cycle_id if instance else 0
    _purge_submission(submission)
    db.session.flush()
    if instance:
        _refresh_instance_status(instance)
    db.session.commit()
    _audit("delete_submission", "report_submission", submission_id, f"cycle={cycle_id}")
    flash("Đã xóa bản nộp.", "success")
    return redirect(url_for("reporting_bp.admin_cycle_detail", cycle_id=cycle_id))


@reporting_bp.route("/admin/reports/cycles/<int:cycle_id>/reset-data", methods=["POST"])
def reset_cycle_data(cycle_id):
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    cycle = db.session.get(ReportCycle, cycle_id)
    if not cycle:
        return "Not Found", 404

    instances = ReportInstance.query.filter_by(cycle_id=cycle.id).all()
    submission_ids = []
    for instance in instances:
        submissions = ReportSubmission.query.filter_by(instance_id=instance.id).all()
        for submission in submissions:
            submission_ids.append(submission.id)
            _purge_submission(submission)

    db.session.flush()

    for instance in instances:
        _refresh_instance_status(instance)
        if cycle.is_locked or cycle.status == "closed":
            instance.locked_at = cycle.close_at or instance.locked_at

    db.session.commit()
    _audit(
        "reset_cycle_data",
        "report_cycle",
        cycle.id,
        f"submissions={len(submission_ids)}",
    )
    flash(f"Đã reset dữ liệu báo cáo, xóa {len(submission_ids)} bản lưu/gửi.", "success")
    return redirect(url_for("reporting_bp.admin_cycle_detail", cycle_id=cycle.id))


@reporting_bp.route("/admin/reports/cycles", methods=["POST"])
def create_cycle():
    if not _is_admin():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    template_version_id = int(request.form.get("template_version_id") or 0)
    report_type_id = int(request.form.get("report_type_id") or 0)
    name = (request.form.get("name") or "").strip()
    if not template_version_id or not report_type_id or not name:
        flash("Thiếu dữ liệu để mở báo cáo.", "danger")
        return redirect(url_for("reporting_bp.admin_dashboard"))
    report_type = db.session.get(ReportType, report_type_id)
    if not report_type:
        flash("Loại báo cáo không hợp lệ.", "danger")
        return redirect(url_for("reporting_bp.admin_dashboard"))

    template_version = db.session.get(ReportTemplateVersion, template_version_id)
    template = db.session.get(ReportTemplate, template_version.template_id) if template_version else None
    scope_payload = _scope_payload_from_form(request.form)
    if not (scope_payload["unit_ids"] or scope_payload["role_ids"] or scope_payload["unit_groups"]) and template:
        scope_payload = _parse_scope_payload(template.assignment_scope_json)
    unit_ids = [unit.id for unit in _resolve_scope_units(scope_payload)]

    if report_type.code == "daily":
        report_date = _parse_date(request.form.get("report_date")) or datetime.now().date()
        open_at = _start_of_day(report_date)
        due_at = _end_of_day(report_date)
    else:
        due_date = _parse_date(request.form.get("due_date"))
        if report_type.code == "periodic" and not due_date and template:
            due_date = _default_periodic_due_date(template)
        open_at = datetime.now()
        due_at = _end_of_day(due_date) if due_date else None

    cycle = ReportCycle(
        template_version_id=template_version_id,
        report_type_id=report_type_id,
        name=name,
        open_at=open_at,
        close_at=None,
        due_at=due_at,
        auto_lock_at=None,
        status="open",
        scope_json=json.dumps(scope_payload, ensure_ascii=False),
        is_locked=False,
        note=(request.form.get("note") or "").strip(),
    )
    db.session.add(cycle)
    db.session.flush()
    period = _ensure_reporting_period(cycle, report_type=report_type)

    for unit_id in unit_ids:
        unit = db.session.get(ReportUnit, unit_id)
        user = None
        if unit:
            user = (
                User.query.filter(
                    or_(User.unit_key == unit.code, User.unit_area == unit.name)
                )
                .order_by(User.id.asc())
                .first()
            )
        db.session.add(
            ReportInstance(
                template_id=template_version.template_id if template_version else None,
                version_id=template_version.id if template_version else None,
                period_id=period.id if period else None,
                user_id=user.id if user else None,
                org_unit=unit.name if unit else "",
                cycle_id=cycle.id,
                report_unit_id=unit_id,
                assigned_user_id=user.id if user else None,
                status="draft",
            )
        )

    db.session.commit()
    _audit("create_cycle", "report_cycle", cycle.id, name)
    flash("Đã mở báo cáo.", "success")
    return redirect(url_for("reporting_bp.admin_dashboard"))


@reporting_bp.route("/reports")
def user_dashboard():
    if not _require_login():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    _ensure_default_types()
    _sync_units_from_users()
    _ensure_active_cycles_for_ready_templates()

    user = db.session.get(User, session.get("uid"))
    unit = _current_user_report_unit()
    is_admin = _is_admin()
    can_view_cycle_progress = is_admin or _is_pc06_role(user)
    dashboard_report_date = date.today()
    cycles = ReportCycle.query.order_by(ReportCycle.created_at.desc()).all()
    accessible_cycles = [cycle for cycle in cycles if _cycle_accessible(cycle, unit.id if unit else None, is_admin)]
    instances = []
    if unit:
        instances = ReportInstance.query.filter_by(report_unit_id=unit.id).order_by(ReportInstance.opened_at.desc()).all()
    instance_map = {instance.cycle_id: instance for instance in instances}
    latest_submission_map = {
        instance.cycle_id: _latest_submission(instance.id)
        for instance in instances
        if instance and instance.id
    }
    cycle_report_types = {
        cycle.id: _report_type(cycle)
        for cycle in accessible_cycles
    }
    cycle_progress_map = {}
    cycle_status_map = {}
    for cycle in accessible_cycles:
        report_type = cycle_report_types.get(cycle.id)
        progress = _dashboard_cycle_progress(cycle, report_type=report_type, report_date=dashboard_report_date)
        cycle_progress_map[cycle.id] = progress
        cycle_status_map[cycle.id] = _dashboard_cycle_status(
            cycle,
            instance=instance_map.get(cycle.id),
            report_type=report_type,
            report_date=dashboard_report_date,
            progress=progress,
        )
    hero_stats = _dashboard_hero_stats(accessible_cycles, cycle_status_map=cycle_status_map)

    return render_template(
        "reporting_dashboard.html",
        templates=[],
        cycle_groups=_group_cycles_by_professional_unit(accessible_cycles),
        versions=[],
        cycles=accessible_cycles,
        units=[unit] if unit else [],
        report_types=[],
        recent_submissions=[],
        current_versions={},
        is_admin=is_admin,
        current_unit=unit,
        current_user=user,
        instances=instances,
        instance_map=instance_map,
        latest_submission_map=latest_submission_map,
        cycle_report_types=cycle_report_types,
        cycle_progress_map=cycle_progress_map,
        cycle_status_map=cycle_status_map,
        hero_stats=hero_stats,
        can_view_cycle_progress=can_view_cycle_progress,
    )


@reporting_bp.route("/reports/cycles/<int:cycle_id>")
def cycle_workspace(cycle_id):
    if not _require_login():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    _ensure_default_types()
    _sync_units_from_users()

    context = _resolve_cycle_context(cycle_id)
    if not context:
        return "Not Found", 404
    cycle = context["cycle"]
    unit = context["unit"]
    instance = context["instance"]
    template_version = context["template_version"]
    template = context["template"]
    report_type = context["report_type"]
    editable = not cycle.is_locked and cycle.status != "closed"
    metadata = json.loads(template_version.metadata_json or "{}")
    workbook = load_workbook(template_version.source_path, data_only=False)
    report_date = context["report_date"]
    latest_submission = context["latest_submission"]
    submission_history = context["submission_history"]
    existing_values = context.get("entry_values", {})
    available_units = _cycle_units(cycle) if _is_admin() else []
    sheet_views = []
    for sheet_meta in metadata.get("sheets", []):
        ws = workbook[sheet_meta["sheet_name"]]
        sheet_fields = ReportTemplateField.query.filter_by(version_id=template_version.id, sheet_name=sheet_meta["sheet_name"]).order_by(ReportTemplateField.display_order.asc()).all()
        editable_fields = [field for field in sheet_fields if field.is_visible and field.is_editable]
        if not editable_fields:
            editable_fields = [field for field in sheet_fields if field.is_editable]
        _, header_end_row = _sheet_header_range(sheet_meta)
        unit_start_row = int(sheet_meta.get("unit_start_row") or sheet_meta.get("data_start_row") or header_end_row + 1)
        unit_end_row = int(sheet_meta.get("unit_end_row") or sheet_meta.get("data_end_row") or unit_start_row)
        row_entries = []
        for row_index in range(unit_start_row, unit_end_row + 1):
            if ws.row_dimensions[row_index].hidden:
                continue
            if not _is_admin() and unit and not _row_matches_unit(
                ws,
                sheet_fields,
                existing_values,
                sheet_meta["sheet_name"],
                row_index,
                unit,
            ):
                continue
            inputs = []
            for field in editable_fields:
                coord = f"{field.column_letter}{row_index}"
                value = existing_values.get(sheet_meta["sheet_name"], {}).get(coord, ws[coord].value)
                inputs.append({
                    "cell_address": coord,
                    "value": _cell_display_value(value),
                    "field_code": field.field_code,
                    "field_label": _field_display_name(field),
                    "field_path": " / ".join(_field_levels(field)[:-1]) if len(_field_levels(field)) > 1 else "",
                })
            if inputs:
                row_entries.append({
                    "excel_row": row_index,
                    "title": _row_context_label(ws, sheet_fields, existing_values, sheet_meta["sheet_name"], row_index),
                    "inputs": inputs,
                })
        sheet_views.append({
            "sheet_name": sheet_meta["sheet_name"],
            "field_count": len(sheet_meta.get("fields", [])),
            "input_count": len(editable_fields),
            "rows": row_entries,
            "config": {
                "header_start_row": sheet_meta.get("header_start_row"),
                "header_end_row": sheet_meta.get("header_end_row"),
                "unit_start_row": sheet_meta.get("unit_start_row") or sheet_meta.get("data_start_row"),
                "unit_end_row": sheet_meta.get("unit_end_row") or sheet_meta.get("data_end_row"),
                "total_start_row": sheet_meta.get("total_start_row"),
                "total_end_row": sheet_meta.get("total_end_row"),
                "start_column": sheet_meta.get("start_column"),
                "end_column": sheet_meta.get("end_column"),
            },
        })

    return render_template(
        "report_cycle_form.html",
        cycle=cycle,
        template=template,
        template_version=template_version,
        report_type=report_type,
        schedule_text=_report_schedule_text(cycle, report_type=report_type, report_date=report_date),
        instance=instance,
        sheet_views=sheet_views,
        latest_submission=latest_submission,
        latest_timeliness=_submission_timeliness(cycle, latest_submission, report_type=report_type),
        submission_history=submission_history,
        report_date=report_date,
        current_unit=unit,
        is_admin=_is_admin(),
        available_units=available_units,
        editable=editable,
    )


@reporting_bp.route("/reports/cycles/<int:cycle_id>/view")
def cycle_preview(cycle_id):
    if not _require_login():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    _ensure_default_types()
    _sync_units_from_users()

    context = _resolve_cycle_context(cycle_id)
    if not context:
        return "Not Found", 404
    cycle = context["cycle"]
    unit = context["unit"]
    template_version = context["template_version"]
    template = context["template"]
    report_type = context["report_type"]
    latest_submission = context["view_submission"]
    report_date = context["report_date"]
    metadata = json.loads(template_version.metadata_json or "{}")
    admin_all_units = _is_admin() and not request.args.get("unit_id")
    is_daily_cumulative = bool(report_type and report_type.code == "daily" and report_date)
    if admin_all_units:
        instances = ReportInstance.query.filter_by(cycle_id=cycle.id).order_by(ReportInstance.report_unit_id.asc()).all()
        if is_daily_cumulative:
            existing_values = {}
            latest_submissions = []
            for instance in instances:
                submissions = _daily_snapshot_submissions_through_date(instance.id, report_date)
                if submissions:
                    latest_submissions.append(submissions[-1])
                    unit_values = _cumulative_submission_cell_values(submissions, template_version.id)
                    for sheet_name, cells in unit_values.items():
                        existing_values.setdefault(sheet_name, {}).update(cells)
            latest_submission = latest_submissions[0] if latest_submissions else None
        else:
            latest_submissions = []
            for instance in instances:
                submission = _latest_submission(instance.id)
                if submission:
                    latest_submissions.append(submission)
            existing_values = _merged_submission_cell_values(latest_submissions)
            latest_submission = latest_submissions[0] if latest_submissions else None
        unit = None
    elif is_daily_cumulative:
        existing_values = context.get("working_values", {})
        submissions = context.get("daily_submissions", [])
        latest_submission = submissions[-1] if submissions else None
    else:
        existing_values = context.get("working_values", {})
    download_url = (
        url_for("reporting_bp.download_submission", submission_id=latest_submission.id)
        if latest_submission and not admin_all_units and not is_daily_cumulative
        else ""
    )
    if g.get("is_mobile"):
        if download_url:
            flash("Trên mobile, hệ thống sẽ tải file báo cáo để xem kết quả.", "info")
            return redirect(download_url)
        flash("Trên mobile, giao diện xem báo cáo đã được ẩn. Vui lòng theo dõi tiến độ hoặc tải file khi có sẵn.", "info")
        if _is_admin():
            fallback_values = {}
            if report_date:
                fallback_values["report_date"] = report_date.strftime("%Y-%m-%d")
            return redirect(url_for("reporting_bp.admin_cycle_detail", cycle_id=cycle.id, **fallback_values))
        return redirect(url_for("reporting_bp.user_dashboard"))
    workbook, formula_values = build_preview_workbook(template_version.source_path, existing_values)
    preview_sheets = []
    for sheet_meta in metadata.get("sheets", []):
        ws = workbook[sheet_meta["sheet_name"]]
        start_row, header_end_row = _sheet_header_range(sheet_meta)
        unit_end_row = int(sheet_meta.get("unit_end_row") or sheet_meta.get("data_end_row") or header_end_row + 1)
        total_end_row = int(sheet_meta.get("total_end_row") or 0)
        start_col = column_index_from_string(sheet_meta.get("start_column") or "A")
        end_col = column_index_from_string(sheet_meta.get("end_column") or sheet_meta.get("start_column") or "A")
        sheet_values = {
            **existing_values.get(sheet_meta["sheet_name"], {}),
            **formula_values.get(sheet_meta["sheet_name"], {}),
        }
        sticky_col = _preferred_sticky_column(sheet_meta, ws, sheet_values)
        preview_sheets.append({
            "sheet_name": sheet_meta["sheet_name"],
            "html": render_sheet_html(
                ws,
                editable_values=sheet_values,
                field_lookup={},
                editable=False,
                start_row=start_row,
                end_row=max(unit_end_row, total_end_row or 0),
                min_col=start_col,
                max_col=end_col,
                header_end_row=header_end_row,
                sticky_first_col=sticky_col,
            ),
        })
    return render_template(
        "report_cycle_preview.html",
        cycle=cycle,
        template=template,
        report_type=report_type,
        report_date=report_date,
        report_date_str=report_date.strftime("%d/%m/%Y") if report_date else "",
        latest_submission=latest_submission,
        preview_sheets=preview_sheets,
        current_unit=unit,
        is_admin=_is_admin(),
        download_url=download_url,
    )


@reporting_bp.route("/reports/cycles/<int:cycle_id>/history")
def cycle_history(cycle_id):
    if not _require_login():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    _ensure_default_types()
    _sync_units_from_users()

    context = _resolve_cycle_context(cycle_id)
    if not context:
        return "Not Found", 404
    history_rows = _history_rows_for_submissions(
        context["view_history"],
        context["template_version"],
        context["report_type"],
        include_unit=False,
        include_actor=False,
        cycle=context["cycle"],
    )
    return render_template(
        "report_cycle_history.html",
        cycle=context["cycle"],
        template=context["template"],
        report_type=context["report_type"],
        report_date=context["report_date"],
        report_date_str=context["report_date"].strftime("%d/%m/%Y") if context["report_date"] else "",
        current_unit=context["unit"],
        latest_submission=context["view_submission"],
        history_rows=history_rows,
        is_admin=_is_admin(),
    )


def _payload_from_request():
    payload = request.form.get("payload_json")
    if payload:
        try:
            return json.loads(payload)
        except Exception:
            return {}
    if request.is_json:
        return request.get_json(silent=True) or {}
    return {}


def _workspace_route_values(cycle, unit=None, report_date=None):
    values = {"cycle_id": cycle.id}
    if unit and _is_admin():
        values["unit_id"] = unit.id
    if report_date:
        values["report_date"] = report_date.strftime("%Y-%m-%d")
    return values


def _resolve_cycle_context(cycle_id):
    cycle = db.session.get(ReportCycle, cycle_id)
    if not cycle:
        return None
    unit = _resolve_cycle_unit(cycle)
    if not _cycle_accessible(cycle, unit.id if unit else None, _is_admin()):
        return None
    user = db.session.get(User, session.get("uid"))
    instance = _get_cycle_instance(cycle, unit, user)
    template_version = db.session.get(ReportTemplateVersion, cycle.template_version_id)
    template = db.session.get(ReportTemplate, template_version.template_id)
    report_type = _report_type(cycle)
    report_date = None
    if report_type and report_type.code == "daily":
        report_date = _parse_date(request.args.get("report_date", "")) or date.today()
    entry_state = _resolve_entry_submission_state(
        instance,
        template_version,
        report_type=report_type,
        report_date=report_date,
    )
    working_state = _resolve_working_submission_state(
        instance,
        template_version,
        report_type=report_type,
        report_date=report_date,
    )
    return {
        "cycle": cycle,
        "unit": unit,
        "user": user,
        "instance": instance,
        "template_version": template_version,
        "template": template,
        "report_type": report_type,
        "report_date": report_date,
        "latest_submission": entry_state["latest_submission"],
        "submission_history": entry_state["history"],
        "entry_values": entry_state["existing_values"],
        "view_submission": working_state["latest_submission"],
        "view_history": working_state["history"],
        "working_values": working_state["existing_values"],
        "daily_submissions": working_state["daily_submissions"],
    }


@reporting_bp.route("/reports/cycles/<int:cycle_id>/save", methods=["POST"])
def save_cycle(cycle_id):
    if not _require_login():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    cycle = db.session.get(ReportCycle, cycle_id)
    unit = _resolve_cycle_unit(cycle) if cycle else None
    report_type = _report_type(cycle) if cycle else None
    report_date = _parse_date(request.form.get("report_date", "")) if report_type and report_type.code == "daily" else None
    workspace_values = _workspace_route_values(cycle, unit=unit, report_date=report_date) if cycle else {"cycle_id": cycle_id}
    if not cycle or not _cycle_accessible(cycle, unit.id if unit else None, _is_admin()):
        return "Forbidden", 403
    if cycle.is_locked or cycle.status == "closed":
        flash("Báo cáo đã khóa, không thể lưu thêm dữ liệu.", "warning")
        return redirect(url_for("reporting_bp.cycle_workspace", **workspace_values))
    user = db.session.get(User, session.get("uid"))
    instance = _get_cycle_instance(cycle, unit, user)
    payload = _payload_from_request()
    submission, errors = _save_submission(instance, payload, final_submit=False, report_date=report_date)
    _audit("save_draft", "report_submission", submission.id, f"cycle={cycle.id}")
    flash("Đã lưu nháp báo cáo.", "success")
    return redirect(url_for("reporting_bp.cycle_workspace", **workspace_values))


@reporting_bp.route("/reports/cycles/<int:cycle_id>/submit", methods=["POST"])
def submit_cycle(cycle_id):
    if not _require_login():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    cycle = db.session.get(ReportCycle, cycle_id)
    unit = _resolve_cycle_unit(cycle) if cycle else None
    report_type = _report_type(cycle) if cycle else None
    report_date = _parse_date(request.form.get("report_date", "")) if report_type and report_type.code == "daily" else None
    workspace_values = _workspace_route_values(cycle, unit=unit, report_date=report_date) if cycle else {"cycle_id": cycle_id}
    if not cycle or not _cycle_accessible(cycle, unit.id if unit else None, _is_admin()):
        return "Forbidden", 403
    if cycle.is_locked or cycle.status == "closed":
        flash("Báo cáo đã khóa, không thể gửi thêm dữ liệu.", "warning")
        return redirect(url_for("reporting_bp.cycle_workspace", **workspace_values))
    user = db.session.get(User, session.get("uid"))
    instance = _get_cycle_instance(cycle, unit, user)
    payload = _payload_from_request()
    submission, errors = _save_submission(instance, payload, final_submit=True, report_date=report_date)
    if errors:
        flash("Còn dữ liệu bắt buộc chưa hoàn tất, hệ thống đã giữ ở trạng thái nháp.", "warning")
        return redirect(url_for("reporting_bp.cycle_workspace", **workspace_values))
    output_path = _export_submission(submission)
    submission.file_path = output_path
    db.session.commit()
    _audit("submit_report", "report_submission", submission.id, f"cycle={cycle.id}")
    flash("Đã gửi báo cáo.", "success")
    return redirect(url_for("reporting_bp.cycle_workspace", **workspace_values))


@reporting_bp.route("/reports/submissions/<int:submission_id>/download")
def download_submission(submission_id):
    if not _require_login():
        return redirect(url_for("auth_bp.login"))
    _ensure_report_schema()
    submission = db.session.get(ReportSubmission, submission_id)
    if not submission:
        return "Not Found", 404
    download_path = _submission_download_path(submission)
    if not download_path:
        submission.file_path = _export_submission(submission)
        db.session.commit()
        download_path = _submission_download_path(submission)
    return send_file(download_path, as_attachment=True, download_name=os.path.basename(download_path))
