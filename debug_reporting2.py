import sys
import io

from app import app
from models_reporting import FormTemplate
from pc06_excel_scanner import scan_excel_structure
from openpyxl.utils import column_index_from_string

with app.app_context():
    template = FormTemplate.query.order_by(FormTemplate.id.desc()).first()
    wb = scan_excel_structure.__globals__['_load_workbook_utf8'](io.BytesIO(template.excel_template_blob) if 'io' in sys.modules else __import__('io').BytesIO(template.excel_template_blob))
    ws = wb.active

    # Detect header rows - find first row with text
    header_candidates = []
    for row in range(1, min(20, ws.max_row + 1)):
        is_data = False
        numeric_count = 0
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row, col).value
            if isinstance(cell_val, (int, float)):
                numeric_count += 1
        if numeric_count >= 3:
            break
            
        has_text = False
        for col in range(1, min(10, ws.max_column + 1)):
            cell_val = ws.cell(row, col).value
            if cell_val and isinstance(cell_val, str) and len(str(cell_val).strip()) > 0:
                if str(cell_val).strip().startswith('='):
                    continue
                has_text = True
                break
        if has_text:
            header_candidates.append(row)

    print("Header candidates:", header_candidates)
    
    total_cols = ws.max_column
    real_header_rows = []
    for r in sorted(header_candidates):
        is_title = False
        for m in ws.merged_cells.ranges:
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(str(m))
            if min_row <= r <= max_row and (max_col - min_col + 1) >= total_cols * 0.5:
                is_title = True
                break
        if not is_title:
            real_header_rows.append(r)

    print("Real header rows:", real_header_rows)
    
    def get_header_text_for_cell(r, c_idx):
        for m in ws.merged_cells.ranges:
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(str(m))
            if min_row <= r <= max_row and min_col <= c_idx <= max_col:
                val = ws.cell(min_row, min_col).value
                return str(val).strip() if val else ""
        
        val = ws.cell(r, c_idx).value
        return str(val).strip() if val else ""

    for col_idx in range(1, 15):
        parts = []
        for r in real_header_rows:
            text = get_header_text_for_cell(r, col_idx)
            if text and text not in parts:
                parts.append(text)
                
        if len(parts) >= 2:
            parts = parts[-2:]
            section = parts[0]
        elif len(parts) == 1:
            section = 'Thông tin chung'
        else:
            parts = [f"Cột {col_idx}"]
            section = 'Thông tin chung'
            
        print(f"Col {col_idx}: section={section}, name={' - '.join(parts)}")
