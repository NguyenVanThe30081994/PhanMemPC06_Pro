# -*- coding: utf-8 -*-

from openpyxl import Workbook

from services.excel_formula_engine import ExcelFormulaEngine


def test_sum_and_iferror_ratio():
    wb = Workbook()
    ws = wb.active

    ws["C13"] = 10
    ws["D13"] = 5
    ws["E13"] = '=IFERROR(D13/C13,"")'
    ws["C12"] = "=SUM(C13:C14)"
    ws["D12"] = "=SUM(D13:D14)"
    ws["E12"] = '=IFERROR(D12/C12,"")'
    ws["C14"] = 0
    ws["D14"] = 0

    engine = ExcelFormulaEngine(wb)

    assert engine.evaluate_cell(ws, "E13") == 0.5
    assert engine.evaluate_cell(ws, "C12") == 10
    assert engine.evaluate_cell(ws, "D12") == 5
    assert engine.evaluate_cell(ws, "E12") == 0.5


def test_iferror_blank_on_division_by_zero():
    wb = Workbook()
    ws = wb.active
    ws["C13"] = 0
    ws["D13"] = 5
    ws["E13"] = '=IFERROR(D13/C13,"")'

    engine = ExcelFormulaEngine(wb)

    assert engine.evaluate_cell(ws, "E13") == ""
