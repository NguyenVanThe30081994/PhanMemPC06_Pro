from app import app
from models_reporting import FormTemplate
from pc06_excel_scanner import scan_excel_structure
from openpyxl.utils import column_index_from_string
from pc06_excel_scanner import _load_workbook_utf8
import io

with app.app_context():
    template = FormTemplate.query.order_by(FormTemplate.id.desc()).first()
    wb = _load_workbook_utf8(io.BytesIO(template.excel_template_blob))
    ws = wb.active
    print("Row 10:")
    for col in range(1, 15):
        print(f"Col {col}: {ws.cell(10, col).value}")
