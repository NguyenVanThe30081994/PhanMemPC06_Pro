# -*- coding: utf-8 -*-
"""
excel_renderer.py
-----------------
Shared utility: renders an openpyxl worksheet as a faithful HTML <table>,
preserving merged cells (colspan/rowspan), cell styles, and optionally
filling in submitted data values.
"""

import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.fills import PatternFill
from markupsafe import Markup
import unicodedata


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def format_excel_number(value, number_format):
    """
    Format number theo number_format của ô Excel.
    Ưu tiên: hiển thị phải theo format gốc, không format mặc định.
    
    Xử lý các format phổ biến:
    - 0, 0.0, 0.00 (số nguyên, 1 chữ số, 2 chữ số thập phân)
    - #,##0, #,##0.00 (có phân tách hàng nghìn)
    - 0%, 0.0%, 0.00% (phần trăm)
    """
    if value is None or value == '':
        return ''
    
    if not isinstance(value, (int, float)):
        return str(value).strip()
    
    if number_format is None:
        # Fallback: format mặc định
        fval = round(float(value), 10)
        if fval.is_integer():
            return str(int(fval))
        return f"{fval:.2f}".rstrip('0').rstrip('.')
    
    fmt = str(number_format).lower().strip()
    
    try:
        # 1. Xử lý phần trăm
        if '%' in fmt:
            decimals = 0
            if '.' in fmt:
                parts = fmt.split('.')
                if len(parts) > 1:
                    decimal_part = parts[1].split('%')[0]
                    decimals = len(decimal_part)
            result = f"{value * 100:.{decimals}f}%"
            return result
        
        # 2. Xử lý số có phân tách hàng nghìn (#,##0)
        if '#,##0' in fmt or ',' in fmt:
            decimals = 0
            if '.' in fmt:
                parts = fmt.split('.')
                if len(parts) > 1:
                    decimal_part = parts[1]
                    decimals = len(decimal_part)
            # Format với dấu phân tách hàng nghìn
            result = f"{value:,.{decimals}f}"
            # Nếu không có chữ số thập phân, cắt bỏ .00
            if decimals == 0:
                result = result.rstrip('0').rstrip('.')
            return result
        
        # 3. Xử lý số thập phân (0.00, 0.0, 0)
        if '0' in fmt:
            if '.' in fmt:
                parts = fmt.split('.')
                if len(parts) > 1:
                    decimal_part = parts[1]
                    decimals = len(decimal_part)
                    return f"{value:.{decimals}f}"
            # Số nguyên - làm tròn đến số nguyên gần nhất
            return str(int(round(value)))
        
        # 4. Fallback: hiển thị giá trị thô
        if float(value).is_integer():
            return str(int(value))
        return str(value)
    
    except Exception:
        # Nếu lỗi, fallback an toàn
        if float(value).is_integer():
            return str(int(value))
        return str(value)


def _fmt_val(val):
    """
    DEPRECATED: Dùng format_excel_number(value, number_format) thay thế.
    Giữ lại cho backward compatibility.
    """
    return format_excel_number(val, None)


def _normalize_nfc(value):
    """Normalize string to NFC form for consistent comparison."""
    return unicodedata.normalize('NFC', str(value)) if value is not None else ""


def _safe_color(color_obj):
    """Return a 6-char hex string or None from an openpyxl Color object."""
    try:
        if not color_obj:
            return None
        # RGB type
        if color_obj.type == 'rgb' and color_obj.rgb:
            rgb = str(color_obj.rgb).upper()
            # Handle 8-char ARGB (common in Excel)
            if len(rgb) == 8:
                # If it's 00000000 or FFFFFFFF, it's usually default/no-fill
                if rgb in ('00000000', 'FFFFFFFF'):
                    return None
                return rgb[2:]
            return rgb
        # Theme type
        if color_obj.type == 'theme' and color_obj.theme is not None:
            return f"THEME_{color_obj.theme}"
    except Exception:
        pass
    return None


def is_input_cell(cell):
    """
    Broadened detection: Any cell with a non-white background is likely an input marker.
    """
    try:
        fill = cell.fill
        if not fill or fill.patternType is None: return False
        
        c = fill.start_color
        if not c: return False
        
        # 1. Match by Theme (Aggressive)
        if c.type == 'theme' and c.theme is not None:
            if c.theme > 0: return True
            if c.theme == 0 and (c.tint and abs(c.tint) > 0.01): return True

        # 2. Match by RGB
        if hasattr(c, 'rgb') and c.rgb:
            rgb = str(c.rgb).upper()
            rgb_6 = rgb[-6:] if len(rgb) >= 6 else rgb
            if rgb_6 not in ['FFFFFF', '000000', '00000000']:
                return True
    except:
        pass
    return False


def _cell_css(cell):
    """Generate inline CSS for a cell based on its openpyxl style."""
    css = []
    
    # 1. Background color - Only apply if patternType is present and not default
    if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
        color = _safe_color(cell.fill.start_color)
        if color and not color.startswith("THEME_"):
            # Skip common default backgrounds that cause "black/white" issues
            if color not in ('FFFFFF', '000000'):
                css.append(f"background-color:#{color};")

    # 2. Font styles
    f = cell.font
    if f:
        if f.bold: css.append("font-weight:bold;")
        if f.italic: css.append("font-style:italic;")
        if f.color and f.color.rgb and isinstance(f.color.rgb, str):
            c = str(f.color.rgb)
            if len(c) == 8: css.append(f"color:#{c[2:]};")
        if f.sz: css.append(f"font-size:{f.sz}pt;")

    # 3. Alignment
    a = cell.alignment
    if a:
        if a.horizontal: css.append(f"text-align:{a.horizontal};")
        if a.vertical: css.append(f"vertical-align:{a.vertical};")
        if a.wrapText: css.append("white-space:normal;")
        else: css.append("white-space:nowrap;")

    return "".join(css)


def _build_merge_lookup(ws):
    """Return (spans, shadows) dicts to handle merged cells in HTML."""
    spans = {}  # (r, c) -> (rowspan, colspan)
    shadows = set() # (r, c) that are covered by a merge
    for mr in ws.merged_cells.ranges:
        s_row, s_col, e_row, e_col = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        spans[(s_row, s_col)] = (e_row - s_row + 1, e_col - s_col + 1)
        for r in range(s_row, e_row + 1):
            for c in range(s_col, e_col + 1):
                if r == s_row and c == s_col: continue
                shadows.add((r, c))
    return spans, shadows


def _col_widths_px(ws):
    """Estimate column widths in pixels."""
    widths = []
    for i in range(1, ws.max_column + 1):
        letter = get_column_letter(i)
        w = ws.column_dimensions[letter].width or 8.43
        widths.append(max(int(w * 7), 45))
    return widths


def _row_height_px(ws, r):
    """Estimate row height in pixels."""
    h = ws.row_dimensions[r].height or 15
    return int(h * 1.33)


def render_range_to_html(ws, start_row, end_row,
                         input_marker_hex='FFE0F2FE',
                         existing_values=None,
                         editable=True,
                         min_col=1, max_col=None):
    """
    Render rows [start_row .. end_row] (1-indexed, inclusive) of `ws`
    as an HTML <tbody> fragment.
    """
    if existing_values is None:
        existing_values = {}
    if max_col is None:
        max_col = ws.max_column

    spans, shadows = _build_merge_lookup(ws)
    html = []
    input_keys = []

    for r in range(start_row, end_row + 1):
        if ws.row_dimensions[r].hidden:
            continue
        rh = _row_height_px(ws, r)
        html.append(f'<tr style="height:{rh}px">')

        for c in range(min_col, max_col + 1):
            if (r, c) in shadows:
                continue

            cell = ws.cell(row=r, column=c)
            rowspan, colspan = spans.get((r, c), (1, 1))
            css = _cell_css(cell)

            is_input = is_input_cell(cell)
            coord = cell.coordinate
            rs_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ''
            cs_attr = f' colspan="{colspan}"' if colspan > 1 else ''
            base_style = 'padding:3px 6px;border:1px solid #d1d5db;overflow:hidden;box-sizing:border-box;'
            if is_input:
                base_style += 'background-color:#e0f2fe;'

            full_css = f'{base_style}{css}'
            td_open = f'<td{rs_attr}{cs_attr} style="{full_css}">'

            if is_input:
                val = existing_values.get(coord, '')
                safe_val = str(val).replace('"', '&quot;')
                input_keys.append(coord)
                td_inner = (
                    f'<input type="text" class="grid-input" '
                    f'data-key="{coord}" data-coord="{coord}" '
                    f'value="{safe_val}" onchange="markDirty()" '
                    f'style="width:100%;height:100%;border:none;'
                    f'background:transparent;padding:2px;font-size:inherit;">'
                )
            else:
                raw_val = cell.value
                if isinstance(raw_val, str) and raw_val.startswith('='):
                    raw_val = ''
                display = format_excel_number(raw_val, cell.number_format)
                td_inner = display

            html.append(f'{td_open}{td_inner}</td>')
        html.append('</tr>')

    return {
        'tbody_html': '\n'.join(html),
        'input_keys': input_keys,
    }
